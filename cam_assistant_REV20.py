import streamlit as st
import json
import math
import pandas as pd
from io import BytesIO
import plotly.graph_objects as go
import numpy as np  # Add this line for 3D simulation
import plotly.express as px
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import os
try:
    from scipy.spatial.distance import cdist
except ImportError:
    # Fallback if scipy is not installed
    def cdist(a, b, metric='euclidean'):
        return np.sqrt(np.sum((a[:, None] - b[None, :])**2, axis=2))
import time

# --- Cookie manager ---
from streamlit_cookies_manager import EncryptedCookieManager

# --- Add import for schematic ---
#from davenport_schematic import davenport_machine_graphic

# Load tool definitions and config (simplified for Davenport only)
try:
    with open("tool_definitions.json", "r") as f:
        TOOL_DEFINITIONS = json.load(f)
        # Convert feed_range lists back to tuples
        for tool_name, tool_data in TOOL_DEFINITIONS.items():
            if "feed_range" in tool_data:
                tool_data["feed_range"] = tuple(tool_data["feed_range"])
    
    with open("config.json", "r") as f:
        CONFIG = json.load(f)
        # Convert string keys to integers for block_setting_ranges
        if "block_setting_ranges" in CONFIG:
            CONFIG["block_setting_ranges"] = {int(k): tuple(v) for k, v in CONFIG["block_setting_ranges"].items()}
except FileNotFoundError:
    TOOL_DEFINITIONS = {}
    CONFIG = {}

# JSON Data Loading Functions
@st.cache_data
def load_drill_charts():
    """Load drill size charts from JSON"""
    try:
        with open("drill_sizes.json", "r") as f:
            return json.load(f)
    except FileNotFoundError:
        st.error("❌ drill_sizes.json not found. Please ensure the file exists.")
        return {}
    except json.JSONDecodeError:
        st.error("❌ Error reading drill_sizes.json. Please check the file format.")
        return {}
# In your Excel export function, add this section:
if hasattr(st.session_state, 'cutoff_data') and st.session_state.cutoff_data:
    # Add cutoff operations sheet
    cutoff_ws = wb.create_sheet("Cutoff Operations")
    cutoff_ws.append(["Position", "Wall Thickness", "Feed Rate", "Required Rise", "Recommended Cam"])
    
    for pos_key, cutoff_info in st.session_state.cutoff_data.items():
        position = pos_key.replace("position_", "")
        best_cam = cutoff_info.get('cam_recommendations', [{}])[0]
        cutoff_ws.append([
            position,
            f"{cutoff_info.get('wall_thickness', 0):.3f}",
            f"{cutoff_info.get('recommended_feed', 0):.4f}",
            f"{cutoff_info.get('required_rise', 0):.3f}",
            best_cam.get('cam_id', 'N/A')
        ])

@st.cache_data
def load_threading_charts():
    """Load threading charts from JSON"""
    try:
        with open("threading_charts.json", "r") as f:
            return json.load(f)
    except FileNotFoundError:
        st.error("❌ threading_charts.json not found. Please ensure the file exists.")
        return {}
    except json.JSONDecodeError:
        st.error("❌ Error reading threading_charts.json. Please check the file format.")
        return {}

def search_drill_sizes(search_term, drill_data):
    """Enhanced drill size search using JSON data"""
    results = []
    search_term = search_term.upper().strip()
    
    # Search all drill categories
    for category_name, drills in drill_data.items():
        if isinstance(drills, list):
            for drill in drills:
                # Search by size name
                if search_term in drill["size"].upper():
                    results.append(drill)
                # Search by decimal value (with tolerance)
                elif search_term.replace("#", "").replace("MM", "").replace(".", "").isdigit():
                    try:
                        search_val = float(search_term.replace("#", "").replace("MM", ""))
                        if abs(drill["decimal"] - search_val) < 0.001:
                            results.append(drill)
                    except ValueError:
                        continue
    
    return results

def search_thread_sizes(search_term, thread_data):
    """Search threading data"""
    results = []
    search_term = search_term.upper().strip()
    
    # Search UNC/UNF threads
    for thread in thread_data.get("unc_unf_threads", []):
        if search_term in thread["thread"].upper():
            results.append(thread)
    
    # Search metric threads
    for thread in thread_data.get("metric_threads", []):
        if search_term in thread["thread"].upper():
            results.append(thread)
    
    return results

# Hardcoded Davenport configuration (simplified)
DAVENPORT_CONFIG = {
    "max_rpm": 4500,
    "positions": 5,
    "cycle_rates": [75, 60, 45],
    "block_ranges": {
        1: (0.8, 1.2),
        2: (0.8, 1.2),
        3: (0.8, 1.2),
        4: (0.8, 1.2),
        5: (0.8, 1.2)
    }
}

# --- BEGIN: Add missing CONFIG and TOOL_DEFINITIONS ---
CONFIG = {
    "block_setting_ranges": DAVENPORT_CONFIG["block_ranges"],
    "threading_tools": ["TAP", "DIE HEAD", "THREAD ROLL"],
    "default_tools_end": ["DRILL", "REAMER", "TAP", "DIE HEAD", "CENTER"],
    "default_tools_side": ["BOXTOOL", "FORM TOOL", "SHAVE", "KNURL", "BROACH"]
}

TOOL_DEFINITIONS = {
    "DRILL": {"type": "TURNING"},
    "REAMER": {"type": "TURNING"},
    "TAP": {"type": "TURNING"},
    "DIE HEAD": {"type": "TURNING"},
    "CENTER": {"type": "TURNING"},
    "BOXTOOL": {"type": "TURNING"},
    "FORM TOOL": {"type": "FORM"},
    "SHAVE": {"type": "FORM"},
    "KNURL": {"type": "FORM"},
    "BROACH": {"type": "TURNING"},
    "THREAD ROLL": {"type": "FORM"}
}

# Tool library constants
DEFAULT_TOOLS_END = ["DRILL", "REAMER", "TAP", "DIE HEAD", "CENTER"]
DEFAULT_TOOLS_SIDE = ["BOXTOOL", "FORM TOOL", "SHAVE", "KNURL", "BROACH"]
THREADING_TOOLS = ["TAP", "DIE HEAD", "THREAD ROLL"]

def save_tool_library(tools, filename):
    """Save tool library to JSON file"""
    with open(filename, 'w') as f:
        json.dump(list(tools), f, indent=2)

def load_tool_library(filename, defaults=None):
    if os.path.exists(filename):
        with open(filename) as f:
            return set(json.load(f))
    return set(defaults) if defaults else set()

def tool_selector(label, tools, session_key, lib_key, filename):
    # Filter out any "Custom" entries that might have been added to prevent duplication
    filtered_tools = [tool for tool in tools if tool.lower() not in ["custom", "custom..."]]
    tool_options = [""] + list(filtered_tools) + ["Custom..."]
    selected = st.selectbox(label, tool_options, key=f"{session_key}_select")
    if selected == "Custom...":
        custom_tool = st.text_input("Enter custom tool name:", key=f"{session_key}_custom")
        if custom_tool:
            # Prevent adding "Custom" variants to avoid duplication with the Custom... option
            if custom_tool.lower() in ["custom", "custom..."]:
                st.error("❌ Cannot use 'Custom' as a tool name - it's reserved for the menu option.")
                return ""
            st.session_state[lib_key].add(custom_tool)
            save_tool_library(st.session_state[lib_key], filename)
            return custom_tool
        return ""
    elif selected:
        return selected
    return ""

def enhanced_tool_selector(label, tools, session_key, lib_key, filename, tool_definitions=None):
    # Filter out any "Custom" entries that might have been added to prevent duplication
    filtered_tools = [tool for tool in tools if tool.lower() not in ["custom", "custom..."]]
    tool_options = [""] + list(filtered_tools) + ["Custom..."]
    
    # Get previously selected value from session state for persistence
    previous_selection = st.session_state.get(session_key, "")
    
    # Ensure the previous selection is in the tool_options, otherwise default to empty
    if previous_selection and previous_selection not in tool_options:
        # If it was a custom tool, add it to the options (but not if it's a "Custom" variant)
        if previous_selection not in ["", "Custom..."] and previous_selection.lower() not in ["custom"]:
            tool_options.insert(-1, previous_selection)  # Insert before "Custom..."
    
    # Set the default index based on previous selection
    try:
        default_index = tool_options.index(previous_selection) if previous_selection in tool_options else 0
    except ValueError:
        default_index = 0
    
    selected = st.selectbox(label, tool_options, index=default_index, key=f"{session_key}_select")
    
    # Store the selected tool in the expected session state key for persistence
    if selected and selected != "Custom...":
        st.session_state[session_key] = selected
    elif selected == "":
        st.session_state[session_key] = ""
    
    # Check for threading tools and show warning
    if selected in THREADING_TOOLS:
        st.warning(f"⚡️ **{selected}** is a threading tool! Please use the **🧮 Thread Calculator** tab for cam specifications.")
    
    # Show tool information if available
    if selected and selected != "Custom..." and tool_definitions and selected in tool_definitions:
        tool_info = tool_definitions[selected]
        with st.expander(f"ℹ️ {selected} Info", expanded=False):
            st.caption(f"**Type:** {tool_info.get('type', 'Unknown')}")
            st.caption(f"**Description:** {tool_info.get('description', 'No description')}")
            feed_range = tool_info.get('feed_range', (0, 0))
            st.caption(f"**Feed Range:** {feed_range[0]:.4f} - {feed_range[1]:.4f} in/rev")
            st.caption(f"**SFM Modifier:** {tool_info.get('sfm_modifier', 1.0):.1f}x")
            st.caption(f"**Default Approach:** {tool_info.get('approach_default', 0.050):.3f} in")
    
    if selected == "Custom...":
        custom_tool = st.text_input("Enter custom tool name:", key=f"{session_key}_custom")
        if custom_tool:
            # Prevent adding "Custom" variants to avoid duplication with the Custom... option
            if custom_tool.lower() in ["custom", "custom..."]:
                st.error("❌ Cannot use 'Custom' as a tool name - it's reserved for the menu option.")
                return ""
            st.session_state[lib_key].add(custom_tool)
            save_tool_library(st.session_state[lib_key], filename)
            st.session_state[session_key] = custom_tool
            return custom_tool
        return ""
    elif selected:
        return selected
    return ""

def load_data():
    with open("materials.json") as f:
        material_data = json.load(f)
    with open("sfm_guidelines.json") as f:
        sfm_guidelines = json.load(f)
    # Load legacy cam data for backward compatibility
    try:
        with open("cams_data.json") as f:
            legacy_cam_data = json.load(f)
    except FileNotFoundError:
        legacy_cam_data = {}
    
    # Add hardcoded Davenport cam
    legacy_cam_data["5-C-792"] = {"size": "3/16", "rise": 0.1650, "type": "TURNING", "cut_start": 0, "cut_end": 45, "dwell_end": 50, "total_spaces": 100}
    
    return material_data, sfm_guidelines, legacy_cam_data

def suggest_davenport_cam(rise_needed, material_data=None, sfm_guidelines=None, legacy_cam_data=None):
    """Suggest appropriate Davenport cam based on rise needed"""
    
    # If data not provided, load it
    if not legacy_cam_data:
        try:
            material_data, sfm_guidelines, legacy_cam_data = load_data()
        except:
            legacy_cam_data = {}
    
    # Add some hardcoded Davenport cams if none exist
    if not legacy_cam_data:
        legacy_cam_data = {
            "5-C-792": {"size": "3/16", "rise": 0.1650, "type": "TURNING", "cut_start": 0, "cut_end": 45, "dwell_end": 50, "total_spaces": 100},
            "5-C-793": {"size": "1/4", "rise": 0.2000, "type": "TURNING", "cut_start": 0, "cut_end": 45, "dwell_end": 50, "total_spaces": 100},
            "5-C-794": {"size": "5/16", "rise": 0.2500, "type": "TURNING", "cut_start": 0, "cut_end": 45, "dwell_end": 50, "total_spaces": 100},
            "5-C-795": {"size": "3/8", "rise": 0.3000, "type": "TURNING", "cut_start": 0, "cut_end": 45, "dwell_end": 50, "total_spaces": 100}
        }
    
    # Find suitable cams
    suitable_cams = []
    
    for cam_name, cam_info in legacy_cam_data.items():
        cam_rise = cam_info.get("rise", 0)
        if cam_rise >= rise_needed:
            suitable_cams.append({
                "name": cam_name,
                "rise": cam_rise,
                "size": cam_info.get("size", "Unknown"),
                "type": cam_info.get("type", "TURNING")
            })
    
    # Sort by rise (smallest first that meets requirement)
    suitable_cams.sort(key=lambda x: x["rise"])
    
    if suitable_cams:
        return suitable_cams[0]  # Return the best match
    else:
        # Return a default/custom cam suggestion
        return {
            "name": "CUSTOM CAM REQUIRED",
            "rise": rise_needed * 1.1,  # Add 10% safety margin
            "size": "Custom",
            "type": "TURNING"
        }

@st.cache_data
def get_threading_cams():
    """Get threading cam data - replacing the old @st.cache_data function"""
    try:
        with open("threading_cams.json", "r") as f:
            return json.load(f)
    except FileNotFoundError:
        # Return hardcoded threading cams if file not found
        return {
            "TAP-001": {"rise": 0.125, "pitch": 0.05, "type": "TAP"},
            "TAP-002": {"rise": 0.250, "pitch": 0.05, "type": "TAP"},
            "DIE-001": {"rise": 0.125, "pitch": 0.05, "type": "DIE HEAD"},
            "DIE-002": {"rise": 0.250, "pitch": 0.05, "type": "DIE HEAD"}
        }

def calculate_threading_rise(thread_pitch, num_threads, safety_factor=1.2):
    """Calculate required rise for threading operation"""
    return thread_pitch * num_threads * safety_factor


# Load actual Davenport cam data from JSON
    
    # Load actual Davenport cam data from JSON
    try:
        with open("davenport_cams.json") as f:
            davenport_cams = json.load(f)
    except FileNotFoundError:
        return "⚠️ Davenport cam database not found"
    
    # Filter only threading cams
    threading_cams = {name: details for name, details in davenport_cams.items() 
                     if "THREADING" in details.get("type", "").upper()}
    
    if not threading_cams:
        return "⚠️ No threading cams found in database"
    
    # Find the best fitting cam (smallest cam that meets or exceeds requirement)
    best_cams = []
    
    for cam_name, cam_info in threading_cams.items():
        cam_rise = cam_info.get("rise", 0)
        if cam_rise >= rise_needed:
            best_cams.append((cam_name, cam_info, abs(cam_rise - rise_needed)))
    
    if not best_cams:
        return "⚠️ Rise exceeds available threading cam range - consider custom cam"
    
    # Sort by closest match (smallest difference)
    best_cams.sort(key=lambda x: x[2])
    
    # Find both brass and steel options for the best size
    primary_suggestion = best_cams[0]
    best_rise = primary_suggestion[1].get("rise", 0)
    cam_size = primary_suggestion[1].get("size", "")
    
    # Get all cams with this rise value (within 0.001" tolerance)
    matching_cams = [cam for cam in best_cams if abs(cam[1].get("rise", 0) - best_rise) < 0.001]
    
    # Separate by material type and extract cam size
    brass_cams = []
    steel_cams = []
    
    for cam_name, cam_info, diff in matching_cams:
        cam_type = cam_info.get("type", "").upper()
        if "BRASS" in cam_type:
            brass_cams.append((cam_name, cam_info))
        elif "STEEL" in cam_type:
            steel_cams.append((cam_name, cam_info))
    
    # Build suggestion string
    suggestions = []
    if brass_cams:
        cam_name, cam_info = brass_cams[0]
        suggestions.append(f"{cam_name} (Brass #{cam_info.get('size', '')})")
    if steel_cams:
        cam_name, cam_info = steel_cams[0]
        suggestions.append(f"{cam_name} (Steel #{cam_info.get('size', '')})")
    
    if len(suggestions) > 1:
        return f"**Suggested Cams:** {' or '.join(suggestions)} (Rise: {best_rise:.4f}\")"
    elif suggestions:
        return f"**Suggested Cam:** {suggestions[0]} (Rise: {best_rise:.4f}\")"
    else:
        return f"**Suggested Cam:** {primary_suggestion[0]} (Rise: {best_rise:.4f}\")"

def load_machine_cam_data():
    """Load Davenport cam database"""
    try:
        with open("davenport_cams.json") as f:
            cam_data = json.load(f)
    except FileNotFoundError:
        # Fallback to legacy cam data
        try:
            with open("cams_data.json") as f:
                cam_data = json.load(f)
        except FileNotFoundError:
            cam_data = {}
    
    # Add hardcoded Davenport cam
    cam_data["5-C-792"] = {"size": "3/16", "rise": 0.1650, "type": "TURNING", "cut_start": 0, "cut_end": 45, "dwell_end": 50, "total_spaces": 100}
    
    return cam_data

def load_gear_table(cpm="75"):
    """Load Davenport gear table"""
    try:
        with open("gears.json") as f:
            data = json.load(f)
    except FileNotFoundError:
        # Return empty if no gear file
        return []
    
    # Handle gear data structure
    if str(cpm) in data:
        if data[str(cpm)] == "same_as_75":
            return data["75"]
        else:
            return data[str(cpm)]
    else:
        # Fallback logic for missing cycle rates
        if str(cpm) in ["60", "45"]:
            return data.get("75", [])
        return data.get(str(cpm), [])

def load_cycle_time_data(cpm_setting):
    """Load cycle time data from JSON files"""
    cpm_files = {
        75: "75_cycle_time_eff_revs_COMPLETE.json",
        60: "60_cycle_time_eff_revs.json",
        45: "45_cycle_time_eff_revs.json"  # ✅ NOW AVAILABLE
    }
    
    if cpm_setting not in cpm_files:
        return None
    
    try:
        with open(cpm_files[cpm_setting]) as f:
            return json.load(f)
    except FileNotFoundError:
        st.warning(f"⚠️ {cpm_files[cpm_setting]} not found")
        return None

def find_closest_rpm_index(target_rpm, rpm_list):
    """Find index of closest RPM in the manual chart"""
    differences = [abs(rpm - target_rpm) for rpm in rpm_list]
    return differences.index(min(differences))

def find_manual_feed_gears(max_effective_revs, setup_rpm, cpm_setting):
    """
    Follow exact Davenport manual methodology:
    1. Calculate effective revolutions needed (already done)
    2. Go to the RPM column that matches your setup RPM
    3. Go DOWN that RPM column to find the closest effective revolutions
    4. Go ACROSS to find the corresponding cycle time and gear configuration
    """
    
    # Load cycle time data
    cycle_data = load_cycle_time_data(cpm_setting)
    if not cycle_data:
        return None
    
    # Find closest RPM in manual charts
    rpm_index = find_closest_rpm_index(setup_rpm, cycle_data["rpm_spindles"])
    closest_manual_rpm = cycle_data["rpm_spindles"][rpm_index]
    
    best_matches = []
    
    # Check each gear ratio - find closest effective revs match in the RPM column
    for gear_ratio in cycle_data["gear_ratios"]:
        # Check if the gear_ratio has effective_revolutions key
        if "effective_revolutions" not in gear_ratio:
            continue
            
        if rpm_index < len(gear_ratio["effective_revolutions"]):
            manual_effective_revs = gear_ratio["effective_revolutions"][rpm_index]
            revs_difference = abs(manual_effective_revs - max_effective_revs)
            
            best_matches.append({
                "manual_cycle_time": gear_ratio["time_seconds"],
                "manual_effective_revs": manual_effective_revs,
                "revs_difference": revs_difference,
                "revs_percentage_diff": (revs_difference / max_effective_revs * 100) if max_effective_revs > 0 else 100,
                "driver": gear_ratio["driver"],
                "driven": gear_ratio.get("driven"),
                "driven_compound": gear_ratio.get("driven_compound"),
                "driver_compound": gear_ratio.get("driver_compound"),
                "production_per_hour": gear_ratio["production_per_hour"],
                "rpm_used": closest_manual_rpm
            })
    
    # Check compound gear configurations if available
    if "compound_gear_configurations" in cycle_data:
        for gear_ratio in cycle_data["compound_gear_configurations"]:
            # Check if the gear_ratio has effective_revolutions key
            if "effective_revolutions" not in gear_ratio:
                continue
                
            if rpm_index < len(gear_ratio["effective_revolutions"]):
                manual_effective_revs = gear_ratio["effective_revolutions"][rpm_index]
                revs_difference = abs(manual_effective_revs - max_effective_revs)
                
                best_matches.append({
                    "manual_cycle_time": gear_ratio["time_seconds"],
                    "manual_effective_revs": manual_effective_revs,
                    "revs_difference": revs_difference,
                    "revs_percentage_diff": (revs_difference / max_effective_revs * 100) if max_effective_revs > 0 else 100,
                    "driver": gear_ratio["driver"],
                    "driven": gear_ratio.get("driven"),
                    "driven_compound": gear_ratio.get("driven_compound"),
                    "driver_compound": gear_ratio.get("driver_compound"),
                    "production_per_hour": gear_ratio["production_per_hour"],
                    "rpm_used": closest_manual_rpm,
                    "is_compound": True
                })
    
    if not best_matches:
        return None
    
    # Sort to prefer values at or above calculated effective revs (never under)
    # Primary: prefer values >= calculated (not under)
    # Secondary: closest to calculated value
    def sort_key(match):
        manual_revs = match["manual_effective_revs"]
        is_under = manual_revs < max_effective_revs
        # If under calculated value, add large penalty to sort it last
        penalty = 999999 if is_under else 0
        return penalty + match["revs_difference"]
    
    best_matches.sort(key=sort_key)
    return best_matches[0]

def calculate_shave_cam_recommendation(target_rise, total_travel, cam_db, machine_profile=None, position=1):
    """
    Calculate SHAVE tool cam recommendation using Excel formula:
    =SQRT((C5/2)^2-(C6/2)^2)+.02
    where C5 = largest diameter, C6 = smallest diameter
    
    This function will be called when user needs to input diameters for SHAVE operations
    """
    # This is a placeholder that returns a FORM cam
    # The actual diameter-based calculation will be done in the UI
    # where the user can input largest_dia and smallest_dia
    
    # For now, return the best FORM cam available
    form_cams = {name: details for name, details in cam_db.items() 
                 if details.get("type", "").strip().upper() == "FORM"}
    
    if not form_cams:
        return None
    
    # Find cam closest to target rise
    closest = None
    min_diff = float("inf")
    
    for cam_name, details in form_cams.items():
        cam_rise = details.get("rise", 0)
        if cam_rise <= 0:
            continue
            
        diff = abs(cam_rise - target_rise)
        if diff < min_diff:
            min_diff = diff
            closest = (cam_name, details)
    
    return closest

def calculate_shave_cam_rise(largest_dia, smallest_dia):
    """
    Calculate SHAVE cam rise using Excel formula:
    =SQRT((largest_dia/2)^2-(smallest_dia/2)^2)+.02
    """
    import math
    
    if largest_dia <= smallest_dia:
        return 0.02  # Minimum rise
    
    largest_radius = largest_dia / 2
    smallest_radius = smallest_dia / 2
    
    # Excel formula: =SQRT((C5/2)^2-(C6/2)^2)+.02
    calculated_rise = math.sqrt(largest_radius**2 - smallest_radius**2) + 0.02
    
    return calculated_rise

def recommend_cam(target_rise, total_travel, tool_type, cam_db, material, position=1, machine_profile=None):
    # Special handling for SHAVE tool - calculate cam using Excel formula
    if tool_type.strip().upper() == "SHAVE":
        return calculate_shave_cam_recommendation(target_rise, total_travel, cam_db, machine_profile, position)
    
    # Use machine profile if provided, otherwise fall back to legacy BLOCK_SETTING_RANGES
    if machine_profile and "block_ranges" in machine_profile:
        min_block, max_block = machine_profile["block_ranges"].get(position, (0.8, 1.2))
    else:
        min_block, max_block = BLOCK_SETTING_RANGES.get(position, (0.8, 1.2))
    
    tool_to_type = {
        "BOXTOOL": "TURNING", "BROACH": "TURNING", "CENTER": "TURNING", "COUNTERBORE": "TURNING",
        "CUSTOM": "TURNING", "DEIHEAD": "TURNING", "DIE HEAD": "TURNING", "DRILL": "TURNING", "HOLLOW MILL": "TURNING",
        "REAMER": "TURNING", "TAP": "TURNING", "TRIPAN": "TURNING",
        "CROSS DRILL": "FORM", "CROSS TAP": "FORM", "CUTOFF": "FORM", "FORM TOOL": "FORM",
        "KNURL": "FORM", "ROLL STAMP": "FORM", "SHAVE": "FORM", "SKIVE": "FORM", "THREAD ROLL": "FORM",
        "BRASS THREADING": "BRASS THREADING", "STEEL THREADING": "STEEL THREADING"
    }
    material_to_thread = {"360 Brass": "BRASS THREADING", "C260 Brass": "BRASS THREADING", "C464 Naval Brass": "BRASS THREADING"}
    cam_type_target = tool_to_type.get(tool_type.strip().upper())
    if tool_type in ["TAP", "DIE HEAD", "THREAD ROLL"] and material in material_to_thread:
        cam_type_target = material_to_thread[material]
    if not cam_type_target:
        return None
    closest = None
    min_diff = float("inf")
    for cam_name, details in cam_db.items():
        cam_type = details.get("type", "").strip().upper()
        cam_rise = details.get("rise", 0)
        if cam_type != cam_type_target or cam_rise <= 0:
            continue
        block_setting = total_travel / cam_rise
        if not (min_block <= block_setting <= max_block):
            continue
        diff = abs(cam_rise - target_rise)
        if diff < min_diff:
            min_diff = diff
            closest = (cam_name, details)
    return closest

def get_manual_threading_gears(threading_method, cpm_setting=75):
    """
    Return official Davenport manual threading gear configurations
    Based on pages 135-151 of the Davenport instruction manual
    """
    
    # Official Davenport manual threading gear configurations
    threading_gears = {
        "6:1": {
            "description": "6:1 Threading (Steel) - Most Common",
            "gear_ratios": [
                {"driver": 32, "driven": 32, "description": "Main spindle ratio 1:1"},
                {"driver": 21, "driven": 28, "description": "Threading spindle ratio 0.75:1"}
            ],
            "combined_ratio": 0.75,
            "rpm_formula": "Work RPM × 0.75",
            "cam_spaces": "0 to 32.5 hundredths (65% of cycle)",
            "typical_applications": ["Steel threading", "Standard die heads", "Most threading operations"]
        },
        "2:1": {
            "description": "2:1 Threading (Brass) - High Speed",
            "gear_ratios": [
                {"driver": 36, "driven": 27, "description": "Threading spindle ratio 1.333:1"}
            ],
            "combined_ratio": 1.333,
            "rpm_formula": "Work RPM × 1.333",
            "cam_spaces": "0 to 25 hundredths (50% of cycle)",
            "typical_applications": ["Brass threading", "High-speed operations", "Short threads"]
        },
        "4:1": {
            "description": "4:1 Threading (Hybrid) - Half Speed",
            "gear_ratios": [
                {"driver": "Variable", "driven": "Variable", "description": "Half-speed method"}
            ],
            "combined_ratio": 0.5,
            "rpm_formula": "Work RPM × 0.5",
            "cam_spaces": "0 to 25 hundredths (50% of cycle)",
            "typical_applications": ["Deep threading", "Long threads", "Special applications"]
        }
    }
    
    # Return the specific threading method requested
    return threading_gears.get(threading_method, threading_gears["6:1"])

def get_all_threading_gears():
    """
    Return all threading gear configurations as a dictionary
    Helper function for code that needs access to all methods
    """
    
    return {
        "6:1": get_manual_threading_gears("6:1"),
        "2:1": get_manual_threading_gears("2:1"),
        "4:1": get_manual_threading_gears("4:1")
    }

def get_threading_method_recommendation(material="Steel", thread_length=0.375, tpi=24.0):
    """
    Recommend threading method based on material and threading requirements
    Following Davenport manual guidelines
    """
    
    recommendations = []
    
    # Material-based recommendations
    if "brass" in material.lower() or "360" in material or "c260" in material.lower():
        recommendations.append({
            "method": "2:1",
            "priority": 1,
            "reason": "Brass material - high speed threading method recommended"
        })
        recommendations.append({
            "method": "6:1", 
            "priority": 2,
            "reason": "Alternative for brass if 2:1 too aggressive"
        })
    else:
        recommendations.append({
            "method": "6:1",
            "priority": 1, 
            "reason": "Steel material - standard 6:1 method recommended"
        })
    
    # Thread length considerations
    if thread_length > 0.5:
        recommendations.append({
            "method": "4:1",
            "priority": 2,
            "reason": f"Long thread length ({thread_length:.3f}\") - consider half-speed method"
        })
    
    # Fine pitch considerations  
    if tpi > 32:
        recommendations.append({
            "method": "4:1",
            "priority": 2,
            "reason": f"Fine pitch ({tpi:.1f} TPI) - half-speed for precision"
        })
    
    return sorted(recommendations, key=lambda x: x["priority"])

def calculate_cpm_from_cycle_time(cycle_time_seconds):
    """Calculate Cycles Per Minute from cycle time in seconds"""
    if cycle_time_seconds <= 0:
        return 0
    return 60.0 / cycle_time_seconds

def calculate_cycle_time_from_cpm(cpm):
    """Calculate cycle time in seconds from Cycles Per Minute"""
    if cpm <= 0:
        return 0
    return 60.0 / cpm

def validate_cycle_time_cpm(cycle_time_seconds, machine_profile=None):
    """Validate cycle time against standard machine CPM rates"""
    cpm = calculate_cpm_from_cycle_time(cycle_time_seconds)
    
    # Default standard CPM rates
    standard_cpms = [75, 60, 45]
    
    # Use machine-specific rates if available
    if machine_profile and "cycle_rates" in machine_profile:
        standard_cpms = machine_profile["cycle_rates"]
    
    # Find closest standard CPM
    closest_cpm = min(standard_cpms, key=lambda x: abs(x - cpm))
    difference = abs(closest_cpm - cpm)
    
    return {
        "calculated_cpm": cpm,
        "closest_standard_cpm": closest_cpm,
        "difference": difference,
        "is_standard": difference < 2.0,  # Within 2 CPM tolerance
        "standard_cpms": standard_cpms
    }

def suggest_collet(diameter, shape, units="in"):
    """Suggest appropriate collet based on diameter, shape, and units"""
    
    if units == "mm":
        # Convert mm to inches for calculations
        dia_inches = diameter / 25.4
        
        # Common metric collet sizes
        metric_collets = [
            (3.0, "3mm"), (4.0, "4mm"), (5.0, "5mm"), (6.0, "6mm"), (8.0, "8mm"), 
            (10.0, "10mm"), (12.0, "12mm"), (16.0, "16mm"), (20.0, "20mm"), (25.0, "25mm")
        ]
        
        # Find closest metric collet
        closest_metric = min(metric_collets, key=lambda x: abs(x[0] - diameter))
        suggestion = f"{closest_metric[1]} RD" if shape == "Round" else f"{closest_metric[1]} {shape.upper()}"
        
    else:
        # Imperial fractional collets
        dia_inches = diameter
        
        # Common fractional sizes
        fractions = [
            (0.0625, "1/16"), (0.09375, "3/32"), (0.125, "1/8"), (0.15625, "5/32"),
            (0.1875, "3/16"), (0.21875, "7/32"), (0.25, "1/4"), (0.28125, "9/32"),
            (0.3125, "5/16"), (0.34375, "11/32"), (0.375, "3/8"), (0.40625, "13/32"),
            (0.4375, "7/16"), (0.46875, "15/32"), (0.5, "1/2"), (0.53125, "17/32"),
            (0.5625, "9/16"), (0.59375, "19/32"), (0.625, "5/8"), (0.6875, "11/16"),
            (0.75, "3/4"), (0.8125, "13/16"), (0.875, "7/8"), (0.9375, "15/16"),
            (1.0, "1"), (1.125, "1-1/8"), (1.25, "1-1/4"), (1.375, "1-3/8"),
            (1.5, "1-1/2"), (1.625, "1-5/8"), (1.75, "1-3/4"), (2.0, "2")
        ]
        
        # Find closest fractional size
        closest_fraction = min(fractions, key=lambda x: abs(x[0] - dia_inches))
        suggestion = f"{closest_fraction[1]} RD" if shape == "Round" else f"{closest_fraction[1]} {shape.upper()}"
    
    return suggestion

def suggest_burr_collet(diameter, shape, units="in"):
    """Suggest burr collet - typically same size as main collet for most applications"""
    
    # Burr collet is usually the same size as the main collet
    # Use the same suggestion logic as the main collet
    return suggest_collet(diameter, shape, units)

def job_setup_section(material_data, sfm_guidelines):
    st.header("Job Setup – Multi-Spindle CAM Machine")
    
    # Get stored setup data if available
    setup_data_stored = st.session_state.get("setup_data", {})
    
    # Simplified to Davenport only for now
    machine_type = "Davenport Model B"
    
    # Use the global Davenport configuration
    machine_config = DAVENPORT_CONFIG
    
    st.info(f"**{machine_type}** - Max RPM: {machine_config['max_rpm']}, Positions: {machine_config['positions']}")
    
    # Davenport Machine Capacity Selection
    machine_capacity = None
    if "Davenport" in machine_type:
        st.markdown("#### 🔧 Machine Capacity")
        capacity_options = {
            "Standard Machine": {
                "round": "7/8\"",
                "hex": "3/4\"", 
                "square": "5/16\"",
                "description": "Standard Davenport capacity"
            },
            "Oversized Machine": {
                "round": "13/16\"",
                "hex": "11/16\"",
                "square": "9/16\"", 
                "description": "Oversized Davenport capacity"
            }
        }
        
        # Use narrower column layout for the selectbox
        capacity_col1, capacity_col2 = st.columns([1, 2])
        with capacity_col1:
            # Handle machine capacity selection with stored value
            capacity_keys = list(capacity_options.keys())
            stored_machine_capacity = setup_data_stored.get("machine_capacity", capacity_keys[0])
            capacity_index = capacity_keys.index(stored_machine_capacity) if stored_machine_capacity in capacity_keys else 0
            machine_capacity = st.selectbox(
                "Select Machine Capacity",
                capacity_keys,
                index=capacity_index,
                key="davenport_capacity",
                help="Choose machine capacity based on your Davenport configuration"
            )
            
            # Display capacity info aligned under the dropdown in the same column
            selected_capacity = capacity_options[machine_capacity]
            
            # Use smaller columns within the left column for alignment
            sub_col1, sub_col2, sub_col3 = st.columns([1, 1, 1])
            with sub_col1:
                st.metric("🔵 Round", selected_capacity["round"])
            with sub_col2:
                st.metric("⬡ Hex", selected_capacity["hex"])
            with sub_col3:
                st.metric("⬜ Square", selected_capacity["square"])
        
        st.caption(f"💡 {selected_capacity['description']}")
    
    # Units selection
    units_col1, units_col2 = st.columns([1, 3])
    with units_col1:
        # Handle units selection with stored value
        units_options = ["in", "mm"]
        stored_units = setup_data_stored.get("units", "in")
        units_index = units_options.index(stored_units) if stored_units in units_options else 0
        units = st.selectbox("Units", units_options, index=units_index, key="setup_units", help="Select measurement units")
    with units_col2:
        if units == "mm":
            st.caption("🔄 All measurements will be in millimeters. Machine calculations remain in inches internally.")
        else:
            st.caption("📏 All measurements in inches (standard for most CAM machines)")
    
    col1, col2 = st.columns(2)
    with col1:
        # Get values from session state if available, otherwise use defaults
        setup_data_stored = st.session_state.get("setup_data", {})
        
        job_name = st.text_input("Part Number", value=setup_data_stored.get("job_name", "44153-36-99"), key="setup_job_name")
        
        # Handle material selection with proper index
        materials_list = list(material_data.keys())
        stored_material = setup_data_stored.get("material", materials_list[0])
        material_index = materials_list.index(stored_material) if stored_material in materials_list else 0
        material = st.selectbox("Material", materials_list, index=material_index, key="setup_material")
        
        # Handle bar shape selection with proper index
        bar_shapes = ["Round", "Hex", "Special", "Tube", "Square"]
        stored_shape = setup_data_stored.get("bar_shape", "Round")
        shape_index = bar_shapes.index(stored_shape) if stored_shape in bar_shapes else 0
        bar_shape = st.selectbox("Bar Shape", bar_shapes, index=shape_index, key="setup_shape")
        
        # Enhanced diameter input with 4 decimal places and unit conversion
        stored_dia = setup_data_stored.get("dia", 0.3125)
        # Ensure stored diameter meets minimum requirement
        stored_dia = max(stored_dia, 0.1000) if stored_dia > 0 else 0.3125
        if units == "mm":
            stored_dia_mm = stored_dia * 25.4 if stored_dia else 7.9375
            stored_dia_mm = max(stored_dia_mm, 2.54)
            dia_mm = st.number_input("Bar Diameter (mm)", min_value=2.54, max_value=50.8, value=stored_dia_mm, step=0.0001, format="%.4f", key="setup_dia_mm")
            dia = dia_mm / 25.4  # Convert to inches for internal calculations
            st.caption(f"💡 Diameter in inches: {dia:.4f}\"")
        else:
            dia = st.number_input("Bar Diameter (in)", min_value=0.1000, max_value=2.0000, value=stored_dia, step=0.0001, format="%.4f", key="setup_dia")
            st.caption(f"💡 Diameter in mm: {dia * 25.4:.4f} mm")
        
        # Enhanced part length with unit conversion
        stored_part_length = setup_data_stored.get("part_length", 0.0)
        # Ensure stored value meets minimum requirement
        stored_part_length = max(stored_part_length, 0.001) if stored_part_length > 0 else 0.001
        if units == "mm":
            stored_part_length_mm = stored_part_length * 25.4 if stored_part_length else 0.025
            stored_part_length_mm = max(stored_part_length_mm, 0.025)
            part_length_mm = st.number_input("Part Length (mm)", min_value=0.025, value=stored_part_length_mm, step=0.001, format="%.3f", key="setup_part_len_mm")
            part_length = part_length_mm / 25.4
        else:
            part_length = st.number_input("Part Length (in)", min_value=0.001, value=stored_part_length, step=0.001, format="%.3f", key="setup_part_len")
        
        # Enhanced cutoff and faceoff with unit conversion
        stored_cutoff = setup_data_stored.get("cutoff", 0.069)
        stored_faceoff = setup_data_stored.get("faceoff", 0.0)
        # Ensure stored values meet minimum requirements
        stored_cutoff = max(stored_cutoff, 0.010) if stored_cutoff > 0 else 0.069
        stored_faceoff = max(stored_faceoff, 0.000)  # faceoff can be 0.000
        if units == "mm":
            stored_cutoff_mm = stored_cutoff * 25.4 if stored_cutoff else 1.753
            stored_faceoff_mm = stored_faceoff * 25.4 if stored_faceoff else 0.0
            stored_cutoff_mm = max(stored_cutoff_mm, 0.254)
            stored_faceoff_mm = max(stored_faceoff_mm, 0.000)
            cutoff_mm = st.number_input("Cutoff Width (mm)", min_value=0.254, step=0.001, value=stored_cutoff_mm, format="%.3f", key="setup_cutoff_mm")
            cutoff = cutoff_mm / 25.4
            faceoff_mm = st.number_input("Faceoff Amount (mm)", min_value=0.000, step=0.001, value=stored_faceoff_mm, format="%.3f", key="setup_faceoff_mm")
            faceoff = faceoff_mm / 25.4
        else:
            cutoff = st.number_input("Cutoff Width (in)", min_value=0.010, step=0.001, value=stored_cutoff, format="%.3f", key="setup_cutoff")
            faceoff = st.number_input("Faceoff Amount (in)", min_value=0.000, step=0.001, value=stored_faceoff, format="%.3f", key="setup_faceoff")
        
        # Smart collet suggestion with proper state handling
        suggested_collet = suggest_collet(dia * 25.4 if units == "mm" else dia, bar_shape, units)
        
        # Initialize session state for collet suggestions
        if "collet_suggestion" not in st.session_state:
            st.session_state.collet_suggestion = suggested_collet
        if "feed_finger_suggestion" not in st.session_state:
            st.session_state.feed_finger_suggestion = suggested_collet
        if "burr_collect_suggestion" not in st.session_state:
            st.session_state.burr_collect_suggestion = suggest_burr_collet(dia * 25.4 if units == "mm" else dia, bar_shape, units)
        
        collet_col1, collet_col2 = st.columns([2, 1])
        with collet_col1:
            # Use the suggestion if the suggest button was pressed, otherwise use stored value
            if st.session_state.get("collet_suggest_pressed", False):
                collet_value = suggested_collet
                st.session_state["collet_suggest_pressed"] = False  # Reset the flag
            else:
                collet_value = setup_data_stored.get("collets", st.session_state.collet_suggestion)
            collets = st.text_input("Collets", value=collet_value, key="setup_collets")
        with collet_col2:
            if st.button("💡 Suggest", key="suggest_collet", help="Auto-suggest collet based on diameter and shape"):
                st.session_state["collet_suggest_pressed"] = True
                st.rerun()
        
        # Enhanced feed finger with suggestion
        feed_finger_col1, feed_finger_col2 = st.columns([2, 1])
        with feed_finger_col1:
            # Use the suggestion if the suggest button was pressed, otherwise use stored value
            if st.session_state.get("feed_finger_suggest_pressed", False):
                feed_finger_value = suggested_collet
                st.session_state["feed_finger_suggest_pressed"] = False  # Reset the flag
            else:
                feed_finger_value = setup_data_stored.get("feed_finger", st.session_state.feed_finger_suggestion)
            feed_finger = st.text_input("Feed Finger", value=feed_finger_value, key="setup_feed_finger")
        with feed_finger_col2:
            if st.button("💡 Suggest", key="suggest_feed_finger", help="Auto-suggest feed finger based on diameter and shape"):
                st.session_state["feed_finger_suggest_pressed"] = True
                st.rerun()
        
        stored_set_pads = setup_data_stored.get("set_pads", "")
        set_pads = st.text_input("Set Pads", value=stored_set_pads, key="setup_set_pads")
        
        # Enhanced burr collect with suggestion
        burr_col1, burr_col2 = st.columns([2, 1])
        with burr_col1:
            # Use the suggestion if the suggest button was pressed, otherwise use stored value
            if st.session_state.get("burr_collect_suggest_pressed", False):
                burr_suggestion = suggest_burr_collet(dia * 25.4 if units == "mm" else dia, bar_shape, units)
                burr_collect_value = burr_suggestion
                st.session_state["burr_collect_suggest_pressed"] = False  # Reset the flag
            else:
                burr_collect_value = setup_data_stored.get("burr_collect", st.session_state.burr_collect_suggestion)
            burr_collect = st.text_input("Burr Collect", value=burr_collect_value, key="setup_burr_collect")
        with burr_col2:
            if st.button("💡 Suggest", key="suggest_burr_collect", help="Auto-suggest burr collet (typically 1-2 sizes larger)"):
                st.session_state["burr_collect_suggest_pressed"] = True
                st.rerun()
    with col2:
        # Get material-specific SFM - always use material default unless user has explicitly set different value
        material_sfm = material_data.get(material, {}).get("sfm", 300)  # Use 300 as reasonable fallback
        
        # Use material's SFM as default (this will update when material changes)
        sfm = st.number_input("Surface Feet per Minute (SFM)", value=material_sfm, step=5, key="setup_sfm")
        st.caption(f"💡 Material Default: {material_sfm} SFM for {material}")
        if dia > 0:
            rpm_calc = (sfm * 12) / (math.pi * dia)  # Correct SFM to RPM formula
            st.markdown(f"**Calculated Machine RPM:** {rpm_calc:.0f}")
            max_rpm = machine_config['max_rpm']
            if rpm_calc > max_rpm:
                st.warning(f"Calculated RPM ({rpm_calc:.0f}) exceeds {machine_type} limit ({max_rpm}). Consider reducing SFM or increasing diameter.")
                rpm_value = max_rpm  # Cap at machine limit
            else:
                rpm_value = int(rpm_calc)
        else:
            rpm_calc = 0
            rpm_value = machine_config['max_rpm']
            st.info("Enter a valid Bar Diameter for RPM calculation.")
        
        # Use calculated RPM as default, but allow user override
        # If user has manually set a different RPM, respect that unless material/diameter changed
        stored_rpm = setup_data_stored.get("rpm", rpm_value)
        
        # Check if calculation inputs changed - if so, use calculated value
        stored_sfm = setup_data_stored.get("sfm", material_sfm)
        stored_dia = setup_data_stored.get("dia", 0.3125)
        stored_material = setup_data_stored.get("material", list(material_data.keys())[0])
        
        # If SFM, diameter, or material changed, use calculated RPM
        if (sfm != stored_sfm or dia != stored_dia or material != stored_material):
            default_rpm = rpm_value
        else:
            # Use stored RPM if inputs haven't changed
            default_rpm = max(stored_rpm, 100) if stored_rpm > 0 else rpm_value
            
        rpm = st.number_input("Machine RPM", min_value=100, max_value=6000, value=default_rpm, step=100, key="setup_rpm")
        
        # Show RPM calculation status
        if dia > 0 and rpm != int(rpm_calc):
            if rpm_calc > machine_config['max_rpm']:
                st.caption(f"💡 Calculated: {rpm_calc:.0f} RPM (exceeds {machine_config['max_rpm']} limit)")
            else:
                st.caption(f"💡 Calculated: {rpm_calc:.0f} RPM, Using: {rpm} RPM (manual override)")
        elif dia > 0:
            st.caption(f"✅ Using calculated RPM: {rpm} RPM")
        
        # Enhanced bar length with unit conversion
        stored_bar_len = setup_data_stored.get("bar_len", 144.0)
        stored_remnant = setup_data_stored.get("remnant", 6.0)
        # Ensure stored values are reasonable (bar_len should be > 0, remnant >= 0)
        stored_bar_len = max(stored_bar_len, 1.0) if stored_bar_len > 0 else 144.0
        stored_remnant = max(stored_remnant, 0.0)
        if units == "mm":
            stored_bar_len_mm = stored_bar_len * 25.4 if stored_bar_len else 3657.6
            stored_remnant_mm = stored_remnant * 25.4 if stored_remnant else 152.4
            stored_bar_len_mm = max(stored_bar_len_mm, 25.4)
            stored_remnant_mm = max(stored_remnant_mm, 0.0)
            bar_len_mm = st.number_input("Bar Length (mm)", value=stored_bar_len_mm, step=25.4, format="%.1f", key="setup_bar_len_mm")
            bar_len = bar_len_mm / 25.4
            st.caption(f"💡 Bar length in inches: {bar_len:.3f}\"")
            remnant_mm = st.number_input("Remnant Length (mm)", min_value=0.0, max_value=bar_len_mm, value=min(stored_remnant_mm, bar_len_mm), step=12.7, format="%.1f", key="setup_remnant_mm")
            remnant = remnant_mm / 25.4
        else:
            bar_len = st.number_input("Bar Length (in)", value=stored_bar_len, step=1.0, format="%.3f", key="setup_bar_len")
            remnant = st.number_input("Remnant Length (in)", min_value=0.0, max_value=bar_len, value=min(stored_remnant, bar_len), step=0.500, format="%.3f", key="setup_remnant")
        
        stored_spindle_gears = setup_data_stored.get("spindle_gears", "44-20")
        stored_feed_gears = setup_data_stored.get("feed_gears", "50-30-60")
        spindle_gears = st.text_input("Spindle Gears", value=stored_spindle_gears, key="setup_spindle_gears")
        feed_gears = st.text_input("Feed Gears", value=stored_feed_gears, key="setup_feed_gears")
        thread_gears_val = st.session_state.get("recommended_threading_gears", setup_data_stored.get("thread_gears", ""))
        thread_gears = st.text_input("Threading Gears", value=thread_gears_val, key="setup_thread_gears")
        
        # Enhanced cycle time with CPM validation
        cycle_col1, cycle_col2 = st.columns([1, 1])
        with cycle_col1:
            stored_cycle_time = setup_data_stored.get("cycle_time", 1.6)
            # Ensure stored cycle time meets minimum requirement
            stored_cycle_time = max(stored_cycle_time, 0.1) if stored_cycle_time > 0 else 1.6
            cycle_time = st.number_input("Cycle Time (sec)", min_value=0.1, value=stored_cycle_time, step=0.1, format="%.2f", key="setup_cycle_time")
        with cycle_col2:
            if cycle_time > 0:
                validation = validate_cycle_time_cpm(cycle_time, machine_config)
                cpm = validation["calculated_cpm"]
                if validation["is_standard"]:
                    st.success(f"✅ {cpm:.1f} CPM")
                else:
                    st.warning(f"⚠️ {cpm:.1f} CPM (Non-standard)")
                st.caption(f"Standard: {', '.join(map(str, validation['standard_cpms']))}")
            
        stored_machine_code = setup_data_stored.get("machine_code", "A,B,O")
        machine_code = st.text_input("Machine Code", value=stored_machine_code, key="setup_machine_code")
    
    # Enhanced summary with units display
    usable_bar_len = bar_len - remnant
    per_part_len = part_length + cutoff + faceoff
    parts_per_bar = usable_bar_len / per_part_len if per_part_len > 0 else 0
    bar_weight = usable_bar_len * material_data.get(material, {}).get("density", 0.307)
    
    if units == "mm":
        st.caption(f"Usable Bar Length: **{usable_bar_len:.3f} in** ({usable_bar_len * 25.4:.1f} mm)")
        st.caption(f"Per Part Length: **{per_part_len:.3f} in** ({per_part_len * 25.4:.1f} mm)")
    else:
        st.caption(f"Usable Bar Length: **{usable_bar_len:.3f} in**")
        st.caption(f"Per Part Length: **{per_part_len:.3f} in**")
    
    st.caption(f"Estimated Parts per Bar: **{parts_per_bar:.2f}**")
    st.caption(f"Usable Bar Weight: **{bar_weight:.2f} lbs**")
    
    # Store all key data in session state for cross-tab persistence
    st.session_state["cycle_time_from_tab1"] = cycle_time
    st.session_state["setup_data"] = {
        "job_name": job_name, "material": material, "bar_shape": bar_shape, "dia": dia, 
        "part_length": part_length, "cutoff": cutoff, "faceoff": faceoff, "collets": collets, 
        "feed_finger": feed_finger, "set_pads": set_pads, "burr_collect": burr_collect, 
        "sfm": sfm, "rpm": rpm, "bar_len": bar_len, "remnant": remnant,
        "spindle_gears": spindle_gears, "feed_gears": feed_gears, "thread_gears": thread_gears, 
        "cycle_time": cycle_time, "machine_code": machine_code, "parts_per_bar": parts_per_bar, 
        "bar_weight": bar_weight, "machine_type": machine_type, "machine_config": machine_config, 
        "machine_capacity": machine_capacity, "units": units
    }
    
    return st.session_state["setup_data"]

def quote_breakdown_section(parts_per_bar, bar_weight, cycle_time=1.6):
    st.subheader("Quoting Quantities & Production Hours")
    
    # Get stored quote data if available
    quote_data_stored = st.session_state.get("quote_data", {})
    
    # Quantity inputs with stored values
    q1, q2, q3 = st.columns(3)
    with q1: 
        stored_low = quote_data_stored.get("low_qty", 100)
        stored_low = max(stored_low, 1) if stored_low > 0 else 100
        low = st.number_input("Low Quote Qty", value=stored_low, step=100, key="quote_low_qty")
    with q2: 
        stored_mid = quote_data_stored.get("mid_qty", 500)
        stored_mid = max(stored_mid, 1) if stored_mid > 0 else 500
        mid = st.number_input("Mid Quote Qty", value=stored_mid, step=100, key="quote_mid_qty")
    with q3: 
        stored_high = quote_data_stored.get("high_qty", 1000)
        stored_high = max(stored_high, 1) if stored_high > 0 else 1000
        high = st.number_input("High Quote Qty", value=stored_high, step=100, key="quote_high_qty")
    
    # Price inputs with stored values
    p1, p2, p3 = st.columns(3)
    with p1: 
        stored_price_low = quote_data_stored.get("price_low", 5.00)
        price_low = st.number_input("Price/lb – Low Qty", value=stored_price_low, step=0.10, format="%.2f", key="price_low")
    with p2: 
        stored_price_mid = quote_data_stored.get("price_mid", 4.75)
        price_mid = st.number_input("Price/lb – Mid Qty", value=stored_price_mid, step=0.10, format="%.2f", key="price_mid")
    with p3: 
        stored_price_high = quote_data_stored.get("price_high", 4.50)
        price_high = st.number_input("Price/lb – High Qty", value=stored_price_high, step=0.10, format="%.2f", key="price_high")
    
    # Enhanced cycle time input with CPM validation
    st.markdown("#### Production Time Calculations")
    cycle_col1, cycle_col2, cycle_col3 = st.columns([1, 1, 1])
    with cycle_col1:
        # Get cycle time from CAM Operations if available, otherwise use stored value or passed value
        tab2_cycle_time = st.session_state.get("cycle_time_from_tab2", cycle_time)
        stored_cycle_time_input = quote_data_stored.get("cycle_time_input", tab2_cycle_time)
        # Ensure stored cycle time meets minimum requirement
        stored_cycle_time_input = max(stored_cycle_time_input, 0.1) if stored_cycle_time_input > 0 else tab2_cycle_time
        cycle_time_input = st.number_input("Cycle Time (seconds)", value=stored_cycle_time_input, step=0.1, format="%.2f", key="production_cycle_time")
    
    with cycle_col2:
        # Calculate and display CPM from cycle time
        if cycle_time_input > 0:
            calculated_cpm = 60.0 / cycle_time_input
            st.metric("Calculated CPM", f"{calculated_cpm:.1f}")
            
            # Check against standard CPM rates [75, 60, 45]
            standard_cpms = [75, 60, 45]
            closest_cpm = min(standard_cpms, key=lambda x: abs(x - calculated_cpm))
            cpm_diff = abs(closest_cpm - calculated_cpm)
            
            if cpm_diff < 2:  # Within 2 CPM tolerance
                st.success(f"✅ Close to {closest_cpm} CPM")
            else:
                st.warning(f"⚠️ Not standard CPM. Closest: {closest_cpm}")
        else:
            st.error("Invalid cycle time")
    
    with cycle_col3:
        if tab2_cycle_time != cycle_time:
            st.success(f"✅ From CAM Operations: {tab2_cycle_time:.2f} sec")
        else:
            st.caption(f"💡 Default: {cycle_time_input:.2f} sec")
        st.caption("📊 Standard CPMs: 75, 60, 45")
        st.caption("🔧 Formula: CPM = 60 ÷ cycle_time")
    
    def calc_with_production(qty, price_per_lb, cycle_time_sec):
        bars = int((qty / parts_per_bar) + 0.999) if parts_per_bar else 0
        weight = round(qty * (bar_weight / parts_per_bar), 2) if parts_per_bar else 0
        cost = round(weight * price_per_lb, 2) if price_per_lb else 0.0
        
        # Enhanced production time calculations
        # Basic production time (cycle time × quantity)
        basic_production_seconds = qty * cycle_time_sec
        
        # Add setup time per bar change (typically 30-60 seconds per bar)
        setup_time_per_bar = 45  # seconds
        total_setup_time = bars * setup_time_per_bar
        
        # Add daily startup/warmup time (typically 15-30 minutes per 8-hour shift)
        if basic_production_seconds > 0:
            estimated_shifts = (basic_production_seconds + total_setup_time) / (8 * 3600)  # 8-hour shifts
            startup_time = max(1, estimated_shifts) * 20 * 60  # 20 minutes per shift
        else:
            startup_time = 0
        
        # Total time including setup and startup
        total_seconds_with_overhead = basic_production_seconds + total_setup_time + startup_time
        
        # Calculate hours at different efficiency rates
        total_hours_100 = total_seconds_with_overhead / 3600  # 100% theoretical
        total_hours_85 = total_hours_100 / 0.85  # 85% efficiency (good operator)
        total_hours_70 = total_hours_100 / 0.70  # 70% efficiency (realistic with breaks/issues)
        
        return bars, weight, cost, total_hours_100, total_hours_85, total_hours_70
    
    # Display results
    c1, c2, c3 = st.columns(3)
    for col, (label, qty, price, color) in zip(
        [c1, c2, c3],
        [("Low Quote", low, price_low, "#cce5ff"),
         ("Mid Quote", mid, price_mid, "#fff3cd"),
         ("High Quote", high, price_high, "#d4edda")]
    ):
        bars, weight, cost, hours_100, hours_85, hours_70 = calc_with_production(qty, price, cycle_time_input)
        with col:
            st.markdown(f"""
                <div style='background-color:{color}; padding:12px; border-radius:10px; margin-bottom:10px'>
                <strong>{label}</strong><br>
                Quantity: {qty:,}<br>
                Bars: {bars}<br>
                Weight: {weight:.2f} lbs<br>
                Est. Mat.Cost: ${cost:.2f}<br>
                Mat.Cost/Part: ${cost/qty:.4f}
                </div>
            """, unsafe_allow_html=True)
            
            # Enhanced production hours with multiple efficiency levels
            st.markdown(f"""
                <div style='background-color:#f8f9fa; padding:10px; border-radius:8px; border-left:4px solid #28a745'>
                <strong>📅 Production Hours</strong><br>
                <small>(Includes setup & startup time)</small><br>
                100% Theoretical: <strong>{hours_100:.1f} hrs</strong><br>
                85% Efficiency: <strong>{hours_85:.1f} hrs</strong><br>
                70% Efficiency: <strong>{hours_70:.1f} hrs</strong><br>
                Days @ 8hrs (70%): <strong>{hours_70/8:.1f} days</strong>
                </div>
            """, unsafe_allow_html=True)
    
    # CPM Reference Table
    st.markdown("#### 📊 CPM Reference (Cycles Per Minute)")
    ref_col1, ref_col2 = st.columns(2)
    with ref_col1:
        st.markdown("""
        **Standard CPM → Index Time:**
        - 75 CPM = 0.4 seconds idle time
        - 60 CPM = 0.5 seconds idle time  
        - 45 CPM = 0.7 seconds idle time
        """)
    with ref_col2:
        st.markdown("""
        **Index Time:** Non-productive time per cycle when machine indexes to next position
        
        **Total Cycle Time** = Index Time + Longest Operation Time
        
        **Setup Times (included above):**
        - Bar change: ~45 seconds
        - Daily startup: ~20 minutes/shift
        """)
    
    # Store quote data in session state for cross-tab persistence
    st.session_state["quote_data"] = {
        "low_qty": low, "mid_qty": mid, "high_qty": high,
        "price_low": price_low, "price_mid": price_mid, "price_high": price_high,
        "cycle_time_input": cycle_time_input, "parts_per_bar": parts_per_bar, "bar_weight": bar_weight
    }
    
    # Return quote data for cross-tab access
    return st.session_state["quote_data"]
        
def thread_calculator_section(setup_rpm=600):
    st.header("🧮 Threading Calculator")
    
    # Check for auto-fill from tool selection redirection
    auto_fill_tool = st.session_state.get("threading_tool_type", None)
    auto_fill_position = st.session_state.get("threading_position", None)
    
    if auto_fill_tool:
        st.success(f"🎯 **Auto-filled for {auto_fill_tool}** from {auto_fill_position}")
        if st.button("🔄 Clear Auto-fill", key="clear_autofill"):
            st.session_state.pop("threading_tool_type", None)
            st.session_state.pop("threading_position", None)
            st.rerun()

    # Get setup RPM from Job Setup tab
    current_setup_rpm = st.session_state.get("setup_rpm", setup_rpm)
    cycle_time_from_tab2 = st.session_state.get("cycle_time_from_tab2", 2.4)
    
    # Check if book cycle time recommendation is available from CAM Operations
    manual_cycle_time = st.session_state.get("book_cycle_time_recommendation", None)
    if manual_cycle_time:
        cycle_time_from_tab2 = manual_cycle_time
    
    # Input specifications section
    st.markdown("### 📋 Threading Specifications")
    specs_col1, specs_col2, specs_col3 = st.columns([1, 1, 1])
    
    with specs_col1:
        st.markdown("#### Input Parameters")
        tpi = st.number_input("Threads per Inch (TPI)", value=24.0, step=0.1, format="%.1f", key="thread_calc_tpi")
        thread_len = st.number_input("Thread Length (in)", value=0.375, step=0.001, format="%.3f", key="thread_calc_length")
        cycle_time = st.number_input("Cycle Time (sec)", value=cycle_time_from_tab2, step=0.1, format="%.2f", key="thread_calc_cycle")
        
        cycle_rate_options = ["75 CPM (.4)", "60 CPM (.5)", "45 CPM (.66)"]
        cycle_rate = st.selectbox("Machine Rate", cycle_rate_options, key="thread_calc_rate")
        
        work_rpm = st.number_input("Work RPM", value=current_setup_rpm, step=100, key="thread_calc_rpm")
        if work_rpm == current_setup_rpm:
            st.caption("✅ From Job Setup")
            
        # Extract index time from the selected option and calculate working time dynamically
        index_map = {"75 CPM (.4)": 0.4, "60 CPM (.5)": 0.5, "45 CPM (.66)": 0.66}
        index_time = index_map[cycle_rate]
        working_time = cycle_time - index_time
        
        st.markdown(f"**Working Time:** {working_time:.2f} sec")
        st.caption(f"Cycle Time ({cycle_time:.2f}) - Index Time ({index_time:.2f})")
        
        st.markdown(f"**Index Time:** {index_time:.2f} sec")
        st.caption(f"From {cycle_rate} machine rate")
        
    with specs_col2:
        # Column 2 content moved up to align with column 1
        pass
        
    with specs_col3:
        # Column 3 content moved up to align with column 1
        pass

    # Store specifications in session state for persistence
    st.session_state["thread_cam_data"] = {
        "cam_rpm": work_rpm,
        "cam_thread_length": thread_len,
        "cam_pitch": tpi,
        "cam_cycle_time": cycle_time,
        "cam_cycle_rate": cycle_rate
    }

    # 3-column layout for calculations
    col1, col2, col3 = st.columns(3)

    with col1:
        st.markdown("#### 1️⃣ Calculate Rise Needed")
        st.caption("📐 Determine required cam rise based on threading method")
        
        method_options = [
            "6:1 Threading (Steel)",
            "2:1 Threading (Brass)", 
            "4:1 Threading (Hybrid)"
        ]
        method = st.selectbox("Threading Method", method_options, key="thread_calc_method")
        
        # Calculate threading RPM based on method (using corrected formulas)
        if "6:1" in method:
            # Manual: 32 driver/32 driven × 21 driver/28 driven = 0.75 ratio
            threading_rpm = work_rpm * (32/32) * (21/28)  # 0.75 × work RPM
            cam_spaces = 32.5  # 6:1 uses 0 to 32.5 hundredths
            method_short = "6:1"
        elif "2:1" in method:
            # Manual: 36 driver/27 driven = 1.33 ratio
            threading_rpm = work_rpm * (36/27)  # 1.333 × work RPM per manual
            cam_spaces = 25.0  # 2:1 uses 0 to 25 hundredths  
            method_short = "2:1"
        elif "4:1" in method:
            # Manual: Half speed method
            threading_rpm = work_rpm / 2  # Half speed
            cam_spaces = 25.0  # 4:1 uses brass cam (0 to 25 hundredths)
            method_short = "4:1"
        
        # Calculate differential RPM (manual methodology)
        differential_rpm = abs(work_rpm - threading_rpm)
        
        st.markdown(f"##### 🔧 {method_short} Threading Calculations")
        st.write(f"Threading RPM: {threading_rpm:.0f}")
        st.write(f"Differential RPM: {differential_rpm:.0f}")
        st.write(f"Rev/Second: {differential_rpm/60:.2f}")
        
        # Calculate effective revolutions (manual formula)
        # Manual: Rev/sec × working_time × cam_percentage
        revolutions_per_second = differential_rpm / 60
        total_revolutions_working_time = revolutions_per_second * working_time
        
        # Apply cam space percentage
        cam_percentage = cam_spaces / 50.0  # Convert to percentage of full cycle
        effective_revolutions = total_revolutions_working_time * cam_percentage
        
        st.write(f"Total Revs (working time): {total_revolutions_working_time:.2f}")
        st.write(f"Cam Spaces Used: {cam_spaces} of 50 ({cam_percentage*100:.1f}%)")
        st.write(f"Effective Revolutions: {effective_revolutions:.3f}")
        
        # Calculate rise needed (manual formula)
        rise_needed = effective_revolutions / tpi
        
        st.info(f"**🎯 Rise Needed: {rise_needed:.4f}\"**")
        
        # Manual validation example
        if abs(work_rpm - 1810) < 10 and abs(thread_len - 0.375) < 0.001 and abs(tpi - 24.0) < 0.1:
            st.markdown("##### ✅ Manual Example Validation")
            # Manual: 1810 → 453 differential → 7.54 rev/sec → 18.1 total → 11.765 effective → 0.490 rise
            manual_differential = 1810 - (1810 * 0.75)  # 453 RPM
            manual_rev_sec = manual_differential / 60  # 7.54
            manual_total = manual_rev_sec * 2.4  # 18.1 (using 2.4 sec)
            manual_effective = manual_total * 0.65  # 11.765 (65% for 32.5/50)
            manual_rise = manual_effective / 24  # 0.490
            
            st.caption(f"📖 **Manual Example (1810 RPM, 0.375\", 24 TPI):**")
            st.caption(f"• Expected: 453 diff → 7.54 rev/sec → 18.1 total → 11.765 eff → 0.490 rise")
            st.caption(f"• Calculated: {differential_rpm:.0f} diff → {revolutions_per_second:.2f} rev/sec → {total_revolutions_working_time:.1f} total → {effective_revolutions:.3f} eff → {rise_needed:.3f} rise")
            
            if abs(rise_needed - 0.490) < 0.020:
                st.success("✅ **Matches manual example within tolerance!**")
            else:
                st.warning(f"⚠️ **Difference from manual: {abs(rise_needed - 0.490):.3f}\"**")
        
        # Cam selection using JSON database
        cam_suggestion = suggest_davenport_cam(rise_needed)
        st.success(f"🔧 **Recommended Cam:** {cam_suggestion}")
        
        # Extract cam rise for block setting calculation
        if "Rise:" in cam_suggestion:
            try:
                cam_rise_text = cam_suggestion.split("Rise: ")[1].split(")")[0]
                actual_cam_rise = float(cam_rise_text)
                block_setting = rise_needed / actual_cam_rise
                
                st.markdown("##### 📐 Block Setting Calculation")
                st.write(f"Rise Needed: {rise_needed:.4f}\"")
                st.write(f"Actual Cam Rise: {actual_cam_rise:.4f}\"")
                st.info(f"**Block Setting: {block_setting:.3f}**")
                
                # Block setting validation
                if 0.8 <= block_setting <= 1.2:
                    st.success("✅ Block setting within optimal range (0.8 - 1.2)")
                elif block_setting < 0.8:
                    st.warning("⚠️ Block setting low - consider smaller cam")
                else:
                    st.warning("⚠️ Block setting high - consider larger cam")
                    
                # Manual block setting validation
                if abs(rise_needed - 0.490) < 0.010 and abs(actual_cam_rise - 0.452) < 0.010:
                    expected_block = 0.490 / 0.452  # 1.084
                    st.caption(f"📖 Manual example: 0.490 ÷ 0.452 = 1.084")
                    if abs(block_setting - expected_block) < 0.010:
                        st.success("✅ Block setting matches manual example!")
                
                # Store calculated values for time calculation
                st.session_state["thread_rise_data"] = {
                    "rise_needed": rise_needed,
                    "cam_suggestion": cam_suggestion,
                    "actual_cam_rise": actual_cam_rise,
                    "block_setting": block_setting,
                    "effective_revolutions": effective_revolutions,
                    "threading_rpm": threading_rpm,
                    "differential_rpm": differential_rpm,
                    "method": method
                }
                
            except (ValueError, IndexError):
                st.caption("📐 Block Setting: Manual calculation needed")
                # Store basic data even without block setting
                st.session_state["thread_rise_data"] = {
                    "rise_needed": rise_needed,
                    "cam_suggestion": cam_suggestion,
                    "effective_revolutions": effective_revolutions,
                    "threading_rpm": threading_rpm,
                    "differential_rpm": differential_rpm,
                    "method": method
                }
        else:
            st.caption("📐 Block Setting: Manual calculation needed")

    with col2:
        st.markdown("#### 2️⃣ Calculate Cam Needed")
        st.caption("🔧 Select appropriate cam based on rise needed")
        
        # Get calculated rise data from column 1
        rise_data = st.session_state.get("thread_rise_data", {})
        specs_data = st.session_state.get("thread_cam_data", {})
        
        if not rise_data:
            st.info("💡 Complete rise calculations in column 1 first")
            st.markdown("##### Summary")
            st.write("Thread Length: Enter in specifications")
            st.write("TPI (Pitch): Enter in specifications")
            st.write("Rise Needed: Calculate in column 1")
            st.write("Cam Selection: Will show here")
        else:
            # Extract data from calculations
            rise_needed = rise_data.get("rise_needed", 0)
            method = rise_data.get("method", "6:1 Threading (Steel)")
            cam_suggestion = rise_data.get("cam_suggestion", "")
            
            st.write(f"📊 **Using Calculated Data:**")
            st.write(f"• Thread Length: {specs_data.get('cam_thread_length', 0.375):.3f}\"")
            st.write(f"• Pitch: {specs_data.get('cam_pitch', 24.0):.1f} TPI")
            st.write(f"• Rise Needed: {rise_needed:.4f}\"")
            st.write(f"• Method: {method.split()[0]}")
            
            # Show cam suggestion
            st.success(f"🔧 **Recommended Cam:** {cam_suggestion}")
            
            # Extract cam details for block setting
            if "Rise:" in cam_suggestion:
                try:
                    cam_rise_text = cam_suggestion.split("Rise: ")[1].split(")")[0]
                    actual_cam_rise = float(cam_rise_text)
                    block_setting = rise_needed / actual_cam_rise
                    
                    st.markdown("##### 📐 Block Setting Calculation")
                    st.write(f"Rise Needed: {rise_needed:.4f}\"")
                    st.write(f"Actual Cam Rise: {actual_cam_rise:.4f}\"")
                    st.info(f"**Block Setting: {block_setting:.3f}**")
                    
                    # Block setting validation
                    if 0.8 <= block_setting <= 1.2:
                        st.success("✅ Block setting within optimal range (0.8 - 1.2)")
                    elif block_setting < 0.8:
                        st.warning("⚠️ Block setting low - consider smaller cam")
                    else:
                        st.warning("⚠️ Block setting high - consider larger cam")
                        
                    # Store cam data for time calculations
                    rise_data["actual_cam_rise"] = actual_cam_rise
                    rise_data["block_setting"] = block_setting
                    st.session_state["thread_rise_data"] = rise_data
                    
                except (ValueError, IndexError):
                    st.caption("📐 Block Setting: Manual calculation needed")
            else:
                st.caption("📐 Block Setting: Manual calculation needed")

    with col3:
        # Convert to expandable info box
        with st.expander("⏱️ Calculate Time Required", expanded=False):
            st.caption("⏱️ Determine actual operation time needed")
            
            # Get calculated rise data from column 2
            rise_data = st.session_state.get("thread_rise_data", {})
            specs_data = st.session_state.get("thread_cam_data", {})
            
            if not rise_data:
                st.info("💡 Complete rise calculations in column 1 first")
                st.markdown("##### Summary")
                st.write("Thread Length: Enter in specifications")
                st.write("Pitch (TPI): Enter in specifications")
                st.write("Rise Needed: Calculate in column 1")
                st.write("Time Required: Will calculate here")
            else:
                # Extract data from calculations
                rise_needed = rise_data.get("rise_needed", 0)
                method = rise_data.get("method", "6:1 Threading (Steel)")
                threading_rpm = rise_data.get("threading_rpm", 0)
                cam_suggestion = rise_data.get("cam_suggestion", "")
                
                # Get specifications
                thread_len = specs_data.get("cam_thread_length", 0.375)
                tpi = specs_data.get("cam_pitch", 24.0)
                cycle_time = specs_data.get("cam_cycle_time", 2.4)
                cycle_rate = specs_data.get("cam_cycle_rate", "75 CPM (.4)")
                
                # Extract index time from cycle rate
                index_map = {"75 CPM (.4)": 0.4, "60 CPM (.5)": 0.5, "45 CPM (.66)": 0.66}
                index_time = index_map.get(cycle_rate, 0.4)
                
                st.write(f"📊 **Using Calculated Data:**")
                st.write(f"• Thread Length: {thread_len:.3f}\"")
                st.write(f"• Pitch: {tpi:.1f} TPI")
                st.write(f"• Rise Needed: {rise_needed:.4f}\"")
                st.write(f"• Method: {method.split()[0]}")
                st.write(f"• Threading RPM: {threading_rpm:.0f}")
                
                # Calculate total thread length (including lead threads)
                total_thread_length = thread_len + (3 / tpi)  # Thread + 3 lead threads
                total_revolutions_needed = total_thread_length * tpi
                
                st.markdown("##### 📏 Thread Requirements")
                st.write(f"Thread Length: {thread_len:.3f}\"")
                st.write(f"Lead Threads: 3 ÷ {tpi:.1f} = {3/tpi:.3f}\"")
                st.write(f"Total Length: {total_thread_length:.3f}\"")
                st.write(f"Total Revolutions: {total_revolutions_needed:.1f}")
                
                # Calculate time required based on method
                if "6:1" in method:
                    # 6:1 threading uses 32.5 cam spaces (65% of cycle)
                    cam_percentage = 32.5 / 50.0
                    effective_threading_rpm = threading_rpm  # Use calculated threading RPM
                elif "2:1" in method:
                    # 2:1 threading uses 25 cam spaces (50% of cycle)
                    cam_percentage = 25.0 / 50.0
                    effective_threading_rpm = threading_rpm
                elif "4:1" in method:
                    # 4:1 threading uses 25 cam spaces (50% of cycle)
                    cam_percentage = 25.0 / 50.0
                    effective_threading_rpm = threading_rpm
                else:
                    cam_percentage = 0.5
                    effective_threading_rpm = threading_rpm
                
                # Time calculation using differential RPM methodology
                if effective_threading_rpm > 0:
                    # Time = (Total Revolutions ÷ Rev/sec) + Index Time
                    revolutions_per_second = effective_threading_rpm / 60
                    threading_time = total_revolutions_needed / revolutions_per_second
                    
                    # Apply cam space factor
                    adjusted_threading_time = threading_time / cam_percentage
                    
                    # Add index time
                    total_time_required = adjusted_threading_time + index_time
                    
                    st.markdown("##### ⏱️ Time Calculation")
                    st.write(f"Threading RPM: {effective_threading_rpm:.0f}")
                    st.write(f"Rev/Second: {revolutions_per_second:.2f}")
                    st.write(f"Threading Time: {threading_time:.2f} sec")
                    st.write(f"Cam Factor: {cam_percentage:.1%} ({cam_percentage*50:.1f}/50 spaces)")
                    st.write(f"Adjusted Time: {adjusted_threading_time:.2f} sec")
                    st.write(f"Index Time: +{index_time:.2f} sec")
                    
                    st.success(f"**⏱️ Total Time Required: {total_time_required:.2f} sec**")
                    
                    # Compare against cycle time
                    if total_time_required <= cycle_time:
                        time_margin = cycle_time - total_time_required
                        st.success(f"✅ **Fits in cycle!** Margin: {time_margin:.2f} sec")
                    else:
                        time_over = total_time_required - cycle_time
                        st.error(f"❌ **Exceeds cycle time!** Over by: {time_over:.2f} sec")
                        st.warning("🔧 Consider: Faster threading method, higher RPM, or longer cycle")
                    
                    # Store time calculation results
                    st.session_state["thread_time_data"] = {
                        "total_time_required": total_time_required,
                        "threading_time": threading_time,
                        "adjusted_threading_time": adjusted_threading_time,
                        "total_thread_length": total_thread_length,
                        "total_revolutions_needed": total_revolutions_needed,
                        "fits_in_cycle": total_time_required <= cycle_time
                    }
                    
                else:
                    st.error("❌ Invalid threading RPM - check calculations")
            
            # Summary section
            st.markdown("##### 📋 Threading Summary")
            if rise_data and specs_data:
                st.write(f"**Method:** {method.split()[0]}")
                st.write(f"**Cam:** {cam_suggestion.split('(')[0] if cam_suggestion else 'TBD'}")
                if "block_setting" in rise_data:
                    st.write(f"**Block Setting:** {rise_data['block_setting']:.3f}")
                if "thread_time_data" in st.session_state:
                    time_data = st.session_state["thread_time_data"]
                    st.write(f"**Time Required:** {time_data['total_time_required']:.2f} sec")
                    if time_data["fits_in_cycle"]:
                        st.write("**Status:** ✅ Fits in cycle")
                    else:
                        st.write("**Status:** ❌ Exceeds cycle time")
            else:
                st.write("Complete threading specifications and calculations above")

    # Threading gears section
    st.markdown("### 🔧 Davenport Manual Threading Gear Configurations")
    st.caption("📖 Based on official Davenport instruction manual (pages 135-151)")
    
    # Get threading specifications
    stored_cam_data = st.session_state.get("thread_cam_data", {})
    setup_data = st.session_state.get("setup_data", {})
    
    material = setup_data.get("material", "Steel")
    thread_len = stored_cam_data.get("cam_thread_length", 0.375)
    tpi = stored_cam_data.get("cam_pitch", 24.0)
    
    # Get manual threading gear configurations for all methods
    threading_gears = {
        "6:1": get_manual_threading_gears("6:1"),
        "2:1": get_manual_threading_gears("2:1"), 
        "4:1": get_manual_threading_gears("4:1")
    }
    recommendations = get_threading_method_recommendation(material, thread_len, tpi)
    
    # Display threading method recommendations
    if recommendations:
        st.markdown("#### 🎯 Recommended Threading Method")
        primary_rec = recommendations[0]
        primary_method = threading_gears[primary_rec["method"]]
        
        st.success(f"**Primary Recommendation: {primary_rec['method']} Threading**")
        st.caption(f"💡 {primary_rec['reason']}")
        
        # Show primary method details
        col1, col2 = st.columns(2)
        with col1:
            st.markdown(f"**{primary_method['description']}**")
            st.write(f"**Combined Ratio:** {primary_method['combined_ratio']}")
            st.write(f"**RPM Formula:** {primary_method['rpm_formula']}")
            st.write(f"**Cam Usage:** {primary_method['cam_spaces']}")
        
        with col2:
            st.markdown("**Gear Configuration:**")
            for gear_ratio in primary_method["gear_ratios"]:
                if isinstance(gear_ratio["driver"], int):
                    st.write(f"• {gear_ratio['driver']}-{gear_ratio['driven']} ({gear_ratio['description']})")
                else:
                    st.write(f"• {gear_ratio['description']}")
            
            st.markdown("**Applications:**")
            for app in primary_method["typical_applications"]:
                st.write(f"• {app}")
        
        # Show alternative methods if available
        if len(recommendations) > 1:
            st.markdown("#### 🔄 Alternative Methods")
            for alt_rec in recommendations[1:]:
                alt_method = threading_gears[alt_rec["method"]]
                with st.expander(f"{alt_rec['method']} Threading - {alt_rec['reason']}", expanded=False):
                    col1, col2 = st.columns(2)
                    with col1:
                        st.write(f"**Ratio:** {alt_method['combined_ratio']}")
                        st.write(f"**RPM:** {alt_method['rpm_formula']}")
                        st.write(f"**Cam Usage:** {alt_method['cam_spaces']}")
                    with col2:
                        st.markdown("**Gear Configuration:**")
                        for gear_ratio in alt_method["gear_ratios"]:
                            if isinstance(gear_ratio["driver"], int):
                                st.write(f"• {gear_ratio['driver']}-{gear_ratio['driven']}")
                            else:
                                st.write(f"• {gear_ratio['description']}")
    
    # Manual gear reference table
    st.markdown("#### 📋 Complete Threading Gear Reference")
    with st.expander("View All Threading Methods", expanded=False):
        for method_key, method_data in threading_gears.items():
            st.markdown(f"**{method_data['description']}**")
            
            ref_col1, ref_col2, ref_col3 = st.columns(3)
            with ref_col1:
                st.write(f"**Ratio:** {method_data['combined_ratio']}")
                st.write(f"**RPM:** {method_data['rpm_formula']}")
            with ref_col2:
                st.write(f"**Cam Usage:** {method_data['cam_spaces']}")
            with ref_col3:
                st.markdown("**Gears:**")
                for gear_ratio in method_data["gear_ratios"]:
                    if isinstance(gear_ratio["driver"], int):
                        st.write(f"{gear_ratio['driver']}-{gear_ratio['driven']}")
                    else:
                        st.write(gear_ratio['description'])
            
            st.markdown("**Typical Applications:**")
            for app in method_data["typical_applications"]:
                st.write(f"• {app}")
            st.markdown("---")
    
    # Link to threading calculator results
    rise_data = st.session_state.get("thread_rise_data", {})
    if rise_data:
        calculated_method = rise_data.get("method", "")
        if calculated_method:
            method_short = calculated_method.split()[0]  # "6:1", "2:1", or "4:1"
            st.success(f"✅ **Threading Calculator Result:** {method_short} method recommended")
            st.caption("🔗 This matches the gear configuration shown above")

    # Auto-fill results back to CAM Operations
    if auto_fill_tool:
        st.markdown("---")
        st.markdown("### 🔄 Send Results to CAM Operations")
        
        rise_data = st.session_state.get("thread_rise_data", {})
        time_data = st.session_state.get("thread_time_data", {})
        
        if rise_data and time_data:
            col1, col2 = st.columns(2)
            with col1:
                st.metric("Total Time Required", f"{time_data.get('total_time_required', 0):.2f} sec")
                st.metric("Rise Needed", f"{rise_data.get('rise_needed', 0):.4f}\"")
            with col2:
                st.metric("Threading Method", rise_data.get("method", "").split()[0])
                if "block_setting" in rise_data:
                    st.metric("Block Setting", f"{rise_data['block_setting']:.3f}")
            
            if st.button("📤 Send Results to CAM Operations", key="send_thread_results"):
                # Get recommended gears from manual-based system
                calculated_method = rise_data.get("method", "")
                recommended_gear_string = ""
                method_short = ""
                
                if calculated_method:
                    method_short = calculated_method.split()[0]
                    threading_gears = get_all_threading_gears()
                    
                    if method_short in threading_gears:
                        selected_method = threading_gears[method_short]
                        # Get the main gear configuration
                        for gear_ratio in selected_method["gear_ratios"]:
                            if isinstance(gear_ratio["driver"], int):
                                recommended_gear_string = f"{gear_ratio['driver']}-{gear_ratio['driven']}"
                                break
                
                # Create thread results with proper structure for CAM Operations display
                thread_results = {
                    "tool_type": auto_fill_tool,
                    "position": auto_fill_position,
                    "recommended_gears": recommended_gear_string,
                    "time_required": time_data.get("total_time_required", 0),
                    "rise_needed": rise_data.get("rise_needed", 0),
                    "cam_suggestion": rise_data.get("cam_suggestion", ""),
                    "method": rise_data.get("method", ""),
                    "fits_in_cycle": time_data.get("fits_in_cycle", False)
                }
                
                # Add block setting and timing data based on calculated method
                if "block_setting" in rise_data:
                    if method_short == "6:1":
                        thread_results["block_setting_6to1"] = rise_data["block_setting"]
                        thread_results["6to1_time"] = time_data.get("total_time_required", 0)
                        thread_results["cam_suggestion_6to1"] = rise_data.get("cam_suggestion", "")
                    elif method_short == "2:1":
                        thread_results["block_setting_2to1"] = rise_data["block_setting"]
                        thread_results["2to1_time"] = time_data.get("total_time_required", 0)
                        thread_results["cam_suggestion_2to1"] = rise_data.get("cam_suggestion", "")
                    elif method_short == "4:1":
                        thread_results["block_setting_4to1"] = rise_data["block_setting"]
                        thread_results["4to1_time"] = time_data.get("total_time_required", 0)
                        thread_results["cam_suggestion_4to1"] = rise_data.get("cam_suggestion", "")
                
                st.session_state["thread_calc_results"] = thread_results
                st.success("✅ Results sent to CAM Operations!")
                st.info("🔄 Switch to the CAM Operations tab to see the results")
                
                # Trigger a rerun to refresh the display
                st.rerun()
        else:
            st.info("Complete the calculations above first, then results will appear here.")

    st.markdown("### Total Length of Thread Calculation")
    specs_data = st.session_state.get("thread_specs_data", {})
    if specs_data:
        thread_len = specs_data.get("thread_length", 0.375)
        tpi = specs_data.get("pitch", 24.0)
        total_length = thread_len + (3 / tpi)
        
        st.write(f"**Formula:** Thread Length + (3 ÷ Pitch)")
        st.write(f"**Calculation:** {thread_len:.3f}\" + (3 ÷ {tpi:.1f}) = {thread_len:.3f}\" + {3/tpi:.3f}\" = **{total_length:.4f}\"**")
        
        # Manual validation
        if abs(tpi - 20.0) < 0.1 and abs(thread_len - 0.500) < 0.001:
            expected = 0.500 + (3/20)  # 0.650"
            st.caption(f"✅ **Standard example:** 0.500\" + (3/20) = **{expected:.3f}\"**")
            if abs(total_length - expected) < 0.001:
                st.success("✅ Formula working correctly!")
    else:
        st.info("Enter threading specifications in column 1 to see calculation")

    st.markdown("### 🔧 Threading Gear Setup Instructions")
    st.caption("📖 Use the gear configurations shown above based on your threading method")
    
    # Show selected threading method from calculator
    rise_data = st.session_state.get("thread_rise_data", {})
    if rise_data:
        calculated_method = rise_data.get("method", "")
        if calculated_method:
            method_short = calculated_method.split()[0]  # "6:1", "2:1", or "4:1"
            threading_gears = get_all_threading_gears()
            
            if method_short in threading_gears:
                selected_method = threading_gears[method_short]
                
                st.success(f"**Selected Method: {selected_method['description']}**")
                
                col1, col2 = st.columns(2)
                with col1:
                    st.markdown("**Gear Installation:**")
                    for gear_ratio in selected_method["gear_ratios"]:
                        if isinstance(gear_ratio["driver"], int):
                            gear_string = f"{gear_ratio['driver']}-{gear_ratio['driven']}"
                            st.write(f"• **{gear_string}** ({gear_ratio['description']})")
                            
                            # Store recommended gears for CAM Operations
                            if st.button(f"📤 Send {gear_string} to CAM Operations", key=f"send_gear_{gear_string}"):
                                st.session_state["recommended_threading_gears"] = gear_string
                                st.success(f"✅ Gear configuration {gear_string} sent to CAM Operations!")
                        else:
                            st.write(f"• {gear_ratio['description']}")
                
                with col2:
                    st.markdown("**Setup Verification:**")
                    st.write(f"• Combined ratio: {selected_method['combined_ratio']}")
                    st.write(f"• Threading RPM: {selected_method['rpm_formula']}")
                    st.write(f"• Cam timing: {selected_method['cam_spaces']}")
                    
                    # Calculate actual threading RPM if work RPM is available
                    work_rpm = rise_data.get("work_rpm", 0)
                    if work_rpm > 0:
                        actual_threading_rpm = work_rpm * selected_method['combined_ratio']
                        st.write(f"• **Calculated:** {actual_threading_rpm:.0f} threading RPM")
    else:
        st.info("💡 Complete threading calculations above to see specific gear recommendations")
        
        # Show general manual reference
        st.markdown("**Manual Reference - Threading Gear Overview:**")
        threading_gears = get_all_threading_gears()
        
        for method_key, method_data in threading_gears.items():
            with st.expander(f"{method_data['description']}", expanded=False):
                for gear_ratio in method_data["gear_ratios"]:
                    if isinstance(gear_ratio["driver"], int):
                        st.write(f"**Gears:** {gear_ratio['driver']}-{gear_ratio['driven']} ({gear_ratio['description']})")
                    else:
                        st.write(f"**Setup:** {gear_ratio['description']}")
                st.write(f"**Result:** {method_data['rpm_formula']}")
                st.write(f"**Applications:** {', '.join(method_data['typical_applications'])}")
    
    # Auto-fill results back to CAM Operations
    if auto_fill_tool:
        st.markdown("---")
        st.markdown("### 🔄 Send Results to CAM Operations")
        
        # Check if calculations have been done
        calc_done = 'time_required_6to1' in locals() and 'recommended_gears' in locals()
        
        if calc_done:
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("6:1 Threading Time", f"{time_required_6to1:.2f} sec")
            with col2:
                st.metric("2:1 Threading Time", f"{time_required_2to1:.2f} sec")
            with col3:
                st.metric("4:1 Threading Time", f"{time_required_4to1:.2f} sec")
            
            send_col1, send_col2 = st.columns(2)
            with send_col1:
                if st.button("📤 Send Results to CAM Operations", key="send_thread_results"):
                    # Store calculated results including block settings
                    thread_results = {
                        "tool_type": auto_fill_tool,
                        "position": auto_fill_position,
                        "recommended_gears": recommended_gears[0]['gears'] if recommended_gears else "",
                        "6to1_time": time_required_6to1,
                        "2to1_time": time_required_2to1,
                        "4to1_time": time_required_4to1,
                        "rise_6to1": rise_needed_6to1,
                        "rise_2to1": rise_needed_2to1,
                        "rise_4to1": rise_needed_4to1,
                        "cam_suggestion_6to1": cam_suggestion_6to1,
                        "cam_suggestion_2to1": cam_suggestion_2to1,
                        "cam_suggestion_4to1": cam_suggestion_4to1
                    }
                    
                    # Add block settings if available
                    if 'block_setting_6to1' in locals():
                        thread_results["block_setting_6to1"] = block_setting_6to1
                    if 'block_setting_2to1' in locals():
                        thread_results["block_setting_2to1"] = block_setting_2to1
                    if 'block_setting_4to1' in locals():
                        thread_results["block_setting_4to1"] = block_setting_4to1
                    
                    st.session_state["thread_calc_results"] = thread_results
                    st.success("✅ Results sent to CAM Operations! Switch to CAM Operations tab to see auto-filled data.")
            with send_col2:
                if st.button("🔙 Return to CAM Operations", key="return_to_cam"):
                    st.info("Switch to CAM Operations tab to continue setup.")
        else:
            st.info("Complete the calculations above first, then results will appear here.")

def generate_setup_sheet(setup_data, spindle_data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Setup Sheet"
    header_font = Font(bold=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    center_align = Alignment(horizontal="center")
    ws.merge_cells("A1:K1")
    machine_name = setup_data.get("machine_type", "Davenport Model B")
    ws["A1"] = f"MACHINE: {machine_name}"
    ws["A1"].font = header_font
    ws["A1"].alignment = center_align
    ws.merge_cells("A2:K2")
    ws["A2"] = "Operator/Lead Person Layout Instructions"
    ws["A2"].font = header_font
    ws["A2"].alignment = center_align
    headers = ["Part No.", "Internal Rev.", "Tool No.", "Tool Location", "Job No.", "Orig Date", "", "Updated", "Updated by", "Approved", ""]
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col).value = header
        ws.cell(row=4, column=col).font = header_font
        ws.cell(row=4, column=col).border = border
        ws.cell(row=4, column=col).alignment = center_align
    ws["A5"] = setup_data["job_name"]
    ws["B5"] = "NA"
    ws["I5"] = "TJ"
    headers = ["Material", "", "Size", "Shape", "Feet Per M", "Bars Per M", "Lbs. Per M", "", "Collets", "Feed Finger", "Set Pads", "Burr. Collect"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=6, column=col).value = header
        ws.cell(row=6, column=col).font = header_font
        ws.cell(row=6, column=col).border = border
        ws.cell(row=6, column=col).alignment = center_align
    ws["A7"] = setup_data["material"]
    ws["C7"] = setup_data["dia"]
    ws["D7"] = setup_data["bar_shape"]
    ws["I7"] = setup_data["collets"]
    ws["J7"] = setup_data["feed_finger"]
    ws["K7"] = setup_data["set_pads"]
    ws["L7"] = setup_data["burr_collect"]
    headers = ["Spindle Speed", "", "S.F.M.", "Drill", "Tap", "Spindle Gears", "Machine Code", "", "Sec", "Feed Gears", "Thread Speed", "Threading Gears", "Eff. Rev"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=8, column=col).value = header
        ws.cell(row=8, column=col).font = header_font
        ws.cell(row=8, column=col).border = border
        ws.cell(row=8, column=col).alignment = center_align
    ws["A9"] = setup_data["rpm"]
    ws["C9"] = setup_data["sfm"]
    ws["F9"] = setup_data["spindle_gears"]
    ws["G9"] = setup_data["machine_code"]
    ws["I9"] = setup_data["cycle_time"]
    ws["J9"] = setup_data["feed_gears"]
    ws["L9"] = setup_data["thread_gears"]
    ws["M9"] = 59
    headers = ["Position", "Operation", "CAM", "CAM Spaces", "Feed", "Feed Per Rev.", "Effective Revs", "Location", "", "Tool Slide", "", "Cross Slide"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=10, column=col).value = header
        ws.cell(row=10, column=col).font = header_font
        ws.cell(row=10, column=col).border = border
        ws.cell(row=10, column=col).alignment = center_align
    for i, spindle in enumerate(spindle_data, start=12):
        ws[f"A{i}"] = spindle["position"]
        ws[f"B{i}"] = spindle["operation"]
        ws[f"C{i}"] = spindle["cam"]
        ws[f"D{i}"] = spindle["cam_spaces"]
        ws[f"E{i}"] = spindle["feed"]
        ws[f"F{i}"] = spindle["feed_per_rev"]
        ws[f"G{i}"] = spindle.get("effective_revs", 0)
        ws[f"H{i}"] = spindle["location"]
        ws[f"J{i}"] = spindle["tool_slide"]
        ws[f"L{i}"] = spindle["cross_slide"]
        for col in range(1, 13):
            ws.cell(row=i, column=col).border = border
            ws.cell(row=i, column=col).alignment = center_align
    ws["A31"] = "FORM # 409-6 10-04-11-B"
    ws.merge_cells("D31:K31")
    ws["D31"] = "CONFIDENTIAL DOCUMENT: Distribution outside of KKSP employees is strictly prohibited"
    ws["D31"].font = Font(italic=True)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def threading_prompt_from_cam_ops(spindle_data):
    threading_positions = []
    for spindle in spindle_data:
        op = spindle.get("operation", "").upper().strip()
        if op in THREADING_TOOLS:
            threading_positions.append(spindle.get("position"))
    if threading_positions:
        st.session_state["thread_calc_prompt"] = True
        st.session_state["threading_positions"] = threading_positions
    else:
        st.session_state["thread_calc_prompt"] = False
        st.session_state["threading_positions"] = []
        
def create_enhanced_3d_simulation(setup_data, spindle_data, machine_config):
    """Create enhanced 3D simulation with material removal visualization"""
    st.write("**🎬 Enhanced 3D Machining Simulation with Material Removal**")
    
    # Simulation controls
    col1, col2, col3 = st.columns(3)
    
    with col1:
        simulation_speed = st.slider("Animation Speed", 0.1, 3.0, 1.0, 0.1)
        show_toolpaths = st.checkbox("Show Tool Paths", value=True)
        show_material_removal = st.checkbox("Show Material Removal", value=True)
        
    with col2:
        workpiece_opacity = st.slider("Workpiece Opacity", 0.3, 1.0, 0.7, 0.1)
        tool_scale = st.slider("Tool Size Scale", 0.5, 3.0, 1.0, 0.1)
        animation_steps = st.slider("Animation Detail", 10, 100, 50, 10)
        
    with col3:
        view_angle = st.selectbox("View Angle", ["Isometric", "Front", "Side", "Top"])
        color_scheme = st.selectbox("Color Scheme", ["Default", "High Contrast", "Material Focus"])
        visualization_mode = st.selectbox(
            "Visualization Mode",
            ["Simple Point Cloud", "Advanced 3D Mesh"],
            index=0,  # Default to Simple Point Cloud
            help="Simple Point Cloud works better for material removal animation"
        )
        debug_mode = st.checkbox("🔍 Debug Mode", help="Show detailed debugging information")
        
    # Debug mode information
    if debug_mode:
        st.info(f"🔍 Debug Info: Bar {setup_data.get('dia', 0.5):.3f}\" x {setup_data.get('part_length', 2.0):.3f}\", Operations: {len([op for op in spindle_data if op.get('operation')])}")
        
    # Create the enhanced 3D plot
    fig = create_material_removal_simulation(
        setup_data, spindle_data, machine_config,
        show_toolpaths, show_material_removal, workpiece_opacity,
        tool_scale, animation_steps, view_angle, color_scheme, visualization_mode
    )
    
    st.plotly_chart(fig, use_container_width=True)
    
    # Add simulation stats
    if spindle_data:
        st.subheader("📊 Simulation Analysis")
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            total_operations = len([op for op in spindle_data if op.get('operation')])
            st.metric("Active Operations", total_operations)
            
        with col2:
            max_feed = max([op.get('feed', 0) for op in spindle_data if op.get('feed', 0) > 0], default=0)
            st.metric("Max Feed Rate", f"{max_feed:.4f} in/rev")
            
        with col3:
            total_material_removed = calculate_material_removal_volume(setup_data, spindle_data)
            st.metric("Material Removed", f"{total_material_removed:.3f} in³")
            
        with col4:
            if setup_data.get("rpm"):
                max_revs = max([op.get("effective_revs", 0) for op in spindle_data], default=0)
                cycle_time = max_revs * 60 / setup_data["rpm"] if max_revs > 0 else 0
                st.metric("Cycle Time", f"{cycle_time:.2f} sec")

def create_material_removal_simulation(setup_data, spindle_data, machine_config, 
                                     show_toolpaths=True, show_material_removal=True,
                                     workpiece_opacity=0.7, tool_scale=1.0, 
                                     animation_steps=50, view_angle="Isometric", 
                                     color_scheme="Default", visualization_mode="Simple Point Cloud"):
    """Create detailed material removal simulation"""
    
    # Show info about visualization mode
    if visualization_mode == "Advanced 3D Mesh":
        st.info("🔧 Using Advanced 3D Mesh mode - may have animation issues")
    else:
        st.success("✅ Using Simple Point Cloud mode - better for animation")
    
    # Get workpiece parameters
    bar_diameter = setup_data.get("dia", 0.5)
    bar_length = setup_data.get("part_length", 2.0)
    
    # Create figure
    fig = go.Figure()
    
    # Color schemes
    color_schemes = {
        "Default": {"workpiece": "lightsteelblue", "removed": "indianred", "tool": "gold"},
        "High Contrast": {"workpiece": "darkslateblue", "removed": "orangered", "tool": "yellow"},
        "Material Focus": {"workpiece": "steelblue", "removed": "crimson", "tool": "silver"}
    }
    colors = color_schemes.get(color_scheme, color_schemes["Default"])
    
    # Create initial workpiece (cylinder)
    workpiece_mesh = create_cylindrical_workpiece(bar_diameter, bar_length)
    
    # Add original workpiece
    fig.add_trace(go.Mesh3d(
        x=workpiece_mesh['x'],
        y=workpiece_mesh['y'], 
        z=workpiece_mesh['z'],
        i=workpiece_mesh['i'],
        j=workpiece_mesh['j'],
        k=workpiece_mesh['k'],
        name="Original Stock",
        color=colors["workpiece"],
        opacity=workpiece_opacity,
        showscale=False,
        lighting=dict(ambient=0.6, diffuse=0.8, specular=0.2),
        lightposition=dict(x=100, y=200, z=0)
    ))
    
    # Process each operation and show material removal
    if show_material_removal:
        removed_volume = create_material_removal_visualization(
            fig, setup_data, spindle_data, machine_config, colors, animation_steps
        )
    
    # Add tool paths
    if show_toolpaths:
        add_enhanced_toolpaths(fig, setup_data, spindle_data, machine_config, colors, tool_scale)
    
    # Set up view angle
    camera_settings = get_camera_settings(view_angle)
    
    # Configure layout with enhanced visualization
    fig.update_layout(
        title=f"3D Machining Simulation - {setup_data.get('job_name', 'Part')}",
        scene=dict(
            xaxis_title="X (inches)",
            yaxis_title="Y (inches)", 
            zaxis_title="Z (inches)",
            aspectmode="data",
            camera=camera_settings,
            bgcolor="white"
        ),
        height=700,
        showlegend=True,
        legend=dict(x=0.02, y=0.98),
        annotations=[
            dict(
                text="🔴 Red dots = material removal | 🔵 Blue cylinder = workpiece | ▶️ Click play button above for step-by-step animation",
                showarrow=False,
                xref="paper", yref="paper",
                x=0.5, y=0.02, xanchor='center', yanchor='bottom',
                font=dict(size=11, color="darkblue"),
                bgcolor="lightyellow",
                bordercolor="orange",
                borderwidth=1
            )
        ] if show_material_removal else []
    )
    
    return fig

def create_cylindrical_workpiece(diameter, length, resolution=20):
    """Create a cylindrical workpiece mesh"""
    radius = diameter / 2
    
    # Create cylinder with end caps
    n_theta = resolution
    n_z = max(8, int(length * 10))  # More z-divisions for longer parts
    
    # Generate coordinates
    theta = np.linspace(0, 2*np.pi, n_theta, endpoint=False)
    z_vals = np.linspace(0, length, n_z)
    
    vertices = []
    
    # Add center points for end caps
    vertices.append([0, 0, 0])       # Bottom center
    vertices.append([0, 0, length])  # Top center
    
    # Add cylindrical surface points
    for z_val in z_vals:
        for theta_val in theta:
            x = radius * np.cos(theta_val)
            y = radius * np.sin(theta_val)
            vertices.append([x, y, z_val])
    
    vertices = np.array(vertices)
    x, y, z = vertices[:, 0], vertices[:, 1], vertices[:, 2]
    
    # Create faces
    faces = []
    
    # Bottom cap (connect to bottom center)
    for i in range(n_theta):
        next_i = (i + 1) % n_theta
        v_center = 0  # Bottom center
        v_curr = 2 + i  # Current bottom edge point
        v_next = 2 + next_i  # Next bottom edge point
        faces.append([v_center, v_next, v_curr])
    
    # Top cap (connect to top center)
    top_layer_offset = 2 + (n_z - 1) * n_theta
    for i in range(n_theta):
        next_i = (i + 1) % n_theta
        v_center = 1  # Top center
        v_curr = top_layer_offset + i  # Current top edge point
        v_next = top_layer_offset + next_i  # Next top edge point
        faces.append([v_center, v_curr, v_next])
    
    # Cylindrical surface
    for z_idx in range(n_z - 1):
        for theta_idx in range(n_theta):
            next_theta = (theta_idx + 1) % n_theta
            
            # Current layer indices
            v1 = 2 + z_idx * n_theta + theta_idx
            v2 = 2 + z_idx * n_theta + next_theta
            
            # Next layer indices
            v3 = 2 + (z_idx + 1) * n_theta + theta_idx
            v4 = 2 + (z_idx + 1) * n_theta + next_theta
            
            # Two triangles per quad
            faces.extend([[v1, v2, v3], [v2, v4, v3]])
    
    faces = np.array(faces)
    
    return {
        'x': x, 'y': y, 'z': z,
        'i': faces[:, 0], 'j': faces[:, 1], 'k': faces[:, 2]
    }

def create_material_removal_visualization(fig, setup_data, spindle_data, machine_config, colors, steps):
    """Create simplified material removal visualization with step-by-step animation"""
    
    removed_volumes = []
    debug_info = []
    bar_diameter = setup_data.get("dia", 0.5)
    bar_length = setup_data.get("part_length", 2.0)
    
    # Get all operations with material removal
    valid_operations = []
    for i, operation in enumerate(spindle_data):
        if not operation.get('operation'):
            debug_info.append(f"Skipped operation {i}: No operation type")
            continue
            
        # Create simple removal visualization
        removal_viz = create_simple_removal_zone(operation, setup_data, machine_config)
        
        if removal_viz and len(removal_viz.get('x', [])) > 0:
            valid_operations.append((operation, removal_viz))
            removed_volumes.append(removal_viz.get('volume', 0))
            debug_info.append(f"✅ Added removal for {operation.get('position', '')}: {operation.get('operation', '')} - Volume: {removal_viz.get('volume', 0):.6f} in³")
        else:
            debug_info.append(f"⚠️ No removal data for {operation.get('position', '')}: {operation.get('operation', '')}")
    
    # Create animation frames if we have operations
    if valid_operations and steps > 1:
        # Create frames for step-by-step animation
        frames = []
        
        for step in range(steps + 1):
            frame_data = []
            
            # Always show workpiece
            workpiece_mesh = create_cylindrical_workpiece(bar_diameter, bar_length)
            frame_data.append(go.Mesh3d(
                x=workpiece_mesh['x'],
                y=workpiece_mesh['y'], 
                z=workpiece_mesh['z'],
                i=workpiece_mesh['i'],
                j=workpiece_mesh['j'],
                k=workpiece_mesh['k'],
                name="Workpiece",
                color=colors["workpiece"],
                opacity=0.7,
                showscale=False
            ))
            
            # Show operations up to current step
            operations_to_show = int((step / steps) * len(valid_operations)) if steps > 0 else len(valid_operations)
            
            for i, (operation, removal_viz) in enumerate(valid_operations[:operations_to_show]):
                # Add removal visualization as scatter points
                frame_data.append(go.Scatter3d(
                    x=removal_viz['x'],
                    y=removal_viz['y'],
                    z=removal_viz['z'],
                    mode='markers',
                    marker=dict(
                        size=6,
                        color=colors["removed"],
                        opacity=0.9,
                        symbol='circle'
                    ),
                    name=f"Removed by {operation.get('position', '')}: {operation.get('operation', '')}",
                    showlegend=(step == steps)  # Only show legend on final frame
                ))
            
            frames.append(go.Frame(data=frame_data, name=f"step_{step}"))
        
        # Set initial data to final frame
        if frames:
            for trace in frames[-1].data:
                fig.add_trace(trace)
            
            # Add animation controls
            fig.frames = frames
            fig.update_layout(
                updatemenus=[{
                    "type": "buttons",
                    "showactive": False,
                    "x": 0.1,
                    "y": 1.15,
                    "buttons": [
                        {
                            "label": "▶️ Play Animation",
                            "method": "animate",
                            "args": [None, {
                                "frame": {"duration": 1000, "redraw": True},
                                "fromcurrent": True,
                                "transition": {"duration": 300}
                            }]
                        },
                        {
                            "label": "⏸️ Pause",
                            "method": "animate",
                            "args": [[None], {
                                "frame": {"duration": 0, "redraw": False},
                                "mode": "immediate",
                                "transition": {"duration": 0}
                            }]
                        }
                    ]
                }],
                sliders=[{
                    "steps": [
                        {
                            "args": [[f"step_{k}"], {
                                "frame": {"duration": 500, "redraw": True},
                                "mode": "immediate",
                                "transition": {"duration": 200}
                            }],
                            "label": f"Step {k}",
                            "method": "animate"
                        } for k in range(steps + 1)
                    ],
                    "active": steps,
                    "currentvalue": {"prefix": "Animation Step: "},
                    "len": 0.8,
                    "x": 0.1,
                    "xanchor": "left",
                    "y": 0.05,
                    "yanchor": "top"
                }]
            )
    else:
        # Static visualization - just add all operations at once
        for operation, removal_viz in valid_operations:
            fig.add_trace(go.Scatter3d(
                x=removal_viz['x'],
                y=removal_viz['y'],
                z=removal_viz['z'],
                mode='markers',
                marker=dict(
                    size=6,
                    color=colors["removed"],
                    opacity=0.9,
                    symbol='circle'
                ),
                name=f"Removed by {operation.get('position', '')}: {operation.get('operation', '')}",
                showlegend=True
            ))
    
    # Display debug information
    if debug_info:
        with st.expander("🔧 Material Removal Debug Info"):
            for info in debug_info:
                if "✅" in info:
                    st.success(info)
                elif "❌" in info:
                    st.error(info)
                elif "⚠️" in info:
                    st.warning(info)
                else:
                    st.info(info)
    
    # Add helpful info
    if len(removed_volumes) > 0:
        if steps > 1:
            st.info(f"🎬 Animation: Click ▶️ Play to see {len(valid_operations)} operations in sequence. Use the slider to step through manually.")
        else:
            st.info(f"🔧 Material Removal: Showing {len(valid_operations)} operations as red point clouds.")
    else:
        st.warning("⚠️ No material removal detected. Check that operations have feed rates and effective revolutions.")
    
    return sum(removed_volumes)

def create_simple_removal_zone(operation, setup_data, machine_config):
    """Create a simple point-based visualization of material removal"""
    
    position = operation.get("location", 1)
    tool_type = operation.get("operation", "")
    effective_revs = operation.get("effective_revs", 0)
    feed = operation.get("feed", 0.005)
    
    if effective_revs == 0 or feed == 0:
        return None
    
    bar_diameter = setup_data.get("dia", 0.5)
    bar_length = setup_data.get("part_length", 2.0)
    bar_radius = bar_diameter / 2
    
    # Determine removal location and size
    if "End" in operation.get("position", ""):
        # End working - create points along axis
        depth = min(effective_revs * feed, bar_length * 0.8)
        n_points = max(10, int(depth * 30))  # More points for better visibility
        
        z_points = np.linspace(bar_length - depth, bar_length, n_points)
        
        # Add some radial spread for drilling operations
        if tool_type.upper() in ["DRILL", "REAMER", "TAP", "DIE HEAD", "CENTER"]:
            tool_radius = min(0.05, bar_radius * 0.4)
            
            # Create spiral pattern for better visualization
            theta = np.linspace(0, 4*np.pi, n_points)  # Two full rotations
            radius_variation = np.linspace(0, tool_radius, n_points)
            
            x_points = radius_variation * np.cos(theta)
            y_points = radius_variation * np.sin(theta)
        else:
            # Center line for other tools
            x_points = np.zeros(n_points)
            y_points = np.zeros(n_points)
        
        volume = np.pi * (0.05)**2 * depth
        
    else:
        # Side working - create points around circumference
        num_positions = machine_config.get("positions", 5)
        z_position = (position - 1) * (bar_length / num_positions)
        
        penetration = min(effective_revs * feed, bar_radius * 0.4)
        removal_radius = bar_radius - penetration/2
        
        n_points = 30  # More points for better visualization
        theta = np.linspace(0, 2*np.pi, n_points)
        
        x_points = removal_radius * np.cos(theta)
        y_points = removal_radius * np.sin(theta)
        z_points = np.full(n_points, z_position)
        
        # Add variation in z for tool width
        tool_width = min(0.1, bar_length / num_positions * 0.5)
        z_variation = np.random.uniform(-tool_width/2, tool_width/2, n_points)
        z_points += z_variation
        
        # Add multiple layers for better visibility
        layers = 3
        all_x, all_y, all_z = [], [], []
        for layer in range(layers):
            layer_radius = removal_radius + (layer * penetration / layers / 2)
            layer_x = layer_radius * np.cos(theta)
            layer_y = layer_radius * np.sin(theta)
            layer_z = z_points + (layer - 1) * tool_width / layers / 3
            
            all_x.extend(layer_x)
            all_y.extend(layer_y)
            all_z.extend(layer_z)
        
        x_points = np.array(all_x)
        y_points = np.array(all_y)
        z_points = np.array(all_z)
        
        volume = 2 * np.pi * removal_radius * tool_width * penetration
    
    return {
        'x': x_points,
        'y': y_points,
        'z': z_points,
        'volume': volume
    }

def calculate_operation_removal(operation, setup_data, machine_config):
    """Calculate the material removal for a specific operation"""
    
    position = operation.get("location", 1)
    tool_type = operation.get("operation", "")
    effective_revs = operation.get("effective_revs", 0)
    feed = operation.get("feed", 0.005)
    
    # Enhanced debugging
    debug_msg = f"🔍 Processing: {operation.get('position', 'Unknown')}, Tool: {tool_type}, Revs: {effective_revs}, Feed: {feed}"
    
    # Debug: Check if we have the required data
    if effective_revs == 0:
        debug_msg += " ❌ No effective revs"
        return None
    if feed == 0:
        debug_msg += " ❌ No feed rate"
        return None
    
    bar_diameter = setup_data.get("dia", 0.5)
    bar_length = setup_data.get("part_length", 2.0)
    
    debug_msg += f" ✅ Bar: {bar_diameter:.3f}\" x {bar_length:.3f}\""
    
    # Calculate removal geometry based on tool type and position
    if "End" in operation.get("position", ""):
        # End-working operation (drilling, tapping, etc.)
        debug_msg += " 🔧 End-working"
        return create_end_working_removal(bar_diameter, bar_length, effective_revs, feed, tool_type)
    else:
        # Side-working operation (turning, forming, etc.)
        debug_msg += " 🔧 Side-working"
        return create_side_working_removal(bar_diameter, bar_length, position, effective_revs, feed, tool_type, machine_config)

def create_end_working_removal(diameter, length, effective_revs, feed, tool_type):
    """Create removal geometry for end-working operations"""
    
    # Tool diameter estimation based on type
    tool_diameters = {
        "DRILL": 0.1, "CENTER": 0.05, "TAP": 0.08, "DIE HEAD": 0.08, "REAMER": 0.12,
        "COUNTERBORE": 0.15, "BROACH": 0.06
    }
    
    tool_dia = tool_diameters.get(tool_type.upper(), 0.1)
    tool_radius = tool_dia / 2
    depth = min(effective_revs * feed, length * 0.8)  # Don't drill through entire part
    
    # Create cylindrical hole with proper mesh
    n_theta = 16
    n_z = 8
    
    # Generate coordinates
    theta = np.linspace(0, 2*np.pi, n_theta, endpoint=False)
    z_vals = np.linspace(length - depth, length, n_z)
    
    vertices = []
    
    # Add center points for end caps
    vertices.append([0, 0, length - depth])  # Bottom center
    vertices.append([0, 0, length])          # Top center
    
    # Add cylindrical surface points
    for z_val in z_vals:
        for theta_val in theta:
            x = tool_radius * np.cos(theta_val)
            y = tool_radius * np.sin(theta_val)
            vertices.append([x, y, z_val])
    
    vertices = np.array(vertices)
    x, y, z = vertices[:, 0], vertices[:, 1], vertices[:, 2]
    
    # Create faces for cylindrical surface and end caps
    faces = []
    
    # Bottom cap (connect to bottom center)
    for i in range(n_theta):
        next_i = (i + 1) % n_theta
        v_center = 0  # Bottom center
        v_curr = 2 + i  # Current bottom edge point
        v_next = 2 + next_i  # Next bottom edge point
        faces.append([v_center, v_next, v_curr])
    
    # Top cap (connect to top center)
    top_layer_offset = 2 + (n_z - 1) * n_theta
    for i in range(n_theta):
        next_i = (i + 1) % n_theta
        v_center = 1  # Top center
        v_curr = top_layer_offset + i  # Current top edge point
        v_next = top_layer_offset + next_i  # Next top edge point
        faces.append([v_center, v_curr, v_next])
    
    # Cylindrical surface
    for z_idx in range(n_z - 1):
        for theta_idx in range(n_theta):
            next_theta = (theta_idx + 1) % n_theta
            
            # Current layer indices
            v1 = 2 + z_idx * n_theta + theta_idx
            v2 = 2 + z_idx * n_theta + next_theta
            
            # Next layer indices
            v3 = 2 + (z_idx + 1) * n_theta + theta_idx
            v4 = 2 + (z_idx + 1) * n_theta + next_theta
            
            # Two triangles per quad
            faces.extend([[v1, v2, v3], [v2, v4, v3]])
    
    faces = np.array(faces)
    volume = np.pi * tool_radius**2 * depth
    
    return {
        'x': x, 'y': y, 'z': z,
        'i': faces[:, 0], 'j': faces[:, 1], 'k': faces[:, 2],
        'volume': volume
    }

def create_side_working_removal(diameter, length, position, effective_revs, feed, tool_type, machine_config):
    """Create removal geometry for side-working operations"""
    
    # Position along the bar
    num_positions = machine_config.get("positions", 5)
    z_position = (position - 1) * (length / num_positions)
    
    # Tool penetration
    penetration = min(effective_revs * feed, diameter * 0.3)  # Limit penetration
    original_radius = diameter / 2
    new_radius = max(0.05, original_radius - penetration)
    
    # Create annular removal volume
    n_theta = 24
    n_z = 6
    z_range = min(0.15, length / num_positions * 0.8)  # Tool width
    
    # Generate coordinates for annular section
    theta = np.linspace(0, 2*np.pi, n_theta, endpoint=False)
    z_vals = np.linspace(z_position - z_range/2, z_position + z_range/2, n_z)
    
    vertices = []
    
    # Create vertices for outer and inner surfaces
    for z_val in z_vals:
        for theta_val in theta:
            # Outer surface (original radius)
            x_outer = original_radius * np.cos(theta_val)
            y_outer = original_radius * np.sin(theta_val)
            vertices.append([x_outer, y_outer, z_val])
            
            # Inner surface (after material removal)
            x_inner = new_radius * np.cos(theta_val)
            y_inner = new_radius * np.sin(theta_val)
            vertices.append([x_inner, y_inner, z_val])
    
    vertices = np.array(vertices)
    x, y, z = vertices[:, 0], vertices[:, 1], vertices[:, 2]
    
    # Create faces for the annular section
    faces = []
    
    for z_idx in range(n_z - 1):
        for theta_idx in range(n_theta):
            next_theta = (theta_idx + 1) % n_theta
            
            # Vertices for current layer
            outer_curr = z_idx * (2 * n_theta) + theta_idx * 2
            inner_curr = z_idx * (2 * n_theta) + theta_idx * 2 + 1
            outer_next = z_idx * (2 * n_theta) + next_theta * 2
            inner_next = z_idx * (2 * n_theta) + next_theta * 2 + 1
            
            # Vertices for next layer
            outer_curr_next = (z_idx + 1) * (2 * n_theta) + theta_idx * 2
            inner_curr_next = (z_idx + 1) * (2 * n_theta) + theta_idx * 2 + 1
            outer_next_next = (z_idx + 1) * (2 * n_theta) + next_theta * 2
            inner_next_next = (z_idx + 1) * (2 * n_theta) + next_theta * 2 + 1
            
            # Outer surface quads
            faces.extend([
                [outer_curr, outer_next, outer_curr_next],
                [outer_next, outer_next_next, outer_curr_next]
            ])
            
            # Inner surface quads (reversed for proper normals)
            faces.extend([
                [inner_curr_next, inner_next, inner_curr],
                [inner_curr_next, inner_next_next, inner_next]
            ])
            
            # Connecting faces (radial faces)
            faces.extend([
                [outer_curr, inner_curr, outer_curr_next],
                [inner_curr, inner_curr_next, outer_curr_next]
            ])
    
    # End caps
    for z_idx in [0, n_z - 1]:  # First and last z layers
        for theta_idx in range(n_theta):
            next_theta = (theta_idx + 1) % n_theta
            
            outer_curr = z_idx * (2 * n_theta) + theta_idx * 2
            inner_curr = z_idx * (2 * n_theta) + theta_idx * 2 + 1
            outer_next = z_idx * (2 * n_theta) + next_theta * 2
            inner_next = z_idx * (2 * n_theta) + next_theta * 2 + 1
            
            if z_idx == 0:  # Bottom cap
                faces.extend([
                    [outer_curr, inner_curr, outer_next],
                    [inner_curr, inner_next, outer_next]
                ])
            else:  # Top cap
                faces.extend([
                    [outer_curr, outer_next, inner_curr],
                    [inner_curr, outer_next, inner_next]
                ])
    
    faces = np.array(faces)
    
    # Calculate volume of material removed
    volume = np.pi * (original_radius**2 - new_radius**2) * z_range
    
    return {
        'x': x, 'y': y, 'z': z,
        'i': faces[:, 0] if len(faces) > 0 else [0],
        'j': faces[:, 1] if len(faces) > 0 else [0], 
        'k': faces[:, 2] if len(faces) > 0 else [0],
        'volume': volume
    }

def add_enhanced_toolpaths(fig, setup_data, spindle_data, machine_config, colors, tool_scale):
    """Add enhanced tool path visualization"""
    
    for i, operation in enumerate(spindle_data):
        if not operation.get('operation'):
            continue
            
        position = operation.get("location", 1)
        tool_type = operation.get("operation", "")
        effective_revs = operation.get("effective_revs", 0)
        feed = operation.get("feed", 0.005)
        
        if effective_revs == 0:
            continue
        
        bar_diameter = setup_data.get("dia", 0.5)
        bar_length = setup_data.get("part_length", 2.0)
        
        # Generate tool path
        if "End" in operation.get("position", ""):
            # End-working tool path
            path_coords = generate_end_working_path(
                bar_diameter, bar_length, effective_revs, feed, tool_scale
            )
        else:
            # Side-working tool path
            path_coords = generate_side_working_path(
                bar_diameter, bar_length, position, effective_revs, feed, machine_config, tool_scale
            )
        
        if path_coords:
            # Add tool path line
            fig.add_trace(go.Scatter3d(
                x=path_coords['x'],
                y=path_coords['y'],
                z=path_coords['z'],
                mode='lines+markers',
                name=f"Path: {operation.get('position', '')}: {tool_type}",
                line=dict(color=colors["tool"], width=4),
                marker=dict(size=3, color=colors["tool"])
            ))
            
            # Add tool representation at current position
            tool_coords = create_tool_geometry(path_coords, tool_type, tool_scale)
            if tool_coords:
                fig.add_trace(go.Scatter3d(
                    x=tool_coords['x'],
                    y=tool_coords['y'],
                    z=tool_coords['z'],
                    mode='markers',
                    name=f"Tool: {tool_type}",
                    marker=dict(
                        size=8 * tool_scale,
                        color=colors["tool"],
                        symbol='diamond',
                        opacity=0.8
                    ),
                    showlegend=False
                ))

def generate_end_working_path(diameter, length, effective_revs, feed, tool_scale):
    """Generate tool path for end-working operations"""
    
    total_depth = effective_revs * feed
    z_start = length + 0.1  # Start above workpiece
    z_end = length - total_depth
    
    # Create drilling path
    z_path = np.linspace(z_start, z_end, 50)
    x_path = np.zeros_like(z_path)
    y_path = np.zeros_like(z_path)
    
    return {'x': x_path, 'y': y_path, 'z': z_path}

def generate_side_working_path(diameter, length, position, effective_revs, feed, machine_config, tool_scale):
    """Generate tool path for side-working operations"""
    
    # Position along the bar
    num_positions = machine_config.get("positions", 5)
    z_position = (position - 1) * (length / num_positions)
    
    # Radial movement
    start_radius = diameter/2 + 0.1  # Start outside workpiece
    end_radius = diameter/2 - (effective_revs * feed)
    
    # Create turning path
    radius_path = np.linspace(start_radius, end_radius, 30)
    x_path = radius_path
    y_path = np.zeros_like(radius_path)
    z_path = np.full_like(radius_path, z_position)
    
    return {'x': x_path, 'y': y_path, 'z': z_path}

def create_tool_geometry(path_coords, tool_type, tool_scale):
    """Create simple tool geometry representation"""
    
    if not path_coords or len(path_coords['x']) == 0:
        return None
    
    # Get tool tip position (last point in path)
    tip_x = path_coords['x'][-1]
    tip_y = path_coords['y'][-1] 
    tip_z = path_coords['z'][-1]
    
    return {'x': [tip_x], 'y': [tip_y], 'z': [tip_z]}

def get_camera_settings(view_angle):
    """Get camera settings for different view angles"""
    
    camera_settings = {
        "Isometric": dict(eye=dict(x=1.5, y=1.5, z=1.5)),
        "Front": dict(eye=dict(x=0, y=3, z=0)),
        "Side": dict(eye=dict(x=3, y=0, z=0)),
        "Top": dict(eye=dict(x=0, y=0, z=3))
    }
    
    return camera_settings.get(view_angle, camera_settings["Isometric"])

def calculate_material_removal_volume(setup_data, spindle_data):
    """Calculate total material removal volume"""
    
    total_volume = 0
    
    for operation in spindle_data:
        if not operation.get('operation'):
            continue
            
        effective_revs = operation.get("effective_revs", 0)
        feed = operation.get("feed", 0)
        
        if effective_revs > 0 and feed > 0:
            # Simplified volume calculation
            if "End" in operation.get("position", ""):
                # Cylindrical removal
                tool_radius = 0.05  # Approximate
                depth = effective_revs * feed
                volume = np.pi * tool_radius**2 * depth
            else:
                # Side turning removal
                bar_radius = setup_data.get("dia", 0.5) / 2
                penetration = effective_revs * feed
                cut_width = 0.1  # Approximate
                volume = 2 * np.pi * bar_radius * penetration * cut_width
            
            total_volume += volume
    
    return total_volume

def simulation_section(setup_data, spindle_data, machine_config):
    """Enhanced simulation section with material removal"""
    st.header("🎬 Advanced 3D Machining Simulation")
    
    # Check if we have operations to simulate
    if not spindle_data:
        st.warning("⚠️ No operations defined. Please configure operations in the CAM Operations tab first.")
        st.info("👈 Go to **CAM Operations** tab to set up your machining operations, then return here for simulation.")
        return
    
    # Simulation type selector
    st.subheader("🎮 Simulation Options")
    
    tab1, tab2, tab3 = st.tabs(["🔥 Material Removal", "📊 Operation Analysis", "📋 Process Summary"])
    
    with tab1:
        # Enhanced 3D simulation with material removal
        create_enhanced_3d_simulation(setup_data, spindle_data, machine_config)
        
        # Real-time simulation controls
        st.subheader("⏯️ Animation Controls")
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            if st.button("▶️ Play Animation"):
                with st.spinner("🎬 Preparing animation..."):
                    animate_machining_process(setup_data, spindle_data, machine_config)
                
        with col2:
            step_mode = st.checkbox("Step Mode", help="Advance one operation at a time")
            if step_mode:
                active_ops = [op for op in spindle_data if op.get('operation')]
                if active_ops:
                    step_number = st.slider("Operation Step", 0, len(active_ops)-1, 0)
                else:
                    st.warning("No operations to step through")
                
        with col3:
            animation_speed = st.slider("Animation Speed", 0.5, 3.0, 1.0, 0.1)
            show_cutting_forces = st.checkbox("Show Cutting Forces")
                
        with col4:
            if st.button("🔄 Reset Animation"):
                if 'animation_step' in st.session_state:
                    del st.session_state['animation_step']
                st.rerun()
        
        # Step-by-step animation or full animation
        if step_mode and 'step_number' in locals():
            create_step_by_step_simulation(setup_data, spindle_data, machine_config, step_number)
        
        # Animation status
        if 'animation_step' in st.session_state:
            st.info(f"🎬 Animation in progress... Step {st.session_state.animation_step}")
        
        # Live animation display area
        animation_placeholder = st.empty()
    
    with tab2:
        # Operation analysis and optimization
        st.subheader("📈 Machining Analysis")
        
        if spindle_data:
            # Create analysis dataframe
            analysis_data = []
            total_material_removed = 0
            
            for i, op in enumerate(spindle_data):
                if op.get('operation'):
                    effective_revs = op.get('effective_revs', 0)
                    feed = op.get('feed', 0)
                    
                    # Calculate removal volume
                    if effective_revs > 0 and feed > 0:
                        if "End" in op.get("position", ""):
                            volume = np.pi * (0.05)**2 * (effective_revs * feed)  # Simplified
                        else:
                            bar_radius = setup_data.get("dia", 0.5) / 2
                            volume = 2 * np.pi * bar_radius * (effective_revs * feed) * 0.1
                        total_material_removed += volume
                    else:
                        volume = 0
                    
                    analysis_data.append({
                        "Position": op.get("position", ""),
                        "Operation": op.get("operation", ""),
                        "Feed Rate": f"{feed:.4f} in/rev",
                        "Effective Revs": f"{effective_revs:.2f}",
                        "Material Removed": f"{volume:.4f} in³",
                        "Cam": op.get("cam", ""),
                        "Status": "✅ Optimized" if volume > 0 else "⚠️ Check Setup"
                    })
            
            df_analysis = pd.DataFrame(analysis_data)
            st.dataframe(df_analysis, use_container_width=True, hide_index=True)
            
            # Summary metrics
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.metric("Total Material Removed", f"{total_material_removed:.3f} in³")
                
            with col2:
                active_ops = len([op for op in spindle_data if op.get('operation')])
                st.metric("Active Operations", active_ops)
                
            with col3:
                if setup_data.get("rpm"):
                    max_revs = max([op.get("effective_revs", 0) for op in spindle_data], default=0)
                    cycle_time = max_revs * 60 / setup_data["rpm"] if max_revs > 0 else 0
                    st.metric("Estimated Cycle Time", f"{cycle_time:.2f} sec")
                    
            with col4:
                efficiency = (active_ops / machine_config.get("positions", 5)) * 100
                st.metric("Spindle Utilization", f"{efficiency:.1f}%")
        
        # Process optimization suggestions
        st.subheader("🎯 Optimization Suggestions")
        
        suggestions = generate_optimization_suggestions(spindle_data, setup_data, machine_config)
        for suggestion in suggestions:
            st.info(f"💡 {suggestion}")
    
    with tab3:
        # Process summary and documentation
        st.subheader("📋 Process Documentation")
        
        if spindle_data:
            # Generate process summary
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("#### 🔧 Tooling Summary")
                tools_used = set()
                for op in spindle_data:
                    if op.get('operation'):
                        tools_used.add(op.get('operation'))
                
                for tool in sorted(tools_used):
                    positions = [op.get('position') for op in spindle_data if op.get('operation') == tool]
                    st.write(f"• **{tool}**: Positions {', '.join(positions)}")
            
            with col2:
                st.markdown("#### ⚙️ Setup Summary")
                st.write(f"• **Machine**: {setup_data.get('machine_type', 'N/A')}")
                st.write(f"• **Material**: {setup_data.get('material', 'N/A')}")
                st.write(f"• **Bar Size**: {setup_data.get('dia', 0):.3f}\" {setup_data.get('bar_shape', '')}")
                st.write(f"• **RPM**: {setup_data.get('rpm', 0)}")
                st.write(f"• **SFM**: {setup_data.get('sfm', 0)}")
        
        # Export options
        st.subheader("📤 Export Options")
        col1, col2, col3 = st.columns(3)
        
        with col1:
            if st.button("📊 Export Analysis Report"):
                st.success("Analysis report exported! (Feature to be implemented)")
                
        with col2:
            if st.button("🎬 Export Simulation Video"):
                st.success("Simulation video exported! (Feature to be implemented)")
                
        with col3:
            if st.button("📋 Export Process Sheet"):
                st.success("Process documentation exported! (Feature to be implemented)")

def animate_machining_process(setup_data, spindle_data, machine_config):
    """Create animated machining process showing progressive material removal"""
    
    # Get active operations
    active_ops = [op for op in spindle_data if op.get('operation')]
    
    if not active_ops:
        st.warning("⚠️ No active operations to animate")
        return
    
    st.success(f"🎬 Starting animation with {len(active_ops)} operations...")
    
    # Create containers for animation
    progress_container = st.container()
    animation_container = st.container()
    
    with progress_container:
        progress_bar = st.progress(0)
        status_text = st.empty()
        operation_details = st.empty()
    
    # Animate each operation sequentially  
    for i, operation in enumerate(active_ops):
        progress = (i + 1) / len(active_ops)
        progress_bar.progress(progress)
        
        operation_name = f"{operation.get('position', '')}: {operation.get('operation', '')}"
        status_text.markdown(f"**🔄 Operation {i+1}/{len(active_ops)}:** {operation_name}")
        
        # Show operation details
        with operation_details.container():
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Feed Rate", f"{operation.get('feed', 0):.4f} in/rev")
            with col2:
                st.metric("Effective Revs", f"{operation.get('effective_revs', 0):.2f}")
            with col3:
                if operation.get('effective_revs', 0) > 0 and setup_data.get('rpm', 0) > 0:
                    op_time = operation['effective_revs'] * 60 / setup_data['rpm']
                    st.metric("Operation Time", f"{op_time:.2f} sec")
        
        # Create and display animation frame
        fig = create_animation_frame(setup_data, active_ops[:i+1], machine_config, i+1)
        
        with animation_container:
            st.plotly_chart(fig, use_container_width=True, key=f"anim_frame_{i}")
        
        # Store animation state
        st.session_state.animation_step = i + 1
        
        # Animation delay based on operation time or default
        if operation.get('effective_revs', 0) > 0 and setup_data.get('rpm', 0) > 0:
            real_time = operation['effective_revs'] * 60 / setup_data['rpm']
            # Scale to reasonable animation time (0.5 to 3 seconds)
            animation_delay = max(0.5, min(3.0, real_time / 5))
        else:
            animation_delay = 1.0
            
        time.sleep(animation_delay)
    
    # Final status
    progress_bar.progress(1.0)
    status_text.markdown("**✅ Machining Complete!** All operations processed successfully.")
    
    # Calculate final results
    total_volume = calculate_material_removal_volume(setup_data, active_ops)
    total_time = sum([op.get('effective_revs', 0) * 60 / setup_data.get('rpm', 1) for op in active_ops if op.get('effective_revs', 0) > 0])
    
    with operation_details.container():
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Total Material Removed", f"{total_volume:.3f} in³")
        with col2:
            st.metric("Total Cycle Time", f"{total_time:.2f} sec")
        with col3:
            if total_time > 0:
                removal_rate = total_volume / total_time * 60  # in³/min
                st.metric("Material Removal Rate", f"{removal_rate:.2f} in³/min")
    
    # Clear animation state
    if 'animation_step' in st.session_state:
        del st.session_state['animation_step']

def create_animation_frame(setup_data, completed_operations, machine_config, frame_number):
    """Create a single frame of the animation showing progressive material removal"""
    
    fig = go.Figure()
    
    # Get workpiece parameters
    bar_diameter = setup_data.get("dia", 0.5)
    bar_length = setup_data.get("part_length", 2.0)
    
    # Color scheme
    colors = {"workpiece": "lightblue", "removed": "red", "active_tool": "gold", "completed": "orange"}
    
    # Create original workpiece
    workpiece_mesh = create_cylindrical_workpiece(bar_diameter, bar_length)
    
    fig.add_trace(go.Mesh3d(
        x=workpiece_mesh['x'],
        y=workpiece_mesh['y'], 
        z=workpiece_mesh['z'],
        i=workpiece_mesh['i'],
        j=workpiece_mesh['j'],
        k=workpiece_mesh['k'],
        name="Stock Material",
        color=colors["workpiece"],
        opacity=0.6,
        showscale=False
    ))
    
    # Add progressive material removal
    for i, operation in enumerate(completed_operations):
        
        # Different color for the currently active operation
        if i == len(completed_operations) - 1:
            removal_color = colors["active_tool"]
            removal_name = f"🔥 CUTTING: {operation.get('operation', '')}"
            opacity = 0.9
        else:
            removal_color = colors["removed"]
            removal_name = f"✅ {operation.get('position', '')}: {operation.get('operation', '')}"
            opacity = 0.7
        
        # Calculate and add removal geometry
        removal_data = calculate_operation_removal(operation, setup_data, machine_config)
        
        if removal_data:
            fig.add_trace(go.Mesh3d(
                x=removal_data['x'],
                y=removal_data['y'],
                z=removal_data['z'],
                i=removal_data['i'],
                j=removal_data['j'],
                k=removal_data['k'],
                name=removal_name,
                color=removal_color,
                opacity=opacity,
                showscale=False
            ))
            
            # Add animated tool position for active operation
            if i == len(completed_operations) - 1:
                add_animated_tool(fig, operation, setup_data, machine_config, colors["active_tool"])
    
    # Configure layout for animation
    fig.update_layout(
        title=f"🎬 Machining Animation - Step {frame_number}/{len(completed_operations)}",
        scene=dict(
            xaxis_title="X (inches)",
            yaxis_title="Y (inches)", 
            zaxis_title="Z (inches)",
            aspectmode="data",
            camera=dict(eye=dict(x=1.5, y=1.5, z=1.5)),
            bgcolor="white"
        ),
        height=600,
        showlegend=True,
        legend=dict(x=0.02, y=0.98)
    )
    
    return fig

def create_step_by_step_simulation(setup_data, spindle_data, machine_config, step_number):
    """Create step-by-step simulation for detailed analysis"""
    
    st.subheader(f"🔍 Step-by-Step Analysis - Operation {step_number + 1}")
    
    # Get active operations
    active_ops = [op for op in spindle_data if op.get('operation')]
    
    if step_number >= len(active_ops):
        st.warning("Invalid step number")
        return
    
    # Show operations up to current step
    current_ops = active_ops[:step_number + 1]
    current_operation = active_ops[step_number]
    
    col1, col2 = st.columns([2, 1])
    
    with col1:
        # Create the step visualization
        fig = create_animation_frame(setup_data, current_ops, machine_config, step_number + 1)
        st.plotly_chart(fig, use_container_width=True)
    
    with col2:
        # Operation details
        st.markdown("#### 🔧 Current Operation")
        st.write(f"**Position:** {current_operation.get('position', 'N/A')}")
        st.write(f"**Tool:** {current_operation.get('operation', 'N/A')}")
        st.write(f"**Feed:** {current_operation.get('feed', 0):.4f} in/rev")
        st.write(f"**Effective Revs:** {current_operation.get('effective_revs', 0):.2f}")
        st.write(f"**Cam:** {current_operation.get('cam', 'N/A')}")
        
        # Calculate cutting parameters
        if current_operation.get('effective_revs', 0) > 0 and setup_data.get('rpm', 0) > 0:
            cutting_time = current_operation['effective_revs'] * 60 / setup_data['rpm']
            st.write(f"**Cutting Time:** {cutting_time:.2f} sec")
        
        # Material removal for this operation
        removal_data = calculate_operation_removal(current_operation, setup_data, machine_config)
        if removal_data:
            st.write(f"**Material Removed:** {removal_data['volume']:.4f} in³")
        
        # Progress indicator
        progress = (step_number + 1) / len(active_ops)
        st.progress(progress)
        st.caption(f"Progress: {step_number + 1}/{len(active_ops)} operations")

def add_animated_tool(fig, operation, setup_data, machine_config, tool_color):
    """Add animated tool representation for the active operation"""
    
    position = operation.get("location", 1)
    tool_type = operation.get("operation", "")
    effective_revs = operation.get("effective_revs", 0)
    feed = operation.get("feed", 0.005)
    
    bar_diameter = setup_data.get("dia", 0.5)
    bar_length = setup_data.get("part_length", 2.0)
    
    # Tool position and size based on operation type
    if "End" in operation.get("position", ""):
        # End-working tool
        depth = effective_revs * feed
        tool_x = [0, 0, 0]
        tool_y = [0, 0, 0] 
        tool_z = [bar_length + 0.2, bar_length, bar_length - depth]
        
        # Add tool shaft
        fig.add_trace(go.Scatter3d(
            x=tool_x,
            y=tool_y,
            z=tool_z,
            mode='lines+markers',
            name=f"🔨 {tool_type} (Active)",
            line=dict(color=tool_color, width=8),
            marker=dict(size=10, color=tool_color, symbol='diamond')
        ))
        
    else:
        # Side-working tool
        num_positions = machine_config.get("positions", 5)
        z_position = (position - 1) * (bar_length / num_positions)
        penetration = effective_revs * feed
        
        tool_x = [bar_diameter/2 + 0.2, bar_diameter/2, bar_diameter/2 - penetration]
        tool_y = [0, 0, 0]
        tool_z = [z_position, z_position, z_position]
        
        # Add tool path
        fig.add_trace(go.Scatter3d(
            x=tool_x,
            y=tool_y,
            z=tool_z,
            mode='lines+markers',
            name=f"🔨 {tool_type} (Active)",
            line=dict(color=tool_color, width=8),
            marker=dict(size=10, color=tool_color, symbol='diamond')
        ))
        
        # Add cutting zone indicator
        cutting_zone_theta = np.linspace(-np.pi/4, np.pi/4, 10)
        cutting_zone_r = bar_diameter/2 - penetration/2
        cutting_x = cutting_zone_r * np.cos(cutting_zone_theta)
        cutting_y = cutting_zone_r * np.sin(cutting_zone_theta)
        cutting_z = np.full_like(cutting_x, z_position)
        
        fig.add_trace(go.Scatter3d(
            x=cutting_x,
            y=cutting_y,
            z=cutting_z,
            mode='markers',
            name="🔥 Cutting Zone",
            marker=dict(size=8, color='red', opacity=0.7),
            showlegend=False
        ))

def generate_optimization_suggestions(spindle_data, setup_data, machine_config):
    """Generate intelligent optimization suggestions"""
    suggestions = []
    
    # Check spindle utilization
    active_positions = len([op for op in spindle_data if op.get('operation')])
    total_positions = machine_config.get("positions", 5)
    utilization = active_positions / total_positions
    
    if utilization < 0.6:
        suggestions.append(f"Consider consolidating operations - only {active_positions}/{total_positions} spindles are active")
    
    # Check feed rates
    feeds = [op.get('feed', 0) for op in spindle_data if op.get('feed', 0) > 0]
    if feeds:
        avg_feed = sum(feeds) / len(feeds)
        if avg_feed < 0.005:
            suggestions.append("Feed rates appear conservative - consider increasing for better productivity")
        elif avg_feed > 0.020:
            suggestions.append("Some feed rates are aggressive - verify surface finish requirements")
    
    # Check for threading operations
    threading_ops = [op for op in spindle_data if op.get('operation', '').upper() in ['TAP', 'DIE HEAD', 'THREAD ROLL']]
    if threading_ops and not setup_data.get('thread_gears'):
        suggestions.append("Threading operations detected - ensure threading gears are properly configured")
    
    # Check effective revs balance
    revs = [op.get('effective_revs', 0) for op in spindle_data if op.get('effective_revs', 0) > 0]
    if revs and (max(revs) / min(revs)) > 3:
        suggestions.append("Large variation in effective revs detected - consider balancing cycle times")
    
    if not suggestions:
        suggestions.append("Setup looks well optimized! ✅")
    
    return suggestions

def reference_charts_section():
    """Reference Charts & Tables section with JSON data loading"""
    st.header("📚 Reference Charts & Tables")
    
    # Load JSON data
    drill_data = load_drill_charts()
    thread_data = load_threading_charts()
    
    # Drill Size Charts Section
    with st.expander("🔧 Drill Size Charts", expanded=False):
        st.markdown("### Complete Drill Size Reference")
        
        # Search functionality
        search_col1, search_col2 = st.columns([2, 1])
        with search_col1:
            drill_search = st.text_input("🔍 Search drill sizes (e.g., '#7', '0.201', '5.1mm'):", key="ref_drill_search")
        with search_col2:
            drill_filter = st.selectbox("Filter by:", ["All", "Number", "Letter", "Fractional", "Metric"], key="ref_drill_filter")
        
        # Display search results
        if drill_search:
            search_results = search_drill_sizes(drill_search, drill_data)
            if search_results:
                st.markdown(f"#### 🎯 Search Results for '{drill_search}':")
                results_data = []
                for drill in search_results:
                    results_data.append({
                        'Size': drill['size'],
                        'Decimal (inches)': f"{drill['decimal']:.4f}",
                        'MM': f"{drill['mm']:.3f}",
                        'Category': drill['category'].title()
                    })
                st.dataframe(pd.DataFrame(results_data), use_container_width=True)
            else:
                st.warning(f"No drill sizes found matching '{drill_search}'")
        
        # Show drill data if available
        if drill_data:
            tab1, tab2, tab3, tab4 = st.tabs(["🔢 Number Drills", "🔤 Letter Drills", "📐 Fractional", "📏 Metric"])
            
            with tab1:
                if 'number_drills' in drill_data:
                    number_df = pd.DataFrame(drill_data['number_drills'])
                    if not number_df.empty:
                        st.dataframe(number_df, use_container_width=True)
                        st.caption(f"📊 Total Number Drills: {len(number_df)}")
                else:
                    st.info("No number drill data available.")
            
            with tab2:
                if 'letter_drills' in drill_data:
                    letter_df = pd.DataFrame(drill_data['letter_drills'])
                    if not letter_df.empty:
                        st.dataframe(letter_df, use_container_width=True)
                        st.caption(f"📊 Total Letter Drills: {len(letter_df)}")
                else:
                    st.info("No letter drill data available.")
            
            with tab3:
                if 'fractional_drills' in drill_data:
                    frac_df = pd.DataFrame(drill_data['fractional_drills'])
                    if not frac_df.empty:
                        st.dataframe(frac_df, use_container_width=True)
                        st.caption(f"📊 Total Fractional Drills: {len(frac_df)}")
                else:
                    st.info("No fractional drill data available.")
            
            with tab4:
                if 'metric_drills' in drill_data:
                    metric_df = pd.DataFrame(drill_data['metric_drills'])
                    if not metric_df.empty:
                        st.dataframe(metric_df, use_container_width=True)
                        st.caption(f"📊 Total Metric Drills: {len(metric_df)}")
                else:
                    st.info("No metric drill data available.")
        else:
            st.error("❌ No drill data loaded. Check that drill_sizes.json exists and is properly formatted.")
    
    # Threading Charts Section
    with st.expander("🧵 Threading Charts", expanded=False):
        st.markdown("### Thread Pitch Charts, Tap Drill Sizes, Threading Specifications")
        
        # Search functionality for threads
        thread_search = st.text_input("🔍 Search threads (e.g., '1/4-20', 'M5', '3/8'):", key="ref_thread_search")
        
        # Display thread search results
        if thread_search:
            search_results = search_thread_sizes(thread_search, thread_data)
            if search_results:
                st.markdown(f"#### 🎯 Search Results for '{thread_search}':")
                results_data = []
                for thread in search_results:
                    results_data.append({
                        'Thread': thread.get('thread', ''),
                        'Tap Drill': thread.get('tap_drill', ''),
                        'Decimal': f"{thread.get('tap_drill_decimal', 0):.4f}\"",
                        'MM': f"{thread.get('tap_drill_mm', 0):.2f}mm",
                        'TPI/Pitch': thread.get('tpi', thread.get('pitch', '')),
                        'Type': thread.get('category', '').upper()
                    })
                st.dataframe(pd.DataFrame(results_data), use_container_width=True)
            else:
                st.warning(f"No threads found matching '{thread_search}'")
        
        # Show thread data if available
        if thread_data:
            thread_tab1, thread_tab2 = st.tabs(["🇺🇸 UNC/UNF Threads", "🌐 Metric Threads"])
            
            with thread_tab1:
                if 'unc_unf_threads' in thread_data and thread_data['unc_unf_threads']:
                    unc_df = pd.DataFrame(thread_data['unc_unf_threads'])
                    if not unc_df.empty:
                        st.dataframe(unc_df, use_container_width=True)
                        st.caption(f"📊 Total UNC/UNF Threads: {len(unc_df)}")
                else:
                    st.info("No UNC/UNF thread data available.")
            
            with thread_tab2:
                if 'metric_threads' in thread_data and thread_data['metric_threads']:
                    metric_df = pd.DataFrame(thread_data['metric_threads'])
                    if not metric_df.empty:
                        st.dataframe(metric_df, use_container_width=True)
                        st.caption(f"📊 Total Metric Threads: {len(metric_df)}")
                else:
                    st.info("No metric thread data available.")
        else:
            st.error("❌ No threading data loaded. Check that threading_charts.json exists and is properly formatted.")
    
    # Machining Reference Section
    with st.expander("🔧 Machining Reference", expanded=False):
        st.markdown("### SFM Charts, Feed Rates, Cutting Speeds for Different Materials")
        st.info("Material cutting data and SFM guidelines will be displayed here.")
    
    # Machine Specifications Section
    with st.expander("⚙️ Machine Specifications", expanded=False):
        st.markdown("### Davenport Model B Specifications")
        
        davenport_specs = {
            "Max RPM": "4,500",
            "Spindle Positions": "5", 
            "Standard Capacity - Round": "7/8\"",
            "Standard Capacity - Hex": "3/4\"",
            "Cycle Rates": "75, 60, 45 CPM"
        }
        
        spec_data = [{"Specification": k, "Value": v} for k, v in davenport_specs.items()]
        st.dataframe(pd.DataFrame(spec_data), use_container_width=True)
    
    # Conversion Tables Section  
    with st.expander("📊 Conversion Tables", expanded=False):
        st.markdown("### Common Unit Conversions")
        
        conversions = [
            ("Inches to MM", "× 25.4"),
            ("MM to Inches", "÷ 25.4"),
            ("IPM to MM/min", "× 25.4"),
            ("MM/min to IPM", "÷ 25.4")
        ]
        
        conversion_data = [{"Conversion": conv[0], "Factor": conv[1]} for conv in conversions]
        st.dataframe(pd.DataFrame(conversion_data), use_container_width=True)





# ------------------------------
# Main App
# ------------------------------

def main():
    # --- Cookie Setup ---
    cookies = EncryptedCookieManager(
        prefix="davenport_",  
        password="REPLACE_THIS_WITH_A_RANDOM_SECRET_STRING"
  )
    if not cookies.ready():
        st.stop()

    st.set_page_config(page_title="Davenport CAM Assistant", layout="wide")
    material_data, sfm_guidelines, legacy_cam_data = load_data()
    st.title("Multi-Spindle CAM Setup Assistant")

    if "tool_library_end" not in st.session_state:
        st.session_state.tool_library_end = load_tool_library("tool_library_end.json", DEFAULT_TOOLS_END)
    if "tool_library_side" not in st.session_state:
        st.session_state.tool_library_side = load_tool_library("tool_library_side.json", DEFAULT_TOOLS_SIDE)
    
    # Check for threading tool context to highlight Thread Calculator tab
    thread_tool_context = st.session_state.get("thread_tool_context", None)
    thread_highlight = " ⚡️" if thread_tool_context else ""
    
    # Create tabs for navigation
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "📋 Quote Builder", 
        "⚙️ CAM Operations", 
        f"🧮 Thread Calculator{thread_highlight}",
        "🎬 Simulation",
        "📚 Reference Charts"
    ])
    
    # Display content based on selected tab
    with tab1:
        setup_data = job_setup_section(material_data, sfm_guidelines)
        quote_data = quote_breakdown_section(setup_data["parts_per_bar"], setup_data["bar_weight"], setup_data.get("cycle_time", 1.6))
        # Store both setup_data and quote_data for other tabs with consistent keys
        # Note: job_setup_section already stores as "setup_data", so keep that as primary
        st.session_state["last_setup_data"] = setup_data  # Backup for compatibility
        st.session_state["last_quote_data"] = quote_data

    with tab2:
        st.subheader("CAM Operations – Spindle Setup")
        
        # Get setup_data from session state - use primary "setup_data" key
        setup_data = st.session_state.get("setup_data", {})
        
        # If no setup_data is found, try the backup key
        if not setup_data.get("job_name"):
            setup_data = st.session_state.get("last_setup_data", {
                "machine_config": DAVENPORT_CONFIG,
                "machine_type": "Davenport Model B",
                "rpm": 600
            })
            if not setup_data.get("job_name"):
                st.warning("⚠️ No job setup data found. Please fill out the Quote Builder tab first.")
            else:
                st.info(f"🔄 Retrieved job setup data for: **{setup_data.get('job_name', 'Unknown Job')}**")
        else:
            st.success(f"✅ Using job setup data for: **{setup_data.get('job_name', 'Unknown Job')}**")
        
        # Display auto-filled threading results if available
        thread_results = st.session_state.get("thread_calc_results", None)
        if thread_results:
            with st.expander("🧮 Auto-filled Threading Results", expanded=True):
                st.success(f"**Threading tool:** {thread_results['tool_type']} at {thread_results['position']}")
                
                # Display recommended gears
                st.info(f"**Recommended Gears:** {thread_results.get('recommended_gears', 'None')}")
                
                # Display timing results
                timing_col1, timing_col2, timing_col3 = st.columns(3)
                with timing_col1:
                    st.metric("6:1 Time", f"{thread_results.get('6to1_time', 0):.2f} sec")
                with timing_col2:
                    st.metric("2:1 Time", f"{thread_results.get('2to1_time', 0):.2f} sec")
                with timing_col3:
                    st.metric("4:1 Time", f"{thread_results.get('4to1_time', 0):.2f} sec")
                
                # Display block settings if available
                if any(key in thread_results for key in ["block_setting_6to1", "block_setting_2to1", "block_setting_4to1"]):
                    st.markdown("**📐 Block Settings:**")
                    block_col1, block_col2, block_col3 = st.columns(3)
                    with block_col1:
                        if "block_setting_6to1" in thread_results:
                            st.success(f"**6:1**: {thread_results['block_setting_6to1']:.3f}")
                        if "cam_suggestion_6to1" in thread_results:
                            st.caption(thread_results['cam_suggestion_6to1'])
                    with block_col2:
                        if "block_setting_2to1" in thread_results:
                            st.success(f"**2:1**: {thread_results['block_setting_2to1']:.3f}")
                        if "cam_suggestion_2to1" in thread_results:
                            st.caption(thread_results['cam_suggestion_2to1'])
                    with block_col3:
                        if "block_setting_4to1" in thread_results:
                            st.success(f"**4:1**: {thread_results['block_setting_4to1']:.3f}")
                        if "cam_suggestion_4to1" in thread_results:
                            st.caption(thread_results['cam_suggestion_4to1'])
                
                if st.button("🗑️ Clear Threading Results", key="clear_thread_results"):
                    st.session_state.pop("thread_calc_results", None)
                    st.rerun()
        
        # Get machine config from setup
        machine_config = setup_data.get("machine_config", DAVENPORT_CONFIG)
        machine_type = setup_data.get("machine_type", "Davenport Model B")
        
        # Load Davenport cam data
        cam_data = load_machine_cam_data()
        
        # Retrieve stored CAM operations data for persistence across tabs
        stored_cam_ops_data = st.session_state.get("cam_operations_data", {})
        
        # Use machine-specific cycle rates
        cycle_rates = machine_config.get("cycle_rates", [75, 60, 45])
        cpm = st.selectbox("Select Machine Cycle Rate (CPM)", options=cycle_rates, index=0)
        
        gear_table = load_gear_table(cpm)
        num_spindles = machine_config["positions"]
        st.info(f"Configuring {num_spindles} positions for {machine_type}")
        
        spindle_data = []
        all_effective_revs = []
        
        # Function to determine if this is the highest effective revs position
        def get_highlight_style(effective_revs, all_revs):
            """Return CSS style for highlighting the highest effective revs position"""
            if not all_revs or effective_revs == 0:
                return ""
            max_revs = max(all_revs)
            if effective_revs == max_revs and effective_revs > 0:
                return "background-color: #ffeb3b; border: 2px solid #ff9800; border-radius: 5px; padding: 5px;"
            return ""
        
        # First pass: collect all effective revolutions for comparison
        temp_effective_revs = []
        for i in range(1, num_spindles + 1):
            # Get current values for this position
            ew_travel = st.session_state.get(f"pos{i}_ew_travel", 0.0)
            ew_approach = st.session_state.get(f"pos{i}_ew_approach", 0.0)
            ew_feed = st.session_state.get(f"pos{i}_ew_feed", 0.0)
            sw_travel = st.session_state.get(f"pos{i}_sw_travel", 0.0)
            sw_approach = st.session_state.get(f"pos{i}_sw_approach", 0.0)
            sw_feed = st.session_state.get(f"pos{i}_sw_feed", 0.0)
            
            # Calculate effective revs for both end and side
            ew_total = ew_travel + ew_approach
            sw_total = sw_travel + sw_approach
            ew_effective_revs = ew_total / ew_feed if ew_feed > 0 else 0
            sw_effective_revs = sw_total / sw_feed if sw_feed > 0 else 0
            
            # Track the higher of the two for this position
            position_max_revs = max(ew_effective_revs, sw_effective_revs)
            if position_max_revs > 0:
                temp_effective_revs.append(position_max_revs)
        
        # Find the overall maximum
        max_effective_revs = max(temp_effective_revs) if temp_effective_revs else 0
        
        for i in range(1, num_spindles + 1):
            # Calculate current position's effective revs for highlighting
            ew_travel = st.session_state.get(f"pos{i}_ew_travel", 0.0)
            ew_approach = st.session_state.get(f"pos{i}_ew_approach", 0.0)
            ew_feed = st.session_state.get(f"pos{i}_ew_feed", 0.0)
            sw_travel = st.session_state.get(f"pos{i}_sw_travel", 0.0)
            sw_approach = st.session_state.get(f"pos{i}_sw_approach", 0.0)
            sw_feed = st.session_state.get(f"pos{i}_sw_feed", 0.0)
            
            # Calculate effective revs for both end and side
            ew_total = ew_travel + ew_approach
            sw_total = sw_travel + sw_approach
            ew_effective_revs = ew_total / ew_feed if ew_feed > 0 else 0
            sw_effective_revs = sw_total / sw_feed if sw_feed > 0 else 0
            position_max_revs = max(ew_effective_revs, sw_effective_revs)
            
            # Determine if this position should be highlighted
            is_highest = (position_max_revs == max_effective_revs and position_max_revs > 0)
            
            # Create expander title with highlight indicator
            title = f"Position {i}"
            if is_highest:
                title += f" 🎯 **HIGHEST EFFECTIVE REVS** ({position_max_revs:.2f})"
            
            with st.expander(title, expanded=is_highest):
                with st.expander("End-Working Tool", expanded=False):
                    ew_col1, ew_col2 = st.columns([2, 3])
                    with ew_col1:
                        ew_tool = enhanced_tool_selector(
                            "Tool Type (End)",
                            sorted(st.session_state.tool_library_end),
                            f"pos{i}_ew_tool",
                            "tool_library_end",
                            "tool_library_end.json",
                            TOOL_DEFINITIONS
                        )
                        # Remove redundant assignment - enhanced_tool_selector handles session state via key
                        ew_travel = st.number_input("Tool Travel (End)", min_value=0.0, step=0.0001, format="%.4f", 
                                                  value=stored_cam_ops_data.get(f"pos{i}_ew_travel", 0.0), key=f"pos{i}_ew_travel")
                        ew_approach = st.number_input("Approach (End)", min_value=0.0, step=0.001, format="%.3f", 
                                                    value=stored_cam_ops_data.get(f"pos{i}_ew_approach", 0.0), key=f"pos{i}_ew_approach")
                        ew_feed = st.number_input("Feed (End)", min_value=0.0, step=0.0001, format="%.4f", 
                                                value=stored_cam_ops_data.get(f"pos{i}_ew_feed", 0.0), key=f"pos{i}_ew_feed")
                        ew_tool_desc = st.text_input("Tool Description (End)", 
                                                   value=stored_cam_ops_data.get(f"pos{i}_ew_tool_desc", ""), key=f"pos{i}_ew_tool_desc")
                    with ew_col2:
                        ew_total = ew_travel + ew_approach
                        min_block, max_block = machine_config["block_ranges"].get(i, (0.8, 1.2))
                        st.caption(f"Valid Block Setting Range for Position {i}: {min_block}–{max_block}")
                        if ew_feed > 0:
                            ew_effective_revs = ew_total / ew_feed
                            # Show effective revs without special highlighting
                            st.info(f"Effective Revs (End): {ew_effective_revs:.2f}")
                        else:
                            ew_effective_revs = 0
                            if ew_total > 0:
                                st.warning("Feed (End) must be > 0 to calculate Effective Revs.")
                        st.caption(f"Total Travel: {ew_total:.3f}")
                        if ew_total == 0.0 or ew_tool.strip() == "":
                            if ew_total != 0.0:
                                st.warning("Please select a tool type for End-Working.")
                        else:
                            ew_cam = recommend_cam(ew_total, ew_total, ew_tool, cam_data, setup_data["material"], i, machine_config)
                            if ew_cam:
                                cam_name, cam_info = ew_cam
                                cam_rise = cam_info.get("rise", 0)
                                ew_block = ew_total / cam_rise if cam_rise else 0
                                feed_per_rev = ew_feed / setup_data["rpm"] if setup_data["rpm"] else 0
                                st.success(f"End Cam: Size {cam_info.get('size', '?')} – Cam #: {cam_name}")
                                st.caption(f"Rise: {cam_rise:.4f}, Type: {cam_info.get('type', '?')}")
                                st.markdown(f"<div style='font-size:16px; font-weight:700; padding-top:2px;'>End Block Setting: {ew_block:.2f}</div>", unsafe_allow_html=True)
                                spindle_data.append({
                                    "position": f"End{i}",
                                    "operation": ew_tool,
                                    "cam": cam_name,
                                    "cam_spaces": cam_info.get("size", ""),
                                    "feed": ew_feed,
                                    "feed_per_rev": feed_per_rev,
                                    "effective_revs": ew_effective_revs,
                                    "location": i,
                                    "tool_slide": ew_tool_desc,
                                    "cross_slide": ""
                                })
                                if ew_effective_revs > 0:
                                    all_effective_revs.append(ew_effective_revs)
                            else:
                                st.warning(f"No viable cam found for End-Working in block range {min_block}–{max_block}")
                with st.expander("Side-Working Tool", expanded=False):
                    sw_col1, sw_col2 = st.columns([2, 3])
                    with sw_col1:
                        sw_tool = enhanced_tool_selector(
                            "Tool Type (Side)",
                            sorted(st.session_state.tool_library_side),
                            f"pos{i}_sw_tool",
                            "tool_library_side",
                            "tool_library_side.json",
                            TOOL_DEFINITIONS
                        )
                        # Remove redundant assignment - enhanced_tool_selector handles session state via key
                        
                        # Special inputs for SHAVE tool
                        if sw_tool.strip().upper() == "SHAVE":
                            st.markdown("**🔧 SHAVE Tool Diameters**")
                            largest_dia = st.number_input(
                                "Largest Diameter (C5)", 
                                min_value=0.001, 
                                step=0.001, 
                                format="%.3f", 
                                key=f"pos{i}_largest_dia",
                                help="Enter the largest diameter for shave calculation"
                            )
                            smallest_dia = st.number_input(
                                "Smallest Diameter (C6)", 
                                min_value=0.001, 
                                step=0.001, 
                                format="%.3f", 
                                key=f"pos{i}_smallest_dia",
                                help="Enter the smallest diameter for shave calculation"
                            )
                            
                            # Calculate recommended rise using Excel formula
                            if largest_dia > smallest_dia:
                                calculated_rise = calculate_shave_cam_rise(largest_dia, smallest_dia)
                                st.info(f"📊 Calculated Rise: {calculated_rise:.4f} in")
                                st.caption("Formula: √((Largest Dia/2)² - (Smallest Dia/2)²) + 0.02")
                            elif largest_dia > 0 and smallest_dia > 0:
                                st.warning("⚠️ Largest diameter must be greater than smallest diameter")
                        
                        sw_travel = st.number_input("Tool Travel (Side)", min_value=0.0, step=0.0001, format="%.4f", 
                                                   value=stored_cam_ops_data.get(f"pos{i}_sw_travel", 0.0), key=f"pos{i}_sw_travel")
                        sw_approach = st.number_input("Approach (Side)", min_value=0.0, step=0.001, format="%.3f", 
                                                    value=stored_cam_ops_data.get(f"pos{i}_sw_approach", 0.0), key=f"pos{i}_sw_approach")
                        sw_feed = st.number_input("Feed (Side)", min_value=0.0, step=0.0001, format="%.4f", 
                                                value=stored_cam_ops_data.get(f"pos{i}_sw_feed", 0.0), key=f"pos{i}_sw_feed")
                        sw_tool_desc = st.text_input("Tool Description (Side)", 
                                                   value=stored_cam_ops_data.get(f"pos{i}_sw_tool_desc", ""), key=f"pos{i}_sw_tool_desc")
                    with sw_col2:
                        sw_total = sw_travel + sw_approach
                        min_block, max_block = machine_config["block_ranges"].get(i, (0.8, 1.2))
                        st.caption(f"Valid Block Setting Range for Position {i}: {min_block}–{max_block}")
                        if sw_feed > 0:
                            sw_effective_revs = sw_total / sw_feed
                            # Show effective revs without special highlighting
                            st.info(f"Effective Revs (Side): {sw_effective_revs:.2f}")
                        else:
                            sw_effective_revs = 0
                            if sw_total > 0:
                                st.warning("Feed (Side) must be > 0 to calculate Effective Revs.")
                        st.caption(f"Total Travel: {sw_total:.3f}")
                        if sw_total == 0.0 or sw_tool.strip() == "":
                            if sw_total != 0.0:
                                st.warning("Please select a tool type for Side-Working.")
                        else:
                            # For SHAVE tools, use calculated rise if diameters are provided
                            target_rise_for_cam = sw_total  # Default
                            if sw_tool.strip().upper() == "SHAVE":
                                largest_dia = st.session_state.get(f"pos{i}_largest_dia", 0)
                                smallest_dia = st.session_state.get(f"pos{i}_smallest_dia", 0)
                                if largest_dia > smallest_dia > 0:
                                    target_rise_for_cam = calculate_shave_cam_rise(largest_dia, smallest_dia)
                                    st.info(f"🎯 Using calculated rise for cam selection: {target_rise_for_cam:.4f} in")
                            
                            sw_cam = recommend_cam(target_rise_for_cam, sw_total, sw_tool, cam_data, setup_data["material"], i, machine_config)
                            if sw_cam:
                                cam_name, cam_info = sw_cam
                                cam_rise = cam_info.get("rise", 0)
                                sw_block = sw_total / cam_rise if cam_rise else 0
                                feed_per_rev = sw_feed / setup_data["rpm"] if setup_data["rpm"] else 0
                                st.success(f"Side Cam: Size {cam_info.get('size', '?')} – Cam #: {cam_name}")
                                st.caption(f"Rise: {cam_rise:.4f}, Type: {cam_info.get('type', '?')}")
                                st.markdown(f"<div style='font-size:16px; font-weight:700; padding-top:2px;'>Side Block Setting: {sw_block:.2f}</div>", unsafe_allow_html=True)
                                spindle_data.append({
                                    "position": f"Side{i}",
                                    "operation": sw_tool,
                                    "cam": cam_name,
                                    "cam_spaces": cam_info.get("size", ""),
                                    "feed": sw_feed,
                                    "feed_per_rev": feed_per_rev,
                                    "effective_revs": sw_effective_revs,
                                    "location": i,
                                    "tool_slide": "",
                                    "cross_slide": sw_tool_desc
                                })
                                if sw_effective_revs > 0:
                                    all_effective_revs.append(sw_effective_revs)
                            else:
                                st.warning(f"No viable cam found for Side-Working in block range {min_block}–{max_block}")

        schematic_station_data = []
        for i in range(1, num_spindles + 1):
            end_tool = st.session_state.get(f"pos{i}_ew_tool", "")
            side_tool = st.session_state.get(f"pos{i}_sw_tool", "")
            schematic_station_data.append({
                "end_operation": end_tool if end_tool.strip() else "",
                "side_operation": side_tool if side_tool.strip() else ""
            })
        st.session_state["schematic_station_data"] = schematic_station_data

        threading_prompt_from_cam_ops(spindle_data)

        # Add summary section showing all effective revolutions
        if any(st.session_state.get(f"pos{i}_ew_feed", 0) > 0 or st.session_state.get(f"pos{i}_sw_feed", 0) > 0 for i in range(1, num_spindles + 1)):
            st.markdown("---")
            st.markdown("### 📊 **Effective Revolutions Summary**")
            
            # Create summary table
            summary_data = []
            for i in range(1, num_spindles + 1):
                ew_travel = st.session_state.get(f"pos{i}_ew_travel", 0.0)
                ew_approach = st.session_state.get(f"pos{i}_ew_approach", 0.0)
                ew_feed = st.session_state.get(f"pos{i}_ew_feed", 0.0)
                sw_travel = st.session_state.get(f"pos{i}_sw_travel", 0.0)
                sw_approach = st.session_state.get(f"pos{i}_sw_approach", 0.0)
                sw_feed = st.session_state.get(f"pos{i}_sw_feed", 0.0)
                
                ew_total = ew_travel + ew_approach
                sw_total = sw_travel + sw_approach
                ew_effective_revs = ew_total / ew_feed if ew_feed > 0 else 0
                sw_effective_revs = sw_total / sw_feed if sw_feed > 0 else 0
                position_max = max(ew_effective_revs, sw_effective_revs)
                
                if position_max > 0:
                    summary_data.append({
                        "Position": f"Position {i}",
                        "End-Working": f"{ew_effective_revs:.2f}" if ew_effective_revs > 0 else "—",
                        "Side-Working": f"{sw_effective_revs:.2f}" if sw_effective_revs > 0 else "—",
                        "Maximum": f"{position_max:.2f}",
                        "Is Critical": "🎯 YES" if position_max == max_effective_revs and position_max > 0 else "No"
                    })
            
            if summary_data:
                import pandas as pd
                df = pd.DataFrame(summary_data)
                st.dataframe(df, use_container_width=True)
                
                # Show the critical position info more subtly
                critical_positions = [row for row in summary_data if "🎯 YES" in row["Is Critical"]]
                if critical_positions:
                    st.info(f"🎯 **Cycle Time Critical Position:** {critical_positions[0]['Position']} - {critical_positions[0]['Maximum']} Effective Revolutions")

        if all_effective_revs and setup_data["rpm"]:
            max_revs = max(all_effective_revs)
            cycle_time = max_revs * 60 / setup_data["rpm"]
            st.session_state["cycle_time_from_tab2"] = cycle_time
            st.success(f"Overall Cycle Time: {cycle_time:.2f} seconds (Max Effective Revs: {max_revs:.2f}, RPM: {setup_data['rpm']})")
            
            # Compare all cycle rates for best production optimization
            st.markdown("---")
            st.markdown("### 🎯 **Cycle Rate Optimization Analysis**")
            
            cycle_rate_results = []
            all_cycle_rates = [75, 60, 45]
            
            for test_cpm in all_cycle_rates:
                test_result = find_manual_feed_gears(max_revs, setup_data["rpm"], test_cpm)
                if test_result:
                    cycle_rate_results.append({
                        'cpm': test_cpm,
                        'production_per_hour': test_result['production_per_hour'],
                        'cycle_time': test_result['manual_cycle_time'],
                        'revs_match': test_result['revs_percentage_diff'],
                        'gear_desc': test_result.get('driver', 'H.S. CLUTCH'),
                        'is_compound': test_result.get('is_compound', False),
                        'full_result': test_result
                    })
            
            if cycle_rate_results:
                # Find best production rate
                best_production = max(cycle_rate_results, key=lambda x: x['production_per_hour'])
                best_match = min(cycle_rate_results, key=lambda x: x['revs_match'])
                
                st.markdown("#### 📊 **Cycle Rate Comparison**")
                
                comparison_data = []
                for result in cycle_rate_results:
                    # Determine status indicators
                    production_indicator = "🏆 BEST" if result == best_production else ""
                    match_indicator = "🎯 BEST FIT" if result == best_match else ""
                    
                    # Combine indicators
                    status_indicators = []
                    if production_indicator:
                        status_indicators.append(production_indicator)
                    if match_indicator:
                        status_indicators.append(match_indicator)
                    
                    status = " & ".join(status_indicators) if status_indicators else ""
                    
                    comparison_data.append({
                        "Cycle Rate": f"{result['cpm']} CPM",
                        "Production/Hour": f"{result['production_per_hour']:,}",
                        "Cycle Time": f"{result['cycle_time']:.1f} sec",
                        "Chart Match": f"{result['revs_match']:.1f}%",
                        "Feed Gears": result['gear_desc'],
                        "Status": status
                    })
                
                # Display comparison table
                import pandas as pd
                df_comparison = pd.DataFrame(comparison_data)
                st.dataframe(df_comparison, use_container_width=True)
                
                # Recommendation logic
                if best_production == best_match:
                    st.success(f"🎯 **OPTIMAL CHOICE: {best_production['cpm']} CPM** - Best production rate AND best chart match!")
                    recommended_cpm = best_production['cpm']
                else:
                    st.info("🤔 **Trade-off Decision Required:**")
                    col1, col2, col3 = st.columns([0.4, 0.4, 0.2])
                    
                    with col1:
                        st.markdown(f"**🏆 Highest Production: {best_production['cpm']} CPM**")
                        st.metric("Production Rate", f"{best_production['production_per_hour']:,} pcs/hr")
                        st.metric("Chart Match", f"{best_production['revs_match']:.1f}%")
                        st.caption(f"Cycle Time: {best_production['cycle_time']:.1f} sec")
                    
                    with col2:
                        st.markdown(f"**🎯 Best Chart Fit: {best_match['cpm']} CPM**")
                        st.metric("Production Rate", f"{best_match['production_per_hour']:,} pcs/hr")
                        st.metric("Chart Match", f"{best_match['revs_match']:.1f}%")
                        st.caption(f"Cycle Time: {best_match['cycle_time']:.1f} sec")
                    
                    with col3:
                        st.markdown("&nbsp;")  # Empty spacing column
                    
                    # Let user choose
                    recommended_cpm = best_production['cpm']  # Default to best production
                
                # User selection option
                st.markdown("#### 🔧 **Select Your Preferred Cycle Rate:**")
                
                cycle_choice_options = []
                for result in cycle_rate_results:
                    choice_label = f"{result['cpm']} CPM - {result['production_per_hour']:,} pcs/hr ({result['cycle_time']:.1f}s cycle)"
                    if result['cpm'] == recommended_cpm:
                        choice_label += " ⭐ RECOMMENDED"
                    cycle_choice_options.append(choice_label)
                
                selected_choice = st.selectbox(
                    "Choose cycle rate configuration:",
                    options=cycle_choice_options,
                    index=next(i for i, result in enumerate(cycle_rate_results) if result['cpm'] == recommended_cpm)
                )
                
                # Extract selected CPM from choice
                selected_cpm = int(selected_choice.split(" CPM")[0])
                selected_result = next(result for result in cycle_rate_results if result['cpm'] == selected_cpm)
                
                # Update the current selection if different
                if selected_cpm != cpm:
                    st.info(f"💡 **Consider switching from {cpm} CPM to {selected_cpm} CPM** for better performance!")
                    
                    if st.button(f"🔄 Switch to {selected_cpm} CPM Configuration", key="switch_cycle_rate"):
                        st.rerun()
                
                # Show detailed info for selected rate
                st.markdown(f"#### 📋 **Selected Configuration: {selected_cpm} CPM**")
                manual_result = selected_result['full_result']
            else:
                st.warning("⚠️ No suitable configurations found in any cycle rate charts")
                manual_result = None
            
            # Manual Chart Lookup (Following Davenport Manual Methodology) - Current Selection
            if not manual_result:
                manual_result = find_manual_feed_gears(max_revs, setup_data["rpm"], cpm)
            
            if manual_result:
                st.info(f"📖 **Davenport Chart Recommendations - {cpm} CPM:**")
                
                # Store book cycle time for Threading Calculator
                st.session_state["book_cycle_time_recommendation"] = manual_result['manual_cycle_time']
                
                col1, col2 = st.columns(2)
                with col1:
                    # Show book cycle time vs calculated
                    cycle_diff = manual_result['manual_cycle_time'] - cycle_time
                    revs_diff = manual_result['manual_effective_revs'] - max_revs
                    
                    st.metric(
                        "Book Cycle Time", 
                        f"{manual_result['manual_cycle_time']:.1f} sec",
                        f"{cycle_diff:+.1f} vs calculated" if abs(cycle_diff) > 0.05 else "matches calculated"
                    )
                    st.metric(
                        "Book Effective Revs", 
                        f"{manual_result['manual_effective_revs']:.1f}",
                        f"{revs_diff:+.1f} vs needed" if abs(revs_diff) > 0.5 else "matches needed"
                    )
                
                with col2:
                    # Feed gear display
                    gear_parts = []
                    if manual_result['driver']:
                        if str(manual_result['driver']).upper() == "H.S. CLUTCH":
                            gear_parts.append("High Speed Clutch")
                        else:
                            gear_parts.append(f"Driver: {manual_result['driver']}")
                    if manual_result.get('driven_compound'):
                        gear_parts.append(f"Driven Compound: {manual_result['driven_compound']}")
                    if manual_result.get('driver_compound'):
                        gear_parts.append(f"Driver Compound: {manual_result['driver_compound']}")
                    if manual_result.get('driven'):
                        gear_parts.append(f"Driven: {manual_result['driven']}")
                    
                    gear_desc = ", ".join(gear_parts) if gear_parts else "High Speed Clutch"
                    st.info(f"**Feed Gears:** {gear_desc}")
                    
                    st.caption(f"Production: {manual_result['production_per_hour']:,} pieces/hour")
                    st.caption(f"Manual RPM: {manual_result['rpm_used']:,} (closest to {setup_data['rpm']:,})")
                    if manual_result.get('is_compound'):
                        st.caption("🔧 Compound gear setup")
                
                # Accuracy indicator based on percentage difference
                if manual_result['revs_percentage_diff'] <= 5:
                    st.success(f"✅ Excellent match: {manual_result['revs_percentage_diff']:.1f}% difference")
                elif manual_result['revs_percentage_diff'] <= 15:
                    st.info(f"ℹ️ Good match: {manual_result['revs_percentage_diff']:.1f}% difference")
                elif manual_result['revs_percentage_diff'] <= 30:
                    st.warning(f"⚠️ Fair match: {manual_result['revs_percentage_diff']:.1f}% difference")
                else:
                    st.error(f"❌ Poor match: {manual_result['revs_percentage_diff']:.1f}% difference")
                    
                    # Provide specific guidance for large mismatches
                    if manual_result['manual_effective_revs'] < max_revs:
                        shortfall = max_revs - manual_result['manual_effective_revs']
                        st.caption(f"⚠️ **Chart capacity exceeded by {shortfall:.1f} effective revolutions**")
                        st.caption("**Consider:**")
                        st.caption("• Split operations across multiple spindles")
                        st.caption("• Reduce depth of cut or travel distance")
                        st.caption("• Use lower RPM for longer cycle time")
                        st.caption("• Check alternative cycle rates above")
                    else:
                        st.caption("💡 Manual chart provides more capacity than needed")
            else:
                # Handle case where no suitable gear ratio found or manual lookup fails
                st.info("📖 **Davenport Manual Chart:**")
                st.error("❌ No gear ratio provides sufficient effective revolutions")
                st.caption(f"Need: {max_revs:.1f} effective revs at {setup_data['rpm']:,} RPM")
                st.caption("Consider:")
                st.caption("• Reducing operation depth/travel")
                st.caption("• Splitting operations across multiple spindles") 
                st.caption("• Using lower RPM for longer cycle time")
        else:
            st.warning("No effective revs entered yet or RPM missing from Job Setup.")
        if st.button("Generate Setup Sheet"):
            setup_data["cycle_time"] = cycle_time if all_effective_revs and setup_data["rpm"] else setup_data.get("cycle_time", 0)
            output = generate_setup_sheet(setup_data, spindle_data)
            st.download_button(
                label="Download Setup Sheet",
                data=output,
                file_name=f"{setup_data['job_name']}_layout.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        # Store spindle_data for other tabs
        st.session_state["spindle_data"] = spindle_data
        st.session_state["last_spindle_data"] = spindle_data
        
        # Store all CAM operations input data for persistence across tabs
        cam_ops_data = {}
        for i in range(1, num_spindles + 1):
            # Store end-working data
            cam_ops_data[f"pos{i}_ew_travel"] = st.session_state.get(f"pos{i}_ew_travel", 0.0)
            cam_ops_data[f"pos{i}_ew_approach"] = st.session_state.get(f"pos{i}_ew_approach", 0.0)
            cam_ops_data[f"pos{i}_ew_feed"] = st.session_state.get(f"pos{i}_ew_feed", 0.0)
            cam_ops_data[f"pos{i}_ew_tool_desc"] = st.session_state.get(f"pos{i}_ew_tool_desc", "")
            # Store side-working data
            cam_ops_data[f"pos{i}_sw_travel"] = st.session_state.get(f"pos{i}_sw_travel", 0.0)
            cam_ops_data[f"pos{i}_sw_approach"] = st.session_state.get(f"pos{i}_sw_approach", 0.0)
            cam_ops_data[f"pos{i}_sw_feed"] = st.session_state.get(f"pos{i}_sw_feed", 0.0)
            cam_ops_data[f"pos{i}_sw_tool_desc"] = st.session_state.get(f"pos{i}_sw_tool_desc", "")
        st.session_state["cam_operations_data"] = cam_ops_data
        st.session_state["last_spindle_data"] = spindle_data

    with tab3:
        # Check for threading tool context
        thread_tool_context = st.session_state.get("thread_tool_context", None)
        if thread_tool_context:
            st.info(f"Threading tool context: {thread_tool_context}")
            col1, col2 = st.columns(2)
            with col1:
                if st.button("✅ Clear Tool Context", key="clear_thread_context"):
                    del st.session_state["thread_tool_context"]
                    st.rerun()
            with col2:
                if st.button("🔙 Return to CAM Operations", key="return_to_cam_ops"):
                    st.info("👆 **Click the 'CAM Operations' tab above to return to setup**")
        
        # Get setup_data for thread calculator - use primary key first
        setup_data = st.session_state.get("setup_data", {})
        if not setup_data.get("machine_type"):
            setup_data = st.session_state.get("last_setup_data", {
                "machine_type": "Davenport Model B",
                "rpm": 600
            })
        
        machine_type = setup_data.get("machine_type", "Davenport Model B")
        setup_rpm = setup_data.get("rpm", 600)
        thread_calculator_section(setup_rpm)

    with tab4:
        # Get setup_data for simulation - use primary key first
        setup_data = st.session_state.get("setup_data", {})
        if not setup_data.get("machine_config"):
            setup_data = st.session_state.get("last_setup_data", {
                "machine_config": DAVENPORT_CONFIG,
                "machine_type": "Davenport Model B"
            })
        machine_config = setup_data.get("machine_config", DAVENPORT_CONFIG)
        
        # Get spindle_data for simulation - use primary key first  
        spindle_data = st.session_state.get("last_spindle_data", {})
        
        simulation_section(setup_data, spindle_data, machine_config)

        with tab5:
                reference_charts_section()
        
        # Placeholder sections for future reference charts
        st.markdown("### Planned Reference Charts:")
        
        with st.expander("🔩 Drill Size Charts", expanded=False):
            st.info("Drill diameter charts for tap sizes, wire gauge, letter sizes, etc.")
            
        with st.expander("🔧 Threading Charts", expanded=False):
            st.info("Thread pitch charts, tap drill sizes, threading specifications")
            
        with st.expander("📐 Machining Reference", expanded=False):
            st.info("SFM charts, feed rates, cutting speeds for different materials")
            
        with st.expander("⚙️ Machine Specifications", expanded=False):
            st.info("Davenport machine capacities, tooling specifications, collet sizes")
            
        with st.expander("🔢 Conversion Tables", expanded=False):
            st.info("Unit conversions, decimal equivalents, metric/imperial conversions")

   
if __name__ == "__main__":
    main()