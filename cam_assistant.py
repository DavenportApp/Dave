import streamlit as st
import json
import math
import pandas as pd
from io import BytesIO
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

# ------------------------------
# Load External Data
# ------------------------------
def load_data():
    with open("materials.json") as f:
        material_data = json.load(f)
    with open("sfm_guidelines.json") as f:
        sfm_guidelines = json.load(f)
    with open("cams_data.json") as f:
        cam_data = json.load(f)
    # Add missing cam from Excel
    cam_data["5-C-792"] = {"size": "3/16", "rise": 0.1650, "type": "TURNING", "cut_start": 0, "cut_end": 45, "dwell_end": 50, "total_spaces": 100}
    return material_data, sfm_guidelines, cam_data

def load_gear_table(cpm="75"):
    with open("gears.json") as f:
        data = json.load(f)
    if cpm in ["60", "45"]:
        return data["75"]  # "same_as_75" in JSON
    return data[str(cpm)]

# ------------------------------
# CAM Recommendation Logic
# ------------------------------
def recommend_cam(target_rise, total_travel, tool_type, cam_db, material, min_block=0.8, max_block=1.2):
    tool_to_type = {
        "BOXTOOL": "TURNING", "BROACH": "TURNING", "CENTER": "TURNING", "COUNTERBORE": "TURNING",
        "CUSTOM": "TURNING", "DEIHEAD": "TURNING", "DRILL": "TURNING", "HOLLOW MILL": "TURNING",
        "REAMER": "TURNING", "TAP": "TURNING", "TRIPAN": "TURNING",
        "CROSS DRILL": "FORM", "CROSS TAP": "FORM", "CUTOFF": "FORM", "FORM TOOL": "FORM",
        "KNURL": "FORM", "ROLL STAMP": "FORM", "SHAVE": "FORM", "SKIVE": "FORM", "THREAD ROLL": "FORM",
        "BRASS THREADING": "BRASS THREADING", "STEEL THREADING": "STEEL THREADING"
    }
    material_to_thread = {"360 Brass": "BRASS THREADING", "C260 Brass": "BRASS THREADING", "C464 Naval Brass": "BRASS THREADING"}
    
    cam_type_target = tool_to_type.get(tool_type.strip().upper())
    if tool_type in ["TAP", "THREAD ROLL"] and material in material_to_thread:
        cam_type_target = material_to_thread[material]
    
    if not cam_type_target:
        return None

    closest = None
    min_diff = float("inf")

    for cam_name, details in cam_db.items():
        cam_type = details.get("type", "").strip().upper()
        cam_rise = details.get("rise", 0)

        if cam_type != cam_type_target or cam_rise <= 0:
            continue

        block_setting = total_travel / cam_rise
        if not (min_block <= block_setting <= max_block):
            continue

        diff = abs(cam_rise - target_rise)
        if diff < min_diff:
            min_diff = diff
            closest = (cam_name, details)

    return closest

# ------------------------------
# Job Setup Section
# ------------------------------
def job_setup_section(material_data, sfm_guidelines):
    st.header("📋 Job Setup – Davenport Model B")
    col1, col2 = st.columns(2)
    with col1:
        job_name = st.text_input("Part Number", value="44153-36-99", key="setup_job_name")
        material = st.selectbox("Material", list(material_data.keys()), index=0, key="setup_material")
        bar_shape = st.selectbox("Bar Shape", ["Round", "Hex", "Special", "Tube", "Square"], key="setup_shape")
        dia = st.number_input("Bar Diameter (in)", min_value=0.100, max_value=2.000, value=0.3125, step=0.001, format="%.3f", key="setup_dia")
        part_length = st.number_input("Part Length (in)", min_value=0.001, step=0.001, format="%.3f", key="setup_part_len")
        cutoff = st.number_input("Cutoff Width (in)", min_value=0.010, step=0.001, value=0.069, format="%.3f", key="setup_cutoff")
        faceoff = st.number_input("Faceoff Amount (in)", min_value=0.000, step=0.001, format="%.3f", key="setup_faceoff")
        collets = st.text_input("Collets", value="5/16 RD", key="setup_collets")
        feed_finger = st.text_input("Feed Finger", value="5/16 RD", key="setup_feed_finger")
        set_pads = st.text_input("Set Pads", value="", key="setup_set_pads")
        burr_collect = st.text_input("Burr Collect", value="9/32 or 19/64 RD", key="setup_burr_collect")
    with col2:
        material_key = next((k for k, v in sfm_guidelines.items() if v["material_name"].lower() == material.lower()), "Brass")
        sfm_default = material_data[material].get("sfm", 150)
        sfm = st.number_input("Surface Feet per Minute (SFM)", value=sfm_default, step=5, key="setup_sfm")
        rpm = st.number_input("Machine RPM", min_value=100, max_value=6000, value=3300, step=100, key="setup_rpm")
        bar_len = st.number_input("Bar Length (in)", value=144.0, step=1.0, format="%.3f", key="setup_bar_len")
        remnant = st.number_input("Remnant Length (in)", min_value=0.0, max_value=bar_len, value=6.000, step=0.500, format="%.3f", key="setup_remnant")
        spindle_gears = st.text_input("Spindle Gears", value="44-20", key="setup_spindle_gears")
        feed_gears = st.text_input("Feed Gears", value="50-30-60", key="setup_feed_gears")
        thread_gears = st.text_input("Threading Gears", value="", key="setup_thread_gears")
        cycle_time = st.number_input("Cycle Time (sec)", min_value=0.0, value=1.6, step=0.1, key="setup_cycle_time")
        machine_code = st.text_input("Machine Code", value="A,B,O", key="setup_machine_code")
    usable_bar_len = bar_len - remnant
    per_part_len = part_length + cutoff + faceoff
    parts_per_bar = usable_bar_len / per_part_len if per_part_len > 0 else 0
    bar_weight = usable_bar_len * material_data[material].get("density", 0.307)
    st.markdown(f"📏 Usable Bar Length: **{usable_bar_len:.3f} in**")
    st.markdown(f"🧮 Estimated Parts per Bar: **{parts_per_bar:.2f}**")
    st.markdown(f"⚖️ Usable Bar Weight: **{bar_weight:.2f} lbs**")
    return {
        "job_name": job_name, "material": material, "bar_shape": bar_shape, "dia": dia, "part_length": part_length,
        "cutoff": cutoff, "faceoff": faceoff, "collets": collets, "feed_finger": feed_finger, "set_pads": set_pads,
        "burr_collect": burr_collect, "sfm": sfm, "rpm": rpm, "bar_len": bar_len, "remnant": remnant,
        "spindle_gears": spindle_gears, "feed_gears": feed_gears, "thread_gears": thread_gears, "cycle_time": cycle_time,
        "machine_code": machine_code, "parts_per_bar": parts_per_bar, "bar_weight": bar_weight, "material_key": material_key
    }

# ------------------------------
# Quote Breakdown Section
# ------------------------------
def quote_breakdown_section(parts_per_bar, bar_weight):
    st.subheader("📦 Quoting Quantities")
    q1, q2, q3 = st.columns(3)
    with q1: low = st.number_input("Low Quote Qty", value=100, step=100, key="quote_low_qty")
    with q2: mid = st.number_input("Mid Quote Qty", value=500, step=100, key="quote_mid_qty")
    with q3: high = st.number_input("High Quote Qty", value=1000, step=100, key="quote_high_qty")
    p1, p2, p3 = st.columns(3)
    with p1: price_low = st.number_input("💲 Price/lb – Low Qty", value=5.00, step=0.10, format="%.2f", key="price_low")
    with p2: price_mid = st.number_input("💲 Price/lb – Mid Qty", value=4.75, step=0.10, format="%.2f", key="price_mid")
    with p3: price_high = st.number_input("💲 Price/lb – High Qty", value=4.50, step=0.10, format="%.2f", key="price_high")
    def calc(qty, price_per_lb):
        bars = int((qty / parts_per_bar) + 0.999) if parts_per_bar else 0
        weight = round(qty * (bar_weight / parts_per_bar), 2) if parts_per_bar else 0
        cost = round(weight * price_per_lb, 2) if price_per_lb else 0.0
        return bars, weight, cost
    c1, c2, c3 = st.columns(3)
    for col, (label, qty, price, icon, color) in zip(
        [c1, c2, c3],
        [("Low Quote", low, price_low, "🔹", "#cce5ff"),
         ("Mid Quote", mid, price_mid, "🔸", "#fff3cd"),
         ("High Quote", high, price_high, "🔺", "#d4edda")]
    ):
        bars, weight, cost = calc(qty, price)
        with col:
            st.markdown(f"""
                <div style='background-color:{color}; padding:12px; border-radius:10px'>
                <strong>{icon} {label}</strong><br>
                📦 Quantity: {qty:,}<br>
                🧱 Bars: {bars}<br>
                ⚖️ Weight: {weight:.2f} lbs<br>
                💵 Est. Mat.Cost: ${cost:.2f}<br>
                🔹 Mat.Cost/Part: ${cost/qty:.4f}
                </div>
            """, unsafe_allow_html=True)

# ------------------------------
# Charts Section
# ------------------------------
def charts_section(gear_table, spindle_data, cam_data):
    st.subheader("📊 Davenport Model B Charts")
    # Gear Ratio Chart
    gear_df = pd.DataFrame([
        {"Gear Combination": f"{g['driver']}-{g['driven']}", "Ratio": g["ratio"]}
        for g in gear_table
    ])
    fig_gear = px.bar(gear_df, x="Gear Combination", y="Ratio", title="Gear Ratios for Spindle Speed (75 CPM)",
                      color_discrete_sequence=["#1f77b4"])
    st.plotly_chart(fig_gear)
    # Cam Timing Chart
    if spindle_data:
        fig_timing = go.Figure()
        for spindle in spindle_data:
            cam_name = spindle["cam"]
            if cam_name in cam_data:
                cam = cam_data[cam_name]
                fig_timing.add_trace(go.Scatter(
                    x=[cam["cut_start"], cam["cut_end"], cam["dwell_end"], cam["total_spaces"]],
                    y=[spindle["position"]] * 4,
                    mode="lines+markers",
                    name=spindle["position"],
                    text=["Start", "Cut", "Dwell", "End"],
                    line=dict(width=4)
                ))
        fig_timing.update_layout(
            title="Cam Timing Diagram (360° Cycle)",
            xaxis_title="Degrees",
            yaxis_title="Spindle Position",
            xaxis_range=[0, 100],
            showlegend=True
        )
        st.plotly_chart(fig_timing)

# ------------------------------
# Setup Sheet Generation
# ------------------------------
def generate_setup_sheet(setup_data, spindle_data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Setup Sheet"
    # Define styles
    header_font = Font(bold=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    center_align = Alignment(horizontal="center")
    # Header
    ws.merge_cells("A1:K1")
    ws["A1"] = "MACHINE: Davenport Model B"
    ws["A1"].font = header_font
    ws["A1"].alignment = center_align
    ws.merge_cells("A2:K2")
    ws["A2"] = "Operator/Lead Person Layout Instructions"
    ws["A2"].font = header_font
    ws["A2"].alignment = center_align
    # Part Info
    headers = ["Part No.", "Internal Rev.", "Tool No.", "Tool Location", "Job No.", "Orig Date", "", "Updated", "Updated by", "Approved", ""]
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col).value = header
        ws.cell(row=4, column=col).font = header_font
        ws.cell(row=4, column=col).border = border
        ws.cell(row=4, column=col).alignment = center_align
    ws["A5"] = setup_data["job_name"]
    ws["B5"] = "NA"
    ws["I5"] = "TJ"
    # Material Info
    headers = ["Material", "", "Size", "Shape", "Feet Per M", "Bars Per M", "Lbs. Per M", "", "Collets", "Feed Finger", "Set Pads", "Burr. Collect"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=6, column=col).value = header
        ws.cell(row=6, column=col).font = header_font
        ws.cell(row=6, column=col).border = border
        ws.cell(row=6, column=col).alignment = center_align
    ws["A7"] = setup_data["material"]
    ws["C7"] = setup_data["dia"]
    ws["D7"] = setup_data["bar_shape"]
    ws["I7"] = setup_data["collets"]
    ws["J7"] = setup_data["feed_finger"]
    ws["K7"] = setup_data["set_pads"]
    ws["L7"] = setup_data["burr_collect"]
    # Spindle Info
    headers = ["Spindle Speed", "", "S.F.M.", "Drill", "Tap", "Spindle Gears", "Machine Code", "", "Sec", "Feed Gears", "Thread Speed", "Threading Gears", "Eff. Rev"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=8, column=col).value = header
        ws.cell(row=8, column=col).font = header_font
        ws.cell(row=8, column=col).border = border
        ws.cell(row=8, column=col).alignment = center_align
    ws["A9"] = setup_data["rpm"]
    ws["C9"] = setup_data["sfm"]
    ws["F9"] = setup_data["spindle_gears"]
    ws["G9"] = setup_data["machine_code"]
    ws["I9"] = setup_data["cycle_time"]
    ws["J9"] = setup_data["feed_gears"]
    ws["L9"] = setup_data["thread_gears"]
    ws["M9"] = 59  # Static from Excel
    # Operations
    headers = ["Position", "Operation", "CAM", "CAM Spaces", "Feed", "Feed Per Rev.", "Location", "", "Tool Slide", "", "Cross Slide"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=10, column=col).value = header
        ws.cell(row=10, column=col).font = header_font
        ws.cell(row=10, column=col).border = border
        ws.cell(row=10, column=col).alignment = center_align
    for i, spindle in enumerate(spindle_data, start=12):
        ws[f"A{i}"] = spindle["position"]
        ws[f"B{i}"] = spindle["operation"]
        ws[f"C{i}"] = spindle["cam"]
        ws[f"D{i}"] = spindle["cam_spaces"]
        ws[f"E{i}"] = spindle["feed"]
        ws[f"F{i}"] = spindle["feed_per_rev"]
        ws[f"G{i}"] = spindle["location"]
        ws[f"I{i}"] = spindle["tool_slide"]
        ws[f"K{i}"] = spindle["cross_slide"]
        for col in range(1, 12):
            ws.cell(row=i, column=col).border = border
            ws.cell(row=i, column=col).alignment = center_align
    # Footer
    ws["A31"] = "FORM # 409-6 10-04-11-B"
    ws.merge_cells("D31:K31")
    ws["D31"] = "CONFIDENTIAL DOCUMENT: Distribution outside of KKSP employees is strictly prohibited"
    ws["D31"].font = Font(italic=True)
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ------------------------------
# Main App
# ------------------------------
def main():
    st.set_page_config(page_title="Davenport CAM Assistant", layout="wide")
    material_data, sfm_guidelines, cam_data = load_data()
    st.title("🛠️ Davenport Model B CAM Setup Assistant")
    tab1, tab2, tab3 = st.tabs(["💵 Quote Builder", "🧰 CAM Operations", "📊 Charts"])
    # Tab 1: Quote Builder
    with tab1:
        setup_data = job_setup_section(material_data, sfm_guidelines)
        quote_breakdown_section(setup_data["parts_per_bar"], setup_data["bar_weight"])
    # Tab 2: CAM Tooling Setup
    with tab2:
        st.subheader("🌀 CAM Operations – Spindle Setup")
        cpm = st.selectbox("Select Machine Cycle Rate (CPM)", options=[75, 60, 45], index=0)
        gear_table = load_gear_table(cpm)
        num_spindles = 5  # Fixed for Davenport Model B
        spindle_data = []
        for i in range(1, num_spindles + 1):
            with st.expander(f"🔧 Position {i}", expanded=True):
                # End-Working Tool
                with st.expander("🔹 End-Working Tool", expanded=False):
                    ew_col1, ew_col2 = st.columns(2)
                    with ew_col1:
                        ew_tool = st.selectbox("Tool Type (End)", [
                            "BOXTOOL", "BROACH", "CENTER", "COUNTERBORE", "CUSTOM", "DEIHEAD", "DRILL",
                            "HOLLOW MILL", "REAMER", "TAP", "TRIPAN"
                        ], key=f"pos{i}_ew_tool")
                        ew_travel = st.number_input("Tool Travel (End)", min_value=0.0, step=0.001, format="%.3f", key=f"pos{i}_ew_travel")
                        ew_approach = st.number_input("Approach (End)", min_value=0.0, step=0.001, format="%.3f", key=f"pos{i}_ew_approach")
                        ew_feed = st.number_input("Feed (End)", min_value=0.0, step=0.001, format="%.3f", key=f"pos{i}_ew_feed")
                        ew_tool_desc = st.text_input("Tool Description (End)", key=f"pos{i}_ew_tool_desc")
                    with ew_col2:
                        ew_total = ew_travel + ew_approach
                        st.caption(f"🧮 Total Travel: {ew_total:.3f}")
                        if ew_total == 0.0:
                            st.warning("⚠️ Please enter travel and approach for End-Working.")
                        else:
                            ew_cam = recommend_cam(ew_total, ew_total, ew_tool, cam_data, setup_data["material"])
                            if ew_cam:
                                cam_name, cam_info = ew_cam
                                cam_rise = cam_info.get("rise", 0)
                                ew_block = ew_total / cam_rise if cam_rise else 0
                                feed_per_rev = ew_feed / setup_data["rpm"] if setup_data["rpm"] else 0
                                st.success(f"💡 End Cam: Size {cam_info.get('size', '?')} – Cam #: **{cam_name}**")
                                st.caption(f"Rise: {cam_rise:.4f}, Type: {cam_info.get('type', '?')}")
                                st.markdown(f"<div style='font-size:20px; font-weight:700; padding-top:4px;'>🔧 End Block Setting: {ew_block:.2f}</div>", unsafe_allow_html=True)
                                spindle_data.append({
                                    "position": f"End{i}",
                                    "operation": ew_tool,
                                    "cam": cam_name,
                                    "cam_spaces": cam_info.get("size", ""),
                                    "feed": ew_feed,
                                    "feed_per_rev": feed_per_rev,
                                    "location": i,
                                    "tool_slide": ew_tool_desc,
                                    "cross_slide": ""
                                })
                            else:
                                st.warning("⚠️ No viable cam found for End-Working in block range 0.8–1.2")
                # Side-Working Tool
                with st.expander("🔸 Side-Working Tool", expanded=False):
                    sw_col1, sw_col2 = st.columns(2)
                    with sw_col1:
                        sw_tool = st.selectbox("Tool Type (Side)", [
                            "CROSS DRILL", "CROSS TAP", "CUTOFF", "CUSTOM", "FORM TOOL", "KNURL",
                            "ROLL STAMP", "SHAVE", "SKIVE", "THREAD ROLL"
                        ], key=f"pos{i}_sw_tool")
                        sw_travel = st.number_input("Tool Travel (Side)", min_value=0.0, step=0.001, format="%.3f", key=f"pos{i}_sw_travel")
                        sw_approach = st.number_input("Approach (Side)", min_value=0.0, step=0.001, format="%.3f", key=f"pos{i}_sw_approach")
                        sw_feed = st.number_input("Feed (Side)", min_value=0.0, step=0.001, format="%.3f", key=f"pos{i}_sw_feed")
                        sw_tool_desc = st.text_input("Tool Description (Side)", key=f"pos{i}_sw_tool_desc")
                    with sw_col2:
                        sw_total = sw_travel + sw_approach
                        st.caption(f"🧮 Total Travel: {sw_total:.3f}")
                        if sw_total == 0.0:
                            st.warning("⚠️ Please enter travel and approach for Side-Working.")
                        else:
                            sw_cam = recommend_cam(sw_total, sw_total, sw_tool, cam_data, setup_data["material"])
                            if sw_cam:
                                cam_name, cam_info = sw_cam
                                cam_rise = cam_info.get("rise", 0)
                                sw_block = sw_total / cam_rise if cam_rise else 0
                                feed_per_rev = sw_feed / setup_data["rpm"] if setup_data["rpm"] else 0
                                st.success(f"💡 Side Cam: Size {cam_info.get('size', '?')} – Cam #: **{cam_name}**")
                                st.caption(f"Rise: {cam_rise:.4f}, Type: {cam_info.get('type', '?')}")
                                st.markdown(f"<div style='font-size:20px; font-weight:700; padding-top:4px;'>🔧 Side Block Setting: {sw_block:.2f}</div>", unsafe_allow_html=True)
                                spindle_data.append({
                                    "position": f"Side{i}",
                                    "operation": sw_tool,
                                    "cam": cam_name,
                                    "cam_spaces": cam_info.get("size", ""),
                                    "feed": sw_feed,
                                    "feed_per_rev": feed_per_rev,
                                    "location": i,
                                    "tool_slide": "",
                                    "cross_slide": sw_tool_desc
                                })
                            else:
                                st.warning("⚠️ No viable cam found for Side-Working in block range 0.8–1.2")
        # Generate Setup Sheet
        if st.button("Generate Setup Sheet"):
            output = generate_setup_sheet(setup_data, spindle_data)
            st.download_button(
                label="Download Setup Sheet",
                data=output,
                file_name=f"{setup_data['job_name']}_layout.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    # Tab 3: Charts
    with tab3:
        charts_section(gear_table, spindle_data, cam_data)

if __name__ == "__main__":
    main()